from fastapi import APIRouter, Depends, Response, UploadFile, File
from crew_legacy.admin_logic.employee import (
    create_employee_logic,
    get_all_employees_logic,
)
from fastapi import HTTPException
from crew_legacy.admin_logic.dropdown import (
    create_dropdown_logic,
    get_dropdown_by_type_logic
)

from bson import ObjectId
from crew_legacy.database.database_mongo import employee_collection, organization_unit_collection
from crew_legacy.admin_logic.dropdown import dropdown_collection
from crew_legacy.database.database_mongo import DutyLeave_collection, system_settings_collection

from openpyxl import Workbook, load_workbook
from io import BytesIO
from crew_legacy.security_utils import ensure_upload_allowed

from crew_legacy.admin_logic.DutyLeaveType import (
    create_dutyLeave_logic,
    get_dutyLeave_by_type_logic
)
from crew_legacy.admin_logic.auth_utils import require_admin
from crew_legacy.admin_logic.auth_utils import hash_password, validate_password_policy
from datetime import datetime, timedelta




router = APIRouter(tags=["Admin"], dependencies=[Depends(require_admin)])


def normalize_list(value):
    if isinstance(value, list):
        items = value
    elif value in [None, ""]:
        return []
    else:
        items = str(value).split(",")
    return list(dict.fromkeys(
        str(item).strip() for item in items if str(item).strip()
    ))


def resolve_employee_organization(function_ids=None, employee_id=None):
    """Resolve employee hierarchy only from the Organization Master."""
    employee_id = str(employee_id or "").strip()
    units = list(organization_unit_collection.find({"isActive": {"$ne": False}}))
    unit_by_id = {str(unit["_id"]): unit for unit in units}

    manual_function_ids = [
        unit_id for unit_id in normalize_list(function_ids)
        if unit_id in unit_by_id and unit_by_id[unit_id].get("unitType") == "function"
    ]
    role_unit_ids = []
    if employee_id:
        for unit in units:
            role_ids = normalize_list(unit.get("headEmployeeIds")) + normalize_list(unit.get("juniorEmployeeIds"))
            if employee_id in role_ids:
                role_unit_ids.append(str(unit["_id"]))
    role_function_ids = [
        unit_id for unit_id in role_unit_ids
        if unit_by_id[unit_id].get("unitType") == "function"
    ]
    manual_function_ids = [
        unit_id for unit_id in manual_function_ids
        if unit_id not in role_function_ids
    ]

    selected_function_ids = list(dict.fromkeys(manual_function_ids + role_function_ids))
    seed_ids = list(dict.fromkeys(selected_function_ids + role_unit_ids))
    vertical_ids = []
    section_ids = []
    department_ids = []
    reporting_ids = []
    intermediary_candidates = []
    hod_candidates = []

    for seed_id in seed_ids:
        current = unit_by_id.get(seed_id)
        if not current:
            continue

        configured_current_heads = normalize_list(current.get("headEmployeeIds"))
        employee_is_current_head = employee_id in configured_current_heads
        current_heads = [
            value for value in configured_current_heads
            if value and value != employee_id
        ]
        direct_found = False
        if current_heads and not employee_is_current_head:
            reporting_ids.extend(current_heads)
            direct_found = True

        parent_id = str(current.get("parentId")) if current.get("parentId") else ""
        while parent_id and parent_id in unit_by_id:
            parent = unit_by_id[parent_id]
            parent_type = parent.get("unitType")
            if parent_type == "vertical":
                vertical_ids.append(parent_id)
            elif parent_type == "section":
                section_ids.append(parent_id)
            elif parent_type == "department":
                department_ids.append(parent_id)

            parent_heads = [
                value for value in normalize_list(parent.get("headEmployeeIds"))
                if value and value != employee_id
            ]
            if parent_heads:
                if not direct_found:
                    reporting_ids.extend(parent_heads)
                    direct_found = True
                elif parent_type != "department":
                    intermediary_candidates.extend(parent_heads)
                if parent_type == "department":
                    hod_candidates.extend(parent_heads)

            parent_id = str(parent.get("parentId")) if parent.get("parentId") else ""

        seed_type = current.get("unitType")
        if seed_type == "vertical":
            vertical_ids.append(seed_id)
        elif seed_type == "section":
            section_ids.append(seed_id)
        elif seed_type == "department":
            department_ids.append(seed_id)

    vertical_ids = list(dict.fromkeys(vertical_ids))
    section_ids = list(dict.fromkeys(section_ids))
    department_ids = list(dict.fromkeys(department_ids))
    reporting_ids = list(dict.fromkeys(reporting_ids))
    intermediary_candidates = [
        value for value in dict.fromkeys(intermediary_candidates)
        if value not in reporting_ids
    ]
    hod_candidates = list(dict.fromkeys(hod_candidates))

    return {
        "manualFunctionIds": manual_function_ids,
        "roleFunctionIds": role_function_ids,
        "functionIds": selected_function_ids,
        "verticalIds": vertical_ids,
        "verticals": [unit_by_id[value].get("name") for value in vertical_ids],
        "sectionIds": section_ids,
        "sections": [unit_by_id[value].get("name") for value in section_ids],
        "departmentIds": department_ids,
        "departments": [unit_by_id[value].get("name") for value in department_ids],
        "department": unit_by_id[department_ids[0]].get("name") if department_ids else None,
        "reportingOfficerIds": reporting_ids,
        "reportingOfficerId": reporting_ids[0] if reporting_ids else None,
        "intermediaryReportingId": intermediary_candidates[0] if intermediary_candidates else None,
        "hodId": hod_candidates[0] if hod_candidates else None,
    }


def sync_employee_organization():
    for employee in employee_collection.find({}, {"userId": 1, "functionIds": 1, "manualFunctionIds": 1}):
        resolved = resolve_employee_organization(
            employee.get("manualFunctionIds", employee.get("functionIds")),
            employee.get("userId"),
        )
        employee_collection.update_one(
            {"_id": employee["_id"]},
            {"$set": {
                **resolved,
                "vertical": (resolved.get("verticals") or [None])[0],
                "updatedAt": datetime.utcnow(),
            }},
        )


def serialize(emp):

    # ðŸ”¥ FETCH RELATED EMPLOYEES
    reporting_ids = normalize_list(
        emp.get("reportingOfficerIds") or emp.get("reportingOfficerId")
    )
    verticals = normalize_list(emp.get("verticals") or emp.get("vertical"))
    reporting_people = list(employee_collection.find(
        {"userId": {"$in": reporting_ids}}, {"userId": 1, "name": 1}
    )) if reporting_ids else []
    reporting_names = {
        person.get("userId"): person.get("name") for person in reporting_people
    }
    function_ids = normalize_list(emp.get("functionIds"))
    function_docs = list(organization_unit_collection.find(
        {"_id": {"$in": [ObjectId(value) for value in function_ids if ObjectId.is_valid(value)]}},
        {"name": 1},
    )) if function_ids else []
    function_names = {str(item["_id"]): item.get("name") for item in function_docs}

    hod = employee_collection.find_one({
        "userId": emp.get("hodId")
    })

    intermediary = employee_collection.find_one({
        "userId": emp.get("intermediaryReportingId")
    })

    return {
        "id": str(emp["_id"]),
        "name": emp.get("name"),
        "nameHindi": emp.get("nameHindi"),
        "designation": emp.get("designation"),
        "designationHindi": emp.get("designationHindi"),
        "userId": emp.get("userId"),
        "phone": emp.get("phone"),
        "gmail": emp.get("gmail"),
        "dutyType": emp.get("dutyType"),
        "category": normalize_list(emp.get("category")),

        # ðŸ”¥ NEW FIELDS
        "verticals": verticals,
        "vertical": verticals[0] if verticals else None,
        "verticalIds": normalize_list(emp.get("verticalIds")),
        "sections": normalize_list(emp.get("sections")),
        "sectionIds": normalize_list(emp.get("sectionIds")),
        "departments": normalize_list(emp.get("departments")),
        "departmentIds": normalize_list(emp.get("departmentIds")),
        "department": emp.get("department"),

        # ðŸ”¥ IDs (keep for logic)
        "reportingOfficerIds": reporting_ids,
        "reportingOfficerId": reporting_ids[0] if reporting_ids else None,
        "intermediaryReportingId": emp.get("intermediaryReportingId"),
        "hodId": emp.get("hodId"),

        # ðŸ”¥ NAMES (for frontend display)
        "reportingOfficerNames": [
            reporting_names.get(officer_id, officer_id) for officer_id in reporting_ids
        ],
        "reportingOfficerName": reporting_names.get(reporting_ids[0]) if reporting_ids else None,
        "functionIds": function_ids,
        "functionNames": [function_names.get(value, value) for value in function_ids],
        "intermediaryReportingName": intermediary.get("name") if intermediary else None,
        "hodName": hod.get("name") if hod else None
    }



@router.post("/employees")  #### save employee entry to database
def create_employee(employee: dict):
    employee.update(resolve_employee_organization(
        employee.get("manualFunctionIds", employee.get("functionIds")),
        employee.get("userId"),
    ))
    inserted_id = create_employee_logic(employee)
    return {"message": "Employee added", "id": inserted_id}


@router.get("/employees")  #### fetch employee entry to database
def get_employees():
    employees = get_all_employees_logic()
    return [serialize(emp) for emp in employees]

@router.post("/dropdown")    ###### create new dropdown entry
def create_dropdown(data: dict):
    return create_dropdown_logic(data)


@router.get("/dropdown/{dropdown_type}")    ###### fetch dropdown entry
def get_dropdown(dropdown_type: str):
    data = get_dropdown_by_type_logic(dropdown_type)
    return [
        {
            "id": str(item["_id"]),
            "value": item["value"]
        }
        for item in data
    ]



ORG_PARENT_TYPES = {
    "department": set(),
    "vertical": {"department"},
    "section": {"vertical"},
    "function": {"vertical", "section"},
}


def serialize_org_unit(unit, employee_names=None, parent_names=None):
    employee_names = employee_names or {}
    parent_names = parent_names or {}
    head_ids = normalize_list(unit.get("headEmployeeIds"))
    junior_ids = normalize_list(unit.get("juniorEmployeeIds"))
    parent_id = str(unit.get("parentId")) if unit.get("parentId") else None
    return {
        "id": str(unit["_id"]),
        "name": unit.get("name"),
        "unitType": unit.get("unitType"),
        "parentId": parent_id,
        "parentName": parent_names.get(parent_id),
        "headEmployeeIds": head_ids,
        "headEmployeeNames": [employee_names.get(value, value) for value in head_ids],
        "juniorEmployeeIds": junior_ids,
        "juniorEmployeeNames": [employee_names.get(value, value) for value in junior_ids],
        "isActive": unit.get("isActive", True),
    }


def validate_org_unit(data, current_id=None):
    name = str(data.get("name") or "").strip()
    unit_type = str(data.get("unitType") or "").strip().lower()
    parent_id = data.get("parentId") or None
    if not name or unit_type not in ORG_PARENT_TYPES:
        raise HTTPException(400, "Name and a valid unit type are required")

    current_unit = None
    if current_id and ObjectId.is_valid(current_id):
        current_unit = organization_unit_collection.find_one({"_id": ObjectId(current_id)})

    parent = None
    if parent_id:
        if not ObjectId.is_valid(parent_id):
            raise HTTPException(400, "Invalid parent unit")
        if current_id and parent_id == current_id:
            raise HTTPException(400, "A unit cannot be its own parent")
        parent = organization_unit_collection.find_one({"_id": ObjectId(parent_id)})
        if not parent:
            raise HTTPException(404, "Parent unit not found")

    allowed = ORG_PARENT_TYPES[unit_type]
    if allowed and (not parent or parent.get("unitType") not in allowed):
        if not (
            current_unit
            and unit_type == "section"
            and str(current_unit.get("parentId")) == parent_id
        ):
            raise HTTPException(400, f"{unit_type.title()} must belong to a {' or '.join(sorted(allowed))}")
    if not allowed and parent:
        raise HTTPException(400, "Department is a top-level unit")

    return {
        "name": name,
        "unitType": unit_type,
        "parentId": parent.get("_id") if parent else None,
        "headEmployeeIds": normalize_list(data.get("headEmployeeIds")),
        "juniorEmployeeIds": normalize_list(data.get("juniorEmployeeIds")) if unit_type == "function" else [],
        "isActive": bool(data.get("isActive", True)),
        "updatedAt": datetime.utcnow(),
    }


@router.get("/organization/units")
def get_organization_units():
    units = list(organization_unit_collection.find({}).sort([("unitType", 1), ("name", 1)]))
    employee_names = {
        item.get("userId"): item.get("name")
        for item in employee_collection.find({}, {"userId": 1, "name": 1})
    }
    parent_names = {str(item["_id"]): item.get("name") for item in units}
    return [serialize_org_unit(item, employee_names, parent_names) for item in units]


@router.post("/organization/resolve-employee")
def resolve_organization_employee(data: dict):
    resolved = resolve_employee_organization(
        data.get("manualFunctionIds", data.get("functionIds")),
        data.get("userId"),
    )
    related_ids = list(dict.fromkeys(
        resolved["reportingOfficerIds"]
        + [resolved.get("intermediaryReportingId"), resolved.get("hodId")]
    ))
    related_ids = [value for value in related_ids if value]
    employee_names = {
        item.get("userId"): item.get("name")
        for item in employee_collection.find(
            {"userId": {"$in": related_ids}},
            {"userId": 1, "name": 1},
        )
    }
    resolved.update({
        "reportingOfficerNames": [
            employee_names.get(value, value) for value in resolved["reportingOfficerIds"]
        ],
        "intermediaryReportingName": employee_names.get(resolved.get("intermediaryReportingId")),
        "hodName": employee_names.get(resolved.get("hodId")),
    })
    return resolved


@router.post("/organization/units")
def create_organization_unit(data: dict):
    unit = validate_org_unit(data)
    if organization_unit_collection.find_one({
        "name": unit["name"], "unitType": unit["unitType"], "parentId": unit["parentId"]
    }):
        raise HTTPException(409, "This organization unit already exists")
    unit["createdAt"] = datetime.utcnow()
    result = organization_unit_collection.insert_one(unit)
    sync_employee_organization()
    return {"message": "Organization unit created", "id": str(result.inserted_id)}


@router.put("/organization/units/{unit_id}")
def update_organization_unit(unit_id: str, data: dict):
    if not ObjectId.is_valid(unit_id):
        raise HTTPException(400, "Invalid organization unit")
    if not organization_unit_collection.find_one({"_id": ObjectId(unit_id)}):
        raise HTTPException(404, "Organization unit not found")
    organization_unit_collection.update_one(
        {"_id": ObjectId(unit_id)}, {"$set": validate_org_unit(data, unit_id)}
    )
    sync_employee_organization()
    return {"message": "Organization unit updated"}


@router.delete("/organization/units/{unit_id}")
def delete_organization_unit(unit_id: str):
    if not ObjectId.is_valid(unit_id):
        raise HTTPException(400, "Invalid organization unit")
    object_id = ObjectId(unit_id)
    if organization_unit_collection.find_one({"parentId": object_id}):
        raise HTTPException(409, "Remove or move child units first")
    if employee_collection.find_one({"functionIds": unit_id}):
        raise HTTPException(409, "Remove employee function assignments first")
    result = organization_unit_collection.delete_one({"_id": object_id})
    if not result.deleted_count:
        raise HTTPException(404, "Organization unit not found")
    sync_employee_organization()
    return {"message": "Organization unit deleted"}


@router.get("/organization/tree")
def get_organization_tree():
    units = list(organization_unit_collection.find({"isActive": {"$ne": False}}))
    if not units:
        return []
    child_map = {}
    for unit in units:
        parent_id = str(unit.get("parentId")) if unit.get("parentId") else "ROOT"
        child_map.setdefault(parent_id, []).append(unit)

    employees = list(employee_collection.find({}))
    employee_map = {item.get("userId"): item for item in employees if item.get("userId")}
    members = {}
    for employee in employees:
        for function_id in normalize_list(employee.get("functionIds")):
            members.setdefault(function_id, []).append(employee)

    def person_node(employee):
        return {
            "expanded": True,
            "type": "person",
            "data": {
                "userId": employee.get("userId"),
                "name": employee.get("name") or employee.get("userId"),
                "title": employee.get("designation") or "",
                "role": "Member",
                "image": employee.get("profilePhoto"),
            },
            "children": [],
        }

    def unit_node(unit, path=None):
        path = path or set()
        unit_id = str(unit["_id"])
        if unit_id in path:
            return None
        head_ids = normalize_list(unit.get("headEmployeeIds"))
        junior_ids = normalize_list(unit.get("juniorEmployeeIds"))
        children = [unit_node(child, {*path, unit_id}) for child in child_map.get(unit_id, [])]
        if unit.get("unitType") == "function":
            excluded = set(head_ids + junior_ids)
            children.extend(
                person_node(employee) for employee in members.get(unit_id, [])
                if employee.get("userId") not in excluded
            )
        return {
            "expanded": True,
            "type": "unit",
            "data": {
                "id": unit_id,
                "name": unit.get("name"),
                "unitType": unit.get("unitType"),
                "heads": [
                    {"userId": value, "name": employee_map.get(value, {}).get("name") or value}
                    for value in head_ids
                ],
                "juniors": [
                    {"userId": value, "name": employee_map.get(value, {}).get("name") or value}
                    for value in junior_ids
                ],
            },
            "children": [child for child in children if child],
        }

    return [unit_node(unit) for unit in child_map.get("ROOT", [])]


@router.put("/employees/{employee_id}")
def update_employee(employee_id: str, data: dict):
    organization = resolve_employee_organization(
        data.get("manualFunctionIds", data.get("functionIds")),
        data.get("userId"),
    )
    verticals = organization["verticals"]
    reporting_officer_ids = organization["reportingOfficerIds"]

    update_data = {
        "name": data.get("name"),
        "nameHindi": data.get("nameHindi"),
        "designation": data.get("designation"),
        "designationHindi": data.get("designationHindi"),
        "userId": data.get("userId"),
        "phone": data.get("phone"),
        "gmail": data.get("gmail"),
        "dutyType": data.get("dutyType"),
        "category": normalize_list(data.get("category")),
        "verticals": verticals,
        "vertical": verticals[0] if verticals else None,
        "verticalIds": organization["verticalIds"],
        "sections": organization["sections"],
        "sectionIds": organization["sectionIds"],
        "departments": organization["departments"],
        "departmentIds": organization["departmentIds"],
        "department": organization["department"],
        "reportingOfficerIds": reporting_officer_ids,
        "reportingOfficerId": reporting_officer_ids[0] if reporting_officer_ids else None,
        "functionIds": organization["functionIds"],
        "manualFunctionIds": organization["manualFunctionIds"],
        "roleFunctionIds": organization["roleFunctionIds"],
        "intermediaryReportingId": organization["intermediaryReportingId"],
        "hodId": organization["hodId"],
    }

    password = data.get("password")
    if password:
        validate_password_policy(password)
        update_data["password"] = password if password.startswith("$2") else hash_password(password)

    employee_collection.update_one(
        {"_id": ObjectId(employee_id)},
        {"$set": update_data}
    )

    return {"message": "Employee updated successfully"}

############## Duty leave tyoe area start ############

def serialize_duty_leave(doc):
    return {
        "id": str(doc["_id"]),
        "dutyLeaveType_cat": doc.get("dutyLeaveType_cat"),
        "value": doc.get("value"),
        "status": doc.get("status", "Active"),
        "order": doc.get("order", 0)
    }


@router.post("/DutyLeaveType")
def create_dutyLeave(data: dict):
    return create_dutyLeave_logic(data)
    


@router.get("/DutyLeaveType/{dutyLeaveType_cat}")
def get_by_type(dutyLeaveType_cat: str):

    data = DutyLeave_collection.find(
        {"dutyLeaveType_cat": dutyLeaveType_cat}
    )

    return [serialize_duty_leave(item) for item in data]



@router.get("/DutyLeaveType")
def get_all_types():

    data = DutyLeave_collection.find()

    return [serialize_duty_leave(item) for item in data]


@router.put("/DutyLeaveType/{item_id}")
def update_duty_leave(item_id: str, data: dict):

    DutyLeave_collection.update_one(
        {"_id": ObjectId(item_id)},
        {
            "$set": {
                "value": data.get("value"),
                "status": data.get("status"),
                "dutyLeaveType_cat": data.get("dutyLeaveType_cat")
            }
        }
    )

    return {"message": "Updated successfully"}



@router.delete("/DutyLeaveType/{DutyLeaveType_id}")
def delete_DutyLeaveType(DutyLeaveType_id: str):

    DutyLeave_collection.delete_one(
        {"_id": ObjectId(DutyLeaveType_id)}
    )

    return {"message": "Deleted successfully"}



############## System setting for 2 person leave ############

@router.get("/settings/group-leave-rule")
def get_group_leave_rule():

    setting = system_settings_collection.find_one(
        {"settingName": "singleLeavePerGroupPerDay"}
    )

    if not setting:
        return {"enabled": False}

    return {"enabled": setting.get("enabled", False)}

@router.put("/settings/group-leave-rule")
def update_group_leave_rule(data: dict):

    enabled = data.get("enabled", False)

    system_settings_collection.update_one(
        {"settingName": "singleLeavePerGroupPerDay"},
        {"$set": {"enabled": enabled}},
        upsert=True
    )

    return {
        "message": "Setting updated",
        "enabled": enabled
    }


@router.get("/employees/export")
def export_employees():

    data = list(employee_collection.find({}, {"_id": 0}))  # remove ObjectId

    return {
        "count": len(data),
        "data": data
    }

@router.post("/employees/import")
def import_employees(data: list):

    # Optional: clear existing data
    employee_collection.delete_many({})

    if data:
        employee_collection.insert_many(data)

    return {
        "message": "Employees imported successfully",
        "count": len(data)
    }


@router.get("/employees/export-excel")
def export_employees_excel():

    data = list(employee_collection.find({}, {"_id": 0}))

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    if not data:
        return {"message": "No data"}

    # headers
    headers = list(data[0].keys())
    ws.append(headers)

    # rows
    for row in data:
        ws.append([
            ", ".join(str(item) for item in row.get(h, []))
            if isinstance(row.get(h), list)
            else row.get(h, "")
            for h in headers
        ])

    stream = BytesIO()
    wb.save(stream)

    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=employees.xlsx"
        }
    )


@router.post("/employees/import-excel")
async def import_employees_excel(file: UploadFile = File(...)):

    contents = ensure_upload_allowed(
        file,
        allowed_content_types={
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        },
        allowed_extensions={"xlsx", "xls"},
        max_bytes=5 * 1024 * 1024,
    )

    wb = load_workbook(BytesIO(contents))
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))

    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]

    employees_processed = 0

    for row in data_rows:

        emp_raw = dict(zip(headers, row))

        user_id = str(emp_raw.get("userId")).strip() if emp_raw.get("userId") else None

        if not user_id:
            continue  # skip invalid rows

        # ============================
        # CLEAN + MAP DATA
        # ============================

        emp_data = {
            "name": emp_raw.get("name"),
            "designation": emp_raw.get("designation"),
            "userId": user_id,
            "phone": emp_raw.get("phone"),
            "gmail": emp_raw.get("gmail"),

            "functionIds": normalize_list(emp_raw.get("functionIds")),

            "updatedAt": datetime.utcnow()
        }
        emp_data.update(resolve_employee_organization(emp_data["functionIds"], user_id))
        emp_data["vertical"] = (emp_data.get("verticals") or [None])[0]

        # ============================
        # UPSERT (UPDATE OR INSERT)
        # ============================

        employee_collection.update_one(
            {"userId": user_id},
            {
                "$set": emp_data,
                "$setOnInsert": {
                    "createdAt": datetime.utcnow()
                }
            },
            upsert=True
        )

        employees_processed += 1

    return {
        "message": "Excel processed successfully",
        "processed": employees_processed
    }



