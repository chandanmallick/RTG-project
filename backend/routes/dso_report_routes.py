import io
import re
from datetime import date, datetime, time, timedelta
from typing import Optional

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi.responses import StreamingResponse

from services.db_handler import MongoService
from routes.pipeline_routes import (
    build_unit_lookup,
    fetch_generation_outage_history_rows,
    make_legacy_session,
    to_float,
)


router = APIRouter(prefix="/api/dso-reports", tags=["DSO Report Preparation"])
COLLECTION = "dso_reports"
STATES = ("BIHAR", "JHARKHAND", "DVC", "ODISHA", "WB", "SIKKIM")

STATE_COLUMNS = {
    "BIHAR": {
        "schedule": ("ABTSC_PG - BSEB SCHD", "ABTSCPGBSEBSCHD"),
        "actual": ("SUDO__PG - BH_DRAWAL", "SUDOPGBHDRAWAL"),
    },
    "JHARKHAND": {
        "schedule": ("ABTSC_PG - JSEB SCHD", "ABTSCPGJSEBSCHD"),
        "actual": ("SUDO__PG - JH_DRAWAL", "SUDOPGJHDRAWAL"),
    },
    "DVC": {
        "schedule": ("ABTSC_PG - DVC SCHD", "ABTSCPGDVCSCHD"),
        "actual": ("SUDO__PG - DV_DRAWAL", "SUDOPGDVDRAWAL"),
    },
    "ODISHA": {
        "schedule": ("ABTSC_PG - GRIDCO SCHD", "ABTSCPGGRIDCOSCHD"),
        "actual": ("SUDO__PG - GR_DRAWAL", "SUDOPGGRDRAWAL"),
    },
    "WB": {
        "schedule": ("ABTSC_PG - WBSEB SCHD", "ABTSCPGWBSEBSCHD"),
        "actual": ("SUDO__PG - WB_DRAWAL", "SUDOPGWBDRAWAL"),
    },
    "SIKKIM": {
        "schedule": ("ABTSC_PG - SIKKIM SCHD", "ABTSCPGSIKKIMSCHD"),
        "actual": ("SUDO__PG - SI_DRAWAL", "SUDOPGSIDRAWAL"),
    },
}


def collection():
    return MongoService().db[COLLECTION]


def master_collection():
    return MongoService().db["pipeline_config"]


def compact(value):
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def json_safe(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def numeric(value):
    if value is None or value == "":
        return None
    try:
        result = float(value)
        return result if result == result else None
    except (TypeError, ValueError):
        match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
        return float(match.group()) if match else None


def display_time(value):
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and 0 <= value < 1:
        minutes = round(float(value) * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"
    text = str(value or "").strip()
    for pattern in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M", "%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, pattern).strftime("%H:%M")
        except ValueError:
            pass
    match = re.search(r"\b(\d{1,2}):(\d{2})", text)
    return f"{int(match.group(1)):02d}:{match.group(2)}" if match else text


def match_header(value):
    key = compact(value)
    if not key:
        return None
    if key in {"TIME", "TIMESTAMP", "DATETIME", "DATEANDTIME"} or key.startswith("TIME"):
        return "time"
    if "FREQUENCY" in key or key in {"FREQ", "HZ"}:
        return "frequency"
    if ("DEMAND" in key and ("EXCLUDINGESS" in key or "REGDEMAND" in key or "ERDEMAND" in key)):
        return "demand"
    for state, fields in STATE_COLUMNS.items():
        for kind, aliases in fields.items():
            if any(compact(alias) in key or key in compact(alias) for alias in aliases):
                return f"{state}.{kind}"
    return None


def locate_headers(sheet):
    best = None
    for row_number in range(1, min(sheet.max_row, 40) + 1):
        mapping = {}
        originals = {}
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row_number, column).value
            canonical = match_header(value)
            if canonical and canonical not in mapping:
                mapping[canonical] = column
                originals[canonical] = str(value)
        # SCADA PC exports place measurement names on row 1, but put only the
        # Date Time label (and measurement key numbers) on row 2. Treat column
        # A as Date Time and combine A2 with the row-1 measurement headers.
        if row_number == 1:
            a1_key = match_header(sheet.cell(1, 1).value)
            a2_key = match_header(sheet.cell(2, 1).value)
            if a1_key == "time" or a2_key == "time":
                mapping["time"] = 1
                originals["time"] = str(sheet.cell(1 if a1_key == "time" else 2, 1).value)
        score = len(mapping) + (4 if "time" in mapping else 0) + (4 if "frequency" in mapping else 0)
        if best is None or score > best[0]:
            best = (score, row_number, mapping, originals)
    if not best or "time" not in best[2] or "frequency" not in best[2]:
        raise ValueError("Could not locate Time and Frequency headers in the uploaded SCADA workbook.")
    return best[1], best[2], best[3]


def parse_scada(contents, report_type):
    # Normal mode is intentional here: repeated random cell access against an
    # openpyxl read-only worksheet is extremely slow for the 1,020-row SCADA
    # export, while this template is small enough to load safely in memory.
    workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
    candidates = []
    for sheet in workbook.worksheets:
        try:
            header_row, mapping, originals = locate_headers(sheet)
            candidates.append((len(mapping), sheet, header_row, mapping, originals))
        except ValueError:
            continue
    if not candidates:
        raise ValueError("No worksheet contains recognizable SCADA Time and Frequency columns.")
    _, sheet, header_row, mapping, originals = max(candidates, key=lambda item: item[0])
    missing = [name for name in ("demand",) if name not in mapping]
    if missing:
        raise ValueError("Required SCADA column is missing: Reg/ER Demand Excluding ESS Demand.")

    rows = []
    # Row 2 is always metadata/key numbers in the SCADA PC template, never
    # operational data. Actual samples therefore begin at row 3.
    for row_number in range(max(header_row + 1, 3), sheet.max_row + 1):
        raw_time = sheet.cell(row_number, mapping["time"]).value
        row_time = display_time(raw_time)
        if not row_time:
            continue
        try:
            hour, minute = [int(part) for part in row_time[:5].split(":")]
        except (TypeError, ValueError):
            continue
        minute_of_day = hour * 60 + minute
        if report_type == "evening" and minute_of_day > 16 * 60 + 59:
            continue
        record = {"time": row_time}
        for name, column in mapping.items():
            if name != "time":
                record[name] = numeric(sheet.cell(row_number, column).value)
        if record.get("frequency") is not None:
            rows.append(record)
    if not rows:
        raise ValueError("No valid SCADA rows were found in the report time window.")
    return rows, {
        "sheet": sheet.title,
        "header_row": header_row,
        "matched_headers": originals,
        "processed_rows": len(rows),
        "first_time": rows[0]["time"],
        "last_time": rows[-1]["time"],
    }


MORNING_HEADERS = {
    "demand": ("REGDEMANDEXCLUDINGESSDEMAND",),
    "frequency": ("FREQUENCY",),
    "alipurduar_pole_3": ("ALIPURDUARPOLE3",),
    "alipurduar_pole_4": ("ALIPURDUARPOLE4",),
    "bnc_pole_1": ("BNCPOLE1",),
    "bnc_pole_2": ("BNCPOLE2",),
    "agra_pole_1": ("AGRAPOLE1",),
    "agra_pole_2": ("AGRAPOLE2",),
    "talcher_kolar": ("TALCHERKOLARNET",),
    "jeypore_gazuwaka_1": ("JEYPOREGAZUAKA1",),
    "jeypore_gazuwaka_2": ("JEYPOREGAZUAKA2",),
    "gazuwaka_jeypore_1": ("GAZUAKAJEPORE1",),
    "gazuwaka_jeypore_2": ("GAZUAKAJEYPORE2",),
    "bheramara": ("BANGLADESHDRAWLHVDC",),
    "india_demand": ("DEMAND",),
    "solar": ("SOLAR",),
    "wind": ("WIND",),
    "gas": ("GAS",),
    "thermal": ("THERMAL",),
    "hydro": ("HYDRO",),
}


def parse_morning_scada(contents, report_date):
    workbook = openpyxl.load_workbook(io.BytesIO(contents), read_only=False, data_only=True)
    sheet = next((item for item in workbook.worksheets if item.title.lower() != "_config"), None)
    if not sheet:
        raise ValueError("No operational worksheet was found.")
    mapping = {"time": 1}
    originals = {"time": str(sheet.cell(3, 1).value or "Date & Time")}
    for column in range(2, sheet.max_column + 1):
        header = compact(sheet.cell(2, column).value)
        group = compact(sheet.cell(1, column).value)
        for canonical, aliases in MORNING_HEADERS.items():
            if canonical in mapping:
                continue
            if canonical == "india_demand" and group != "ALLINDIA":
                continue
            if canonical in {"solar", "wind", "gas", "thermal", "hydro"} and group != "ALLINDIA":
                continue
            if any(header == alias or alias in header for alias in aliases):
                mapping[canonical] = column
                originals[canonical] = str(sheet.cell(2, column).value)
                break
    required = {"demand", "frequency", "india_demand", "solar", "wind", "gas", "thermal", "hydro"}
    missing = sorted(required - set(mapping))
    if missing:
        raise ValueError(f"Required morning SCADA columns are missing: {', '.join(missing)}")
    yesterday = date.fromisoformat(report_date)
    today = yesterday + timedelta(days=1)
    rows = []
    # Row 3 contains only measurement keys. Operational samples start at row 4.
    for row_number in range(4, sheet.max_row + 1):
        raw_dt = sheet.cell(row_number, 1).value
        if not isinstance(raw_dt, datetime):
            continue
        if raw_dt.date() < yesterday or raw_dt.date() > today:
            continue
        if raw_dt.date() == today and (raw_dt.hour > 6 or (raw_dt.hour == 6 and raw_dt.minute > 59)):
            continue
        record = {"datetime": raw_dt, "time": raw_dt.strftime("%H:%M"), "date": raw_dt.date().isoformat()}
        for canonical, column in mapping.items():
            if canonical != "time":
                record[canonical] = numeric(sheet.cell(row_number, column).value)
        rows.append(record)
    if not rows:
        raise ValueError("No samples were found from yesterday 00:00 through today 06:59.")
    return rows, {
        "sheet": sheet.title,
        "header_rows": [1, 2],
        "skipped_key_row": 3,
        "matched_headers": originals,
        "processed_rows": len(rows),
        "first_timestamp": rows[0]["datetime"].isoformat(),
        "last_timestamp": rows[-1]["datetime"].isoformat(),
    }


def series_extreme(rows, field, mode):
    valid = [(row, row.get(field)) for row in rows if row.get(field) is not None]
    if not valid:
        return {"value": None, "time": ""}
    row, value = (max(valid, key=lambda item: item[1]) if mode == "max" else min(valid, key=lambda item: item[1]))
    return {"value": round(value, 3), "time": row["time"]}


def psp_morning_values(report_date):
    db = MongoService().db
    psp = db["psp_data"].find_one({"date": report_date}, {"_id": 0}) or {}
    curve = db["psp_curve_metrics"].find_one({"date": report_date}, {"_id": 0}) or {}
    requirement = next(
        (row for row in psp.get("pspstatedemandrequirement", []) if compact(row.get("STATE_NAME")) == "REGION"),
        {},
    )
    er_curve = (curve.get("metrics") or {}).get("ER") or {}
    profile = (psp.get("pspFrequencyProfile") or [{}])[0]
    max_min = (psp.get("pspFrequencyProfileMaxMin") or [{}])[0]
    states = {compact(row.get("STATE_NAME")): row for row in psp.get("pspTransnationalExchangeState", [])}
    lines = {compact(row.get("LINE_NAME")): row for row in psp.get("pspTransnationalExchangeLine", [])}
    nepal_line = next((row for key, row in lines.items() if "132KVBIHARNEPAL" in key), {})
    return {
        "available": bool(psp),
        "demand_frequency": {
            "max_demand_mw": numeric(requirement.get("MAX_DEMAND") or er_curve.get("max_demand")),
            "max_demand_time": str(requirement.get("MAX_DEMAND_TIME") or ""),
            "min_demand_mw": numeric(er_curve.get("min_demand")),
            "min_demand_time": str(requirement.get("TIME_AT_MIN_DEMAND") or ""),
            "max_frequency_hz": numeric(max_min.get("MAX_FREQ")),
            "max_frequency_time": str(max_min.get("MAX_TIME") or "")[:5],
            "min_frequency_hz": numeric(max_min.get("MIN_FREQ")),
            "min_frequency_time": str(max_min.get("MIN_TIME") or "")[:5],
        },
        "frequency_distribution": {
            "above_50_05_pct": numeric(profile.get("FREQ7_VALUE")),
            "within_band_pct": numeric(profile.get("FREQ5_VALUE")),
            "below_49_9_pct": numeric(profile.get("FREQ4_VALUE")),
        },
        "international_exchange": {
            "BHUTAN": {
                "schedule_mu": numeric(states.get("BHUTAN", {}).get("Scheduled_EX")),
                "actual_mu": numeric(states.get("BHUTAN", {}).get("ACTUAL_EX")),
            },
            "NEPAL ISTS": {
                "schedule_mu": numeric(states.get("NEPAL", {}).get("Scheduled_EX")),
                "actual_mu": numeric(states.get("NEPAL", {}).get("ACTUAL_EX")),
            },
            "BANGLADESH": {
                "schedule_mu": numeric(states.get("BANGLADESH", {}).get("Scheduled_EX")),
                "actual_mu": numeric(states.get("BANGLADESH", {}).get("ACTUAL_EX")),
            },
            "NEPAL BIHAR": {
                "schedule_mu": numeric(nepal_line.get("ENERGY_EXCHANGE")),
                "actual_mu": numeric(nepal_line.get("ENERGY_EXCHANGE")),
            },
        },
        "source": {
            "collection": "psp_data",
            "curve_collection": "psp_curve_metrics",
            "date": report_date,
            "curve_source": (curve.get("meta") or {}).get("file"),
        },
    }


def build_morning_results(rows, report_date):
    yesterday = report_date
    today = (date.fromisoformat(report_date) + timedelta(days=1)).isoformat()
    night_rows = [row for row in rows if row["date"] == today]
    yesterday_rows = [row for row in rows if row["date"] == yesterday]
    if not night_rows:
        raise ValueError("The upload does not contain today 00:00 through 06:59 data.")
    night = {
        "max_demand": series_extreme(night_rows, "demand", "max"),
        "min_demand": series_extreme(night_rows, "demand", "min"),
        "max_frequency": series_extreme(night_rows, "frequency", "max"),
        "min_frequency": series_extreme(night_rows, "frequency", "min"),
    }
    generation_fields = ("india_demand", "solar", "wind", "gas", "thermal", "hydro")
    today_generation = {field: series_extreme(night_rows, field, "max") for field in generation_fields}
    yesterday_generation = {field: series_extreme(yesterday_rows, field, "max") for field in generation_fields}
    today_generation["india_demand_min"] = series_extreme(night_rows, "india_demand", "min")
    yesterday_generation["india_demand_min"] = series_extreme(yesterday_rows, "india_demand", "min")
    last = night_rows[-1]
    def summed(*keys):
        values = [last.get(key) for key in keys if last.get(key) is not None]
        return round(sum(values), 3) if values else None
    hvdc = {
        "AGRA_ALIPURDUAR": {"label": "± 800 kV Agra-Alipurduar", "value_mw": summed("alipurduar_pole_3", "alipurduar_pole_4"), "region": "NR", "time": last["time"]},
        "AGRA_BNC": {"label": "± 800 kV Agra-BNC", "value_mw": summed("bnc_pole_1", "bnc_pole_2"), "region": "NR", "time": last["time"]},
        "TALCHER_KOLAR": {"label": "± 500 kV Talcher-Kolar", "value_mw": last.get("talcher_kolar"), "region": "SR", "time": last["time"]},
        "GAZUWAKA": {"label": "HVDC Gazuwaka", "value_mw": summed("jeypore_gazuwaka_1", "jeypore_gazuwaka_2"), "region": "SR", "time": last["time"]},
        "BHERAMARA": {"label": "HVDC Bheramara", "value_mw": last.get("bheramara"), "region": "BAN", "time": last["time"]},
    }
    hvdc["AGRA_ALIPURDUAR"]["label"] = "+/- 800 kV Agra-Alipurduar"
    hvdc["AGRA_BNC"]["label"] = "+/- 800 kV Agra-BNC"
    hvdc["TALCHER_KOLAR"]["label"] = "+/- 500 kV Talcher-Kolar"
    psp = psp_morning_values(report_date)
    # PSP curve metrics do not retain the min-demand timestamp; use the matching
    # uploaded yesterday curve time while retaining the PSP minimum value.
    if not psp["demand_frequency"].get("min_demand_time") and psp["demand_frequency"].get("min_demand_mw") is not None and yesterday_rows:
        target = psp["demand_frequency"]["min_demand_mw"]
        nearest = min(yesterday_rows, key=lambda row: abs((row.get("demand") or 0) - target))
        psp["demand_frequency"]["min_demand_time"] = nearest["time"]
        psp["source"]["min_demand_time_fallback"] = "uploaded SCADA nearest PSP curve minimum"
    return {
        "night_shift": night,
        "yesterday_psp": psp,
        "generation": {"today": today_generation, "yesterday": yesterday_generation},
        "hvdc": hvdc,
        "report_date": report_date,
        "today_date": today,
    }

def extrema(rows, field, mode):
    valid = [(index, row[field]) for index, row in enumerate(rows) if row.get(field) is not None]
    if not valid:
        return None, None, None
    index, value = (max(valid, key=lambda item: item[1]) if mode == "max" else min(valid, key=lambda item: item[1]))
    return round(value, 3), rows[index]["time"], index


def parse_external_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip().split("T")[0].split(" ")[0]
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def thermal_availability(report_date):
    db = MongoService()
    rows, from_cache, source_url = fetch_generation_outage_history_rows(
        db, report_date, report_date, session=make_legacy_session()
    )
    unit_lookup = build_unit_lookup(db)
    fuel_lookup = {}
    for unit in db.unit_collection.find({}, {"_id": 0, "Unit_Name": 1, "fuel_type": 1}):
        key = compact(unit.get("Unit_Name"))
        if key:
            fuel_lookup[key] = unit.get("fuel_type")
    revived, outage = {}, {}
    details = {"revived": [], "outage": []}
    target = date.fromisoformat(report_date)

    def occurred_by_1700(row, date_keys, time_keys):
        event_date = next((parse_external_date(row.get(key)) for key in date_keys if row.get(key)), None)
        if event_date != target:
            return False
        raw_time = next((row.get(key) for key in time_keys if row.get(key)), "")
        match = re.search(r"(\d{1,2}):(\d{2})", str(raw_time))
        return not match or (int(match.group(1)), int(match.group(2))) <= (17, 0)

    for row in rows:
        name = row.get("ELEMENT_NAME") or row.get("elementName") or row.get("ElementName") or ""
        unit = unit_lookup.get(" ".join(str(name).strip().upper().split())) or {}
        fuel = compact(fuel_lookup.get(compact(name)) or row.get("FUEL_TYPE") or row.get("FUEL"))
        if fuel and not any(word in fuel for word in ("THERMAL", "COAL", "LIGNITE", "GAS", "DIESEL")):
            continue
        capacity = to_float(
            row.get("INSTALLED_CAPACITY") or row.get("installedCapacity")
            or unit.get("installed_capacity")
        )
        identity = compact(name) or f"ROW{len(details['outage'])}"
        if occurred_by_1700(
            row,
            ("REVIVAL_DATE", "revivalDate", "Synchronization_Time"),
            ("REVIVAL_TIME", "revivalTime", "Synchronization_Time"),
        ):
            revived[identity] = capacity
            details["revived"].append({
                "unit": name,
                "capacity_mw": capacity,
                "reason": row.get("OUT_REASON") or row.get("REASON") or row.get("reason") or "",
            })
        if occurred_by_1700(
            row,
            ("OUTAGE_DATE", "outageDate", "Tripped_Time"),
            ("OUTAGE_TIME", "outageTime", "Tripped_Time"),
        ):
            outage[identity] = capacity
            details["outage"].append({
                "unit": name,
                "capacity_mw": capacity,
                "reason": row.get("OUT_REASON") or row.get("REASON") or row.get("reason") or "",
            })
    revived_mw = round(sum(revived.values()), 3)
    outage_mw = round(sum(outage.values()), 3)
    return {
        "revived_capacity_mw": revived_mw,
        "outage_capacity_mw": outage_mw,
        "net_capacity_change_mw": round(revived_mw - outage_mw, 3),
        "units": details,
        "source_url": source_url,
        "from_cache": from_cache,
    }


def build_results(rows, limits):
    max_demand, max_demand_time, _ = extrema(rows, "demand", "max")
    min_demand, min_demand_time, _ = extrema(rows, "demand", "min")
    max_frequency, max_frequency_time, max_frequency_index = extrema(rows, "frequency", "max")
    min_frequency, min_frequency_time, min_frequency_index = extrema(rows, "frequency", "min")
    count = len(rows)
    frequency_distribution = {
        "above_50_05_pct": round(100 * sum((row.get("frequency") or 0) > 50.05 for row in rows) / count, 3),
        "within_band_pct": round(100 * sum(49.9 <= (row.get("frequency") or 0) <= 50.05 for row in rows) / count, 3),
        "below_49_9_pct": round(100 * sum((row.get("frequency") or 0) < 49.9 for row in rows) / count, 3),
    }
    state_results = {}
    for state in STATES:
        schedule_field, actual_field = f"{state}.schedule", f"{state}.actual"
        max_schedule, max_schedule_time, _ = extrema(rows, schedule_field, "max")
        max_actual, max_actual_time, _ = extrema(rows, actual_field, "max")
        atc = numeric((limits.get(state) or {}).get("atc"))
        if state == "DVC" and atc is not None and atc < 0:
            min_actual, _, _ = extrema(rows, actual_field, "min")
            actual_violation = round(atc - min_actual, 3) if min_actual is not None and min_actual < atc else None
        else:
            actual_violation = round(max_actual - atc, 3) if max_actual is not None and atc is not None and max_actual > atc else None
        low_row = rows[min_frequency_index] if min_frequency_index is not None else {}
        high_row = rows[max_frequency_index] if max_frequency_index is not None else {}
        low_actual, low_schedule = low_row.get(actual_field), low_row.get(schedule_field)
        high_actual, high_schedule = high_row.get(actual_field), high_row.get(schedule_field)
        state_results[state] = {
            "ttc_limit_mw": numeric((limits.get(state) or {}).get("ttc")),
            "atc_limit_mw": atc,
            "max_schedule_mw": max_schedule,
            "max_schedule_time": max_schedule_time,
            "max_actual_mw": max_actual,
            "max_actual_time": max_actual_time,
            "atc_violation_mw": actual_violation,
            "od_at_min_frequency_mw": round(low_actual - low_schedule, 3) if low_actual is not None and low_schedule is not None else None,
            "ud_at_max_frequency_mw": round(high_schedule - high_actual, 3) if high_actual is not None and high_schedule is not None else None,
        }
    return {
        "demand_frequency": {
            "max_demand_mw": max_demand, "max_demand_time": max_demand_time,
            "min_demand_mw": min_demand, "min_demand_time": min_demand_time,
            "max_frequency_hz": max_frequency, "max_frequency_time": max_frequency_time,
            "min_frequency_hz": min_frequency, "min_frequency_time": min_frequency_time,
        },
        "frequency_distribution": frequency_distribution,
        "states": state_results,
    }


@router.get("/master")
async def get_master():
    doc = master_collection().find_one({"config_type": "DSO_TTC_ATC"}, {"_id": 0})
    return {"success": True, "limits": (doc or {}).get("limits") or {state: {"ttc": "", "atc": ""} for state in STATES}}


@router.put("/master")
async def save_master(payload: dict):
    limits = {}
    supplied = payload.get("limits") or {}
    for state in STATES:
        limits[state] = {"ttc": numeric((supplied.get(state) or {}).get("ttc")), "atc": numeric((supplied.get(state) or {}).get("atc"))}
    now = datetime.utcnow().isoformat()
    master_collection().update_one(
        {"config_type": "DSO_TTC_ATC"},
        {"$set": {"config_type": "DSO_TTC_ATC", "limits": limits, "updated_at": now}},
        upsert=True,
    )
    return {"success": True, "limits": limits, "updated_at": now}


@router.post("/process")
async def process_report(
    report_date: str = Form(...),
    report_type: str = Form("evening"),
    important_events: str = Form(""),
    sic_name: str = Form(""),
    overwrite: bool = Form(False),
    file: UploadFile = File(...),
):
    if report_type not in {"evening", "morning"}:
        raise HTTPException(400, "Report type must be evening or morning.")
    try:
        date.fromisoformat(report_date)
        contents = await file.read()
        if not contents:
            raise ValueError("The uploaded SCADA workbook is empty.")
        key = {"doc_type": "processed_report", "report_type": report_type, "report_date": report_date}
        if collection().find_one(key, {"_id": 1}) and not overwrite:
            raise HTTPException(409, "A saved report already exists for this date. Select overwrite to replace it.")
        now = datetime.utcnow().isoformat()
        if report_type == "morning":
            rows, input_summary = parse_morning_scada(contents, report_date)
            document = {
                "doc_type": "processed_report",
                "report_type": report_type,
                "report_date": report_date,
                "report_window": "Yesterday 00:00-Today 06:59",
                "important_events": important_events.strip(),
                "signoff_regards": "Regards",
                "signoff_name": sic_name.strip(),
                "source_file_name": file.filename,
                "input_summary": input_summary,
                "morning_results": build_morning_results(rows, report_date),
                "processed_at": now,
            }
        else:
            rows, input_summary = parse_scada(contents, report_type)
            master = master_collection().find_one({"config_type": "DSO_TTC_ATC"}, {"_id": 0}) or {}
            document = {
                "doc_type": "processed_report",
                "report_type": report_type,
                "report_date": report_date,
                "report_window": "00:00-16:59",
                "important_events": important_events.strip(),
                "signoff_regards": "Regards",
                "signoff_name": sic_name.strip(),
                "source_file_name": file.filename,
                "input_summary": input_summary,
                "results": build_results(rows, master.get("limits") or {}),
                "thermal_availability": thermal_availability(report_date),
                "processed_at": now,
            }
        collection().replace_one(key, document, upsert=True)
        document.pop("_id", None)
        return {"success": True, "report": document}
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(502, f"Report processing failed: {exc}") from exc


@router.get("/{report_type}/{report_date}")
async def get_report(report_type: str, report_date: str):
    doc = collection().find_one(
        {"doc_type": "processed_report", "report_type": report_type, "report_date": report_date},
        {"_id": 0},
    )
    return {"success": True, "report": doc}


@router.delete("/{report_type}/{report_date}")
async def delete_report(report_type: str, report_date: str):
    if report_type not in {"evening", "morning"}:
        raise HTTPException(400, "Report type must be evening or morning.")
    result = collection().delete_one({
        "doc_type": "processed_report",
        "report_type": report_type,
        "report_date": report_date,
    })
    if not result.deleted_count:
        raise HTTPException(404, "Saved report was not found.")
    return {"success": True, "message": "Saved report deleted."}


@router.put("/{report_type}/{report_date}")
async def update_report(report_type: str, report_date: str, payload: dict):
    reports = collection()
    existing = reports.find_one({
        "doc_type": "processed_report",
        "report_type": report_type,
        "report_date": report_date,
    })
    if not existing:
        raise HTTPException(404, "Process the report before editing it.")

    if report_type == "morning":
        morning_results = payload.get("morning_results")
        if not isinstance(morning_results, dict):
            raise HTTPException(400, "Morning report values are required.")
        now = datetime.utcnow().isoformat()
        reports.update_one(
            {"_id": existing["_id"]},
            {"$set": {
                "morning_results": morning_results,
                "important_events": str(payload.get("important_events") or "").strip(),
                "signoff_regards": str(payload.get("signoff_regards") or "Regards").strip(),
                "signoff_name": str(payload.get("signoff_name") or "").strip(),
                "edited_at": now,
            }},
        )
        updated = reports.find_one({"_id": existing["_id"]}, {"_id": 0})
        return {"success": True, "report": updated, "edited_at": now}

    incoming_metrics = payload.get("demand_frequency") or {}
    metrics = {}
    for key in ("max_demand_mw", "min_demand_mw", "max_frequency_hz", "min_frequency_hz"):
        metrics[key] = numeric(incoming_metrics.get(key))
    for key in ("max_demand_time", "min_demand_time", "max_frequency_time", "min_frequency_time"):
        metrics[key] = str(incoming_metrics.get(key) or "").strip()

    incoming_frequency = payload.get("frequency_distribution") or {}
    frequency_distribution = {
        key: numeric(incoming_frequency.get(key))
        for key in ("above_50_05_pct", "within_band_pct", "below_49_9_pct")
    }

    incoming_thermal = payload.get("thermal_availability") or {}
    thermal_updates = {
        "revived_capacity_mw": numeric(incoming_thermal.get("revived_capacity_mw")),
        "outage_capacity_mw": numeric(incoming_thermal.get("outage_capacity_mw")),
        "net_capacity_change_mw": numeric(incoming_thermal.get("net_capacity_change_mw")),
        "revived_details": str(incoming_thermal.get("revived_details") or "").strip(),
        "outage_details": str(incoming_thermal.get("outage_details") or "").strip(),
    }

    incoming_states = payload.get("states") or {}
    states = {}
    numeric_state_keys = (
        "ttc_limit_mw", "atc_limit_mw", "max_schedule_mw", "max_actual_mw",
        "atc_violation_mw", "od_at_min_frequency_mw", "ud_at_max_frequency_mw",
    )
    for state in STATES:
        source = incoming_states.get(state) or {}
        states[state] = {key: numeric(source.get(key)) for key in numeric_state_keys}
        states[state]["max_schedule_time"] = str(source.get("max_schedule_time") or "").strip()
        states[state]["max_actual_time"] = str(source.get("max_actual_time") or "").strip()

    now = datetime.utcnow().isoformat()
    updates = {
        "results.demand_frequency": metrics,
        "results.frequency_distribution": frequency_distribution,
        "results.states": states,
        "thermal_availability.revived_capacity_mw": thermal_updates["revived_capacity_mw"],
        "thermal_availability.outage_capacity_mw": thermal_updates["outage_capacity_mw"],
        "thermal_availability.net_capacity_change_mw": thermal_updates["net_capacity_change_mw"],
        "thermal_availability.revived_details": thermal_updates["revived_details"],
        "thermal_availability.outage_details": thermal_updates["outage_details"],
        "major_od_text": str(payload.get("major_od_text") or "").strip(),
        "major_ud_text": str(payload.get("major_ud_text") or "").strip(),
        "important_events": str(payload.get("important_events") or "").strip(),
        "signoff_regards": str(payload.get("signoff_regards") or "Regards").strip(),
        "signoff_name": str(payload.get("signoff_name") or "").strip(),
        "edited_at": now,
    }
    reports.update_one({"_id": existing["_id"]}, {"$set": updates})
    updated = reports.find_one({"_id": existing["_id"]}, {"_id": 0})
    return {"success": True, "report": updated, "edited_at": now}


def morning_excel_response(doc, report_date):
    from openpyxl.styles import Border, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DSO Morning Report"
    sheet.sheet_view.showGridLines = False
    navy, blue, pale_blue, pale_green, white = "081F5C", "0057B7", "EAF2FF", "E9F8F0", "FFFFFF"
    border = Border(
        left=Side(style="thin", color="94A3B8"),
        right=Side(style="thin", color="94A3B8"),
        top=Side(style="thin", color="94A3B8"),
        bottom=Side(style="thin", color="94A3B8"),
    )

    def style(cell_range, fill=None, bold=False, color="0F172A", center=False, size=11):
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border
                cell.font = Font(name="Arial", size=size, bold=bold, color=color)
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

    def title(cell_range, value, fill=blue, color=white, size=11):
        sheet.merge_cells(cell_range)
        sheet[cell_range.split(":")[0]] = value
        style(cell_range, fill=fill, bold=True, color=color, center=True, size=size)

    results = doc.get("morning_results") or {}
    night = results.get("night_shift") or {}
    psp = results.get("yesterday_psp") or {}
    yesterday = psp.get("demand_frequency") or {}
    frequency = psp.get("frequency_distribution") or {}
    date_text = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    today_text = datetime.strptime(results.get("today_date") or report_date, "%Y-%m-%d").strftime("%d-%m-%Y")

    title("A1:L1", f"DSO Morning Shift Report — {today_text}", fill=navy, size=14)
    title("A3:E3", f"Demand / Frequency — Night Shift after 00:00 Hrs {today_text}", fill=pale_green, color="006845")
    title("F3:J3", f"During Yesterday {date_text} — PSP Report Database", fill=pale_blue, color=navy)
    sheet["A4"], sheet["B4"], sheet["C4"], sheet["D4"], sheet["E4"] = "Parameter", "Value", "Time", "Unit", "Source"
    sheet["F4"], sheet["G4"], sheet["H4"], sheet["I4"], sheet["J4"] = "Parameter", "Value", "Time", "Unit", "Source"
    labels = (
        ("Max demand met", "max_demand", "max_demand_mw", "max_demand_time", "MW"),
        ("Min demand met", "min_demand", "min_demand_mw", "min_demand_time", "MW"),
        ("Max freq.", "max_frequency", "max_frequency_hz", "max_frequency_time", "Hz"),
        ("Min freq.", "min_frequency", "min_frequency_hz", "min_frequency_time", "Hz"),
    )
    for row_number, (label, night_key, psp_value, psp_time, unit) in enumerate(labels, 5):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, (night.get(night_key) or {}).get("value"))
        sheet.cell(row_number, 3, (night.get(night_key) or {}).get("time"))
        sheet.cell(row_number, 4, unit)
        sheet.cell(row_number, 5, "Uploaded SCADA")
        sheet.cell(row_number, 6, label)
        sheet.cell(row_number, 7, yesterday.get(psp_value))
        sheet.cell(row_number, 8, yesterday.get(psp_time))
        sheet.cell(row_number, 9, unit)
        sheet.cell(row_number, 10, "PSP database")
    style("A4:J8")
    style("A4:J4", fill=pale_blue, bold=True, center=True)

    title("K3:L3", f"Frequency During Yesterday {date_text}", fill=pale_green, color="006845")
    sheet["K4"], sheet["L4"] = "Frequency", "% of time"
    for row_number, values in enumerate((
        (">50.05", frequency.get("above_50_05_pct")),
        ("within band", frequency.get("within_band_pct")),
        ("<49.9", frequency.get("below_49_9_pct")),
    ), 5):
        sheet.cell(row_number, 11, values[0])
        sheet.cell(row_number, 12, values[1])
    style("K4:L8")
    style("K4:L4", fill=pale_blue, bold=True, center=True)

    title("A10:F10", "Demand / Generation", fill=pale_green, color="006845")
    generation = results.get("generation") or {}
    sheet["A11"], sheet["B11"], sheet["C11"] = "Parameter", f"Today {today_text}", "Time"
    sheet["D11"], sheet["E11"], sheet["F11"] = "Parameter", f"Yesterday {date_text}", "Time"
    generation_labels = (
        ("All India Demand Max (GW)", "india_demand"),
        ("All India Demand Min (GW)", "india_demand_min"),
        ("All India Solar Generation Max (GW)", "solar"),
        ("All India Wind Generation Max (GW)", "wind"),
        ("All India Gas Generation Max (GW)", "gas"),
        ("All India Thermal Generation Max (GW)", "thermal"),
        ("All India Hydro Generation Max (GW)", "hydro"),
    )
    for row_number, (label, key) in enumerate(generation_labels, 12):
        today_value = ((generation.get("today") or {}).get(key) or {})
        yesterday_value = ((generation.get("yesterday") or {}).get(key) or {})
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, (numeric(today_value.get("value")) / 1000) if numeric(today_value.get("value")) is not None else None)
        sheet.cell(row_number, 3, today_value.get("time"))
        sheet.cell(row_number, 4, label)
        sheet.cell(row_number, 5, (numeric(yesterday_value.get("value")) / 1000) if numeric(yesterday_value.get("value")) is not None else None)
        sheet.cell(row_number, 6, yesterday_value.get("time"))
    style("A11:F18")
    style("A11:F11", fill=pale_blue, bold=True, center=True)

    title("G10:L10", "HVDC Details — Last registered value today", fill=blue)
    sheet["G11"], sheet["H11"], sheet["I11"], sheet["J11"] = "Link", "Value (MW)", "Time", "Region"
    hvdc = results.get("hvdc") or {}
    for row_number, item in enumerate(hvdc.values(), 12):
        sheet.cell(row_number, 7, item.get("label"))
        sheet.cell(row_number, 8, item.get("value_mw"))
        sheet.cell(row_number, 9, item.get("time"))
        sheet.cell(row_number, 10, item.get("region"))
    style("G11:J16")
    style("G11:J11", fill=pale_blue, bold=True, center=True)

    title("A19:F19", "International Exchange During Yesterday (Import +ve / Export -ve)", fill=pale_green, color="006845")
    sheet["A20"], sheet["B20"], sheet["C20"] = "Entity", "Schedule (MU)", "Actual (MU)"
    exchange = psp.get("international_exchange") or {}
    for row_number, name in enumerate(("BHUTAN", "NEPAL ISTS", "BANGLADESH", "NEPAL BIHAR"), 21):
        values = exchange.get(name) or {}
        sheet.cell(row_number, 1, name.title())
        sheet.cell(row_number, 2, values.get("schedule_mu"))
        sheet.cell(row_number, 3, values.get("actual_mu"))
    style("A20:C24")
    style("A20:C20", fill=pale_blue, bold=True, center=True)

    title("A26:L26", "Important Events (FTC/GD/GI/Load crash etc.)", fill=pale_green, color="006845")
    sheet.merge_cells("A27:L29")
    sheet["A27"] = doc.get("important_events") or "NIL"
    style("A27:L29")
    sheet["A31"] = doc.get("signoff_regards") or "Regards"
    sheet["A32"] = doc.get("signoff_name") or "Ashoke Kumar Basak, SIC ERLDC"
    sheet["A31"].font = Font(name="Arial", size=10, bold=True)
    sheet["A32"].font = Font(name="Arial", size=10)

    for letter in "ABCDEFGHIJKL":
        sheet.column_dimensions[letter].width = 19
    sheet.column_dimensions["A"].width = 31
    sheet.column_dimensions["D"].width = 31
    sheet.column_dimensions["G"].width = 27
    for row_number in range(5, 33):
        sheet.row_dimensions[row_number].height = 23
    sheet.freeze_panes = "A3"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = "A1:L32"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"DSO_Morning_{report_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{report_type}/{report_date}/excel")
async def download_report(report_type: str, report_date: str):
    doc = collection().find_one(
        {"doc_type": "processed_report", "report_type": report_type, "report_date": report_date},
        {"_id": 0},
    )
    if not doc:
        raise HTTPException(404, "Processed report was not found.")
    if report_type == "morning":
        return morning_excel_response(doc, report_date)
    from openpyxl.styles import Border, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DSO Report"
    sheet.sheet_view.showGridLines = False
    navy, blue, pale_blue, pale_green, white = "081F5C", "0057B7", "EAF2FF", "E9F8F0", "FFFFFF"
    thin = Side(style="thin", color="1F2937")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_range(cell_range, fill=None, bold=False, color="0F172A", center=False, size=11):
        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border
                cell.font = Font(name="Arial", size=size, bold=bold, color=color)
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

    def merged_title(cell_range, text, fill=blue, font_color=white, size=11):
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":")[0]]
        cell.value = text
        style_range(cell_range, fill=fill, bold=True, color=font_color, center=True, size=size)

    title_date = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    merged_title(
        "A1:L1",
        f"Important information update for ER Grid (since 00:00 hrs to 17:00 hrs) for {title_date}",
        fill=navy,
        size=14,
    )
    sheet.row_dimensions[1].height = 27

    metrics = doc["results"]["demand_frequency"]
    freq = doc["results"]["frequency_distribution"]
    sheet["A3"], sheet["B3"], sheet["C3"] = "", "MW/Hz", "Time (Hrs)"
    demand_rows = (
        ("Max demand met", metrics.get("max_demand_mw"), metrics.get("max_demand_time")),
        ("Min demand met", metrics.get("min_demand_mw"), metrics.get("min_demand_time")),
        ("Max freq.", metrics.get("max_frequency_hz"), metrics.get("max_frequency_time")),
        ("Min freq.", metrics.get("min_frequency_hz"), metrics.get("min_frequency_time")),
    )
    for row_number, values in enumerate(demand_rows, 4):
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
    style_range("A3:C7")
    style_range("A3:C3", fill=pale_blue, bold=True, center=True)
    for row_number in range(4, 8):
        sheet.cell(row_number, 1).font = Font(name="Arial", size=10, bold=True)
        sheet.cell(row_number, 2).number_format = "0.000"
        sheet.cell(row_number, 2).alignment = Alignment(horizontal="center")
        sheet.cell(row_number, 3).alignment = Alignment(horizontal="center")

    sheet["F3"], sheet["G3"] = "Frequency", "% of time"
    frequency_rows = (
        (">50.05", freq.get("above_50_05_pct")),
        ("within band", freq.get("within_band_pct")),
        ("<49.9", freq.get("below_49_9_pct")),
    )
    for row_number, values in enumerate(frequency_rows, 4):
        sheet.cell(row_number, 6, values[0])
        sheet.cell(row_number, 7, values[1])
    style_range("F3:G6")
    style_range("F3:G3", fill=pale_blue, bold=True, center=True)
    style_range("F4:G6", center=True)

    merged_title("A9:E9", "Thermal generation availability change from 00:00 hrs to 17:00 hrs", fill=pale_green, font_color="006845")
    sheet.merge_cells("A10:B10")
    sheet.merge_cells("C10:D10")
    sheet["A10"], sheet["C10"], sheet["E10"] = "Revived capacity (MW)", "Outage capacity (MW)", "Net Capacity addition (+)/ reduction (-) (MW)"
    sheet.merge_cells("A11:B11")
    sheet.merge_cells("C11:D11")
    thermal = doc.get("thermal_availability") or {}
    sheet["A11"], sheet["C11"], sheet["E11"] = thermal.get("revived_capacity_mw"), thermal.get("outage_capacity_mw"), thermal.get("net_capacity_change_mw")
    sheet.merge_cells("A12:B17")
    sheet.merge_cells("C12:D17")
    revived_units = thermal.get("units", {}).get("revived", [])
    outage_units = thermal.get("units", {}).get("outage", [])
    calculated_revived_details = "\n".join(f"{item.get('unit') or '-'} ({item.get('capacity_mw', 0):g} MW)" for item in revived_units)
    calculated_outage_details = "\n".join(
        f"{item.get('unit') or '-'} ({item.get('capacity_mw', 0):g} MW)"
        + (f" [{item.get('reason')}]" if item.get("reason") else "")
        for item in outage_units
    )
    sheet["A12"] = thermal.get("revived_details") or calculated_revived_details or "NIL"
    sheet["C12"] = thermal.get("outage_details") or calculated_outage_details or "NIL"
    style_range("A10:E17")
    style_range("A10:E10", fill=pale_blue, bold=True, center=True)
    style_range("A11:E11", bold=True, center=True)
    sheet["A12"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet["C12"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    merged_title("F9:L9", "TTC/ATC Limit (+ve Import, -ve Export)", fill=pale_blue, font_color=navy)
    sheet["F10"] = "STATE"
    for index, state in enumerate(STATES, 7):
        sheet.cell(10, index, state)
    row_labels = ("TTC LIMIT", "ATC LIMIT", "MAX SCHEDULE", "MAX ACTUAL", "VIOLATION")
    for row_number, label in enumerate(row_labels, 11):
        sheet.cell(row_number, 6, label)
    for column, state in enumerate(STATES, 7):
        values = doc["results"]["states"].get(state) or {}
        sheet.cell(11, column, f"{values.get('ttc_limit_mw'):g} (0-24 Hrs.)" if values.get("ttc_limit_mw") is not None else "—")
        sheet.cell(12, column, f"{values.get('atc_limit_mw'):g} (0-24 Hrs.)" if values.get("atc_limit_mw") is not None else "—")
        sheet.cell(13, column, f"{values.get('max_schedule_mw'):.0f}\n({values.get('max_schedule_time') or '—'})" if values.get("max_schedule_mw") is not None else "—")
        sheet.cell(14, column, f"{values.get('max_actual_mw'):.0f}\n({values.get('max_actual_time') or '—'})" if values.get("max_actual_mw") is not None else "—")
        sheet.cell(15, column, f"{values.get('atc_violation_mw'):.0f} MW" if values.get("atc_violation_mw") is not None else "NIL")
    style_range("F10:L15", center=True)
    style_range("F10:L10", fill=pale_blue, bold=True, center=True)
    for row_number in range(11, 16):
        sheet.cell(row_number, 6).font = Font(name="Arial", size=9, bold=True)

    states = doc["results"]["states"]
    od_states = [f"{state.title()} ({values.get('od_at_min_frequency_mw'):.0f} MW)" for state, values in states.items() if (values.get("od_at_min_frequency_mw") or 0) > 0]
    ud_states = [f"{state.title()} ({values.get('ud_at_max_frequency_mw'):.0f} MW)" for state, values in states.items() if (values.get("ud_at_max_frequency_mw") or 0) > 0]
    def formatted(value, digits=3):
        return f"{value:.{digits}f}" if value is not None else "—"

    merged_title("A19:L19", "Major OD/UD by states/generators:", fill=pale_green, font_color="006845")
    sheet.merge_cells("A20:L22")
    default_od_text = (
        f"Freq. touched {formatted(metrics.get('min_frequency_hz'))} Hz at {metrics.get('min_frequency_time') or '—'} Hrs, "
        f"OD by states in ER: {', '.join(od_states) or 'NIL'}"
    )
    default_ud_text = (
        f"Freq. touched {formatted(metrics.get('max_frequency_hz'))} Hz at {metrics.get('max_frequency_time') or '—'} Hrs, "
        f"UD by states in ER: {', '.join(ud_states) or 'NIL'}"
    )
    sheet["A20"] = f"{doc.get('major_od_text') or default_od_text}\n{doc.get('major_ud_text') or default_ud_text}"
    style_range("A20:L22")

    merged_title("A24:L24", "Important Events (FTC/GD/GI/Load crash etc.):", fill=pale_green, font_color="006845")
    sheet.merge_cells("A25:L27")
    sheet["A25"] = doc.get("important_events") or "NIL"
    style_range("A25:L27")
    sheet["A29"] = doc.get("signoff_regards") or "Regards"
    sheet["A30"] = doc.get("signoff_name") or "Ashoke Kumar Basak, SIC ERLDC"
    sheet["A29"].font = Font(name="Arial", size=10, bold=True)
    sheet["A30"].font = Font(name="Arial", size=10)

    widths = {"A": 22, "B": 18, "C": 22, "D": 18, "E": 22, "F": 15}
    for letter in "GHIJKL":
        widths[letter] = 17
    for letter, width in widths.items():
        sheet.column_dimensions[letter].width = width
    for row_number in range(12, 18):
        sheet.row_dimensions[row_number].height = 25
    sheet.row_dimensions[20].height = 32
    sheet.row_dimensions[25].height = 28
    sheet.freeze_panes = "A3"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_area = "A1:L30"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"DSO_{report_type.title()}_{report_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def report_pdf_response(doc, report_type, report_date):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    pdf = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title=f"DSO {report_type.title()} Report {report_date}",
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle("DsoBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=11)
    centered = ParagraphStyle("DsoCentered", parent=body, alignment=TA_CENTER)
    section = ParagraphStyle("DsoSection", parent=body, fontName="Helvetica-Bold", textColor=colors.HexColor("#006845"), fontSize=11, leading=13)
    title_style = ParagraphStyle("DsoTitle", parent=body, fontName="Helvetica-Bold", textColor=colors.white, fontSize=15, leading=18, alignment=TA_CENTER)
    story = []

    def p(value, center=False):
        text = str(value if value not in (None, "") else "-").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(text.replace("\n", "<br/>"), centered if center else body)

    def add_title(text):
        table = Table([[Paragraph(text, title_style)]], colWidths=[277 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#081F5C")),
            ("BOX", (0, 0), (-1, -1), .7, colors.HexColor("#081F5C")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.extend([table, Spacer(1, 5)])

    def add_section(text):
        story.extend([Paragraph(text, section), Spacer(1, 3)])

    def add_table(rows, widths=None):
        converted = [[p(cell, center=column > 0) for column, cell in enumerate(row)] for row in rows]
        table = Table(converted, colWidths=widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF2FF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#004DA8")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), .45, colors.HexColor("#AFC6E5")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.extend([table, Spacer(1, 7)])

    def value(number, digits=2):
        converted = numeric(number)
        return "-" if converted is None else f"{converted:,.{digits}f}"

    formatted_date = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d-%m-%Y")
    add_title(f"DSO {report_type.title()} Report - {formatted_date}")

    if report_type == "morning":
        results = doc.get("morning_results") or {}
        night = results.get("night_shift") or {}
        psp = results.get("yesterday_psp") or {}
        yesterday = psp.get("demand_frequency") or {}
        add_section("Demand and Frequency - Night Shift / During Yesterday")
        rows = [["Parameter", "Night value", "Time", "Yesterday PSP", "Time", "Unit"]]
        for label, night_key, psp_key, time_key, unit in (
            ("Max demand met", "max_demand", "max_demand_mw", "max_demand_time", "MW"),
            ("Min demand met", "min_demand", "min_demand_mw", "min_demand_time", "MW"),
            ("Max frequency", "max_frequency", "max_frequency_hz", "max_frequency_time", "Hz"),
            ("Min frequency", "min_frequency", "min_frequency_hz", "min_frequency_time", "Hz"),
        ):
            rows.append([
                label,
                value((night.get(night_key) or {}).get("value"), 3 if unit == "Hz" else 0),
                (night.get(night_key) or {}).get("time"),
                value(yesterday.get(psp_key), 3 if unit == "Hz" else 0),
                yesterday.get(time_key),
                unit,
            ])
        add_table(rows, [50 * mm, 35 * mm, 27 * mm, 35 * mm, 27 * mm, 20 * mm])

        distribution = psp.get("frequency_distribution") or {}
        add_table([
            ["Frequency during yesterday", "> 50.05 Hz", "Within band", "< 49.9 Hz"],
            ["% of time", value(distribution.get("above_50_05_pct")), value(distribution.get("within_band_pct")), value(distribution.get("below_49_9_pct"))],
        ], [55 * mm, 40 * mm, 40 * mm, 40 * mm])

        add_section("Demand / Generation")
        generation = results.get("generation") or {}
        rows = [["Parameter", "Today 00:00-06:59 (GW)", "Time", "Yesterday 00:00-23:59 (GW)", "Time"]]
        for label, key in (
            ("All India Demand Max", "india_demand"),
            ("All India Demand Min", "india_demand_min"),
            ("Solar Generation Max", "solar"),
            ("Wind Generation Max", "wind"),
            ("Gas Generation Max", "gas"),
            ("Thermal Generation Max", "thermal"),
            ("Hydro Generation Max", "hydro"),
        ):
            current = (generation.get("today") or {}).get(key) or {}
            previous = (generation.get("yesterday") or {}).get(key) or {}
            current_gw = numeric(current.get("value"))
            previous_gw = numeric(previous.get("value"))
            rows.append([
                label,
                value(current_gw / 1000 if current_gw is not None else None, 3),
                current.get("time"),
                value(previous_gw / 1000 if previous_gw is not None else None, 3),
                previous.get("time"),
            ])
        add_table(rows, [62 * mm, 48 * mm, 25 * mm, 55 * mm, 25 * mm])

        add_section("HVDC Details")
        add_table(
            [["Link", "Value (MW)", "Time", "Region"]] + [
                [item.get("label"), value(item.get("value_mw")), item.get("time"), item.get("region")]
                for item in (results.get("hvdc") or {}).values()
            ],
            [85 * mm, 40 * mm, 30 * mm, 30 * mm],
        )
        add_section("International Exchange During Yesterday (Import +ve / Export -ve)")
        exchange = psp.get("international_exchange") or {}
        add_table(
            [["Entity", "Schedule (MU)", "Actual (MU)"]] + [
                [name.title(), value((exchange.get(name) or {}).get("schedule_mu"), 3), value((exchange.get(name) or {}).get("actual_mu"), 3)]
                for name in ("BHUTAN", "NEPAL ISTS", "BANGLADESH", "NEPAL BIHAR")
            ],
            [65 * mm, 48 * mm, 48 * mm],
        )
    else:
        results = doc.get("results") or {}
        metrics = results.get("demand_frequency") or {}
        add_section("Demand and Frequency Summary")
        add_table([
            ["Parameter", "MW / Hz", "Time (Hrs)"],
            ["Max demand met", value(metrics.get("max_demand_mw"), 0), metrics.get("max_demand_time")],
            ["Min demand met", value(metrics.get("min_demand_mw"), 0), metrics.get("min_demand_time")],
            ["Max frequency", value(metrics.get("max_frequency_hz"), 3), metrics.get("max_frequency_time")],
            ["Min frequency", value(metrics.get("min_frequency_hz"), 3), metrics.get("min_frequency_time")],
        ], [65 * mm, 45 * mm, 40 * mm])
        frequency = results.get("frequency_distribution") or {}
        add_table([
            ["Frequency", "> 50.05 Hz", "Within band", "< 49.9 Hz"],
            ["% of time", value(frequency.get("above_50_05_pct")), value(frequency.get("within_band_pct")), value(frequency.get("below_49_9_pct"))],
        ], [50 * mm, 38 * mm, 38 * mm, 38 * mm])

        thermal = doc.get("thermal_availability") or {}
        add_section("Thermal Generation Availability Change (00:00 hrs to 17:00 hrs)")
        add_table([
            ["Revived capacity (MW)", "Outage capacity (MW)", "Net capacity change (MW)"],
            [value(thermal.get("revived_capacity_mw"), 0), value(thermal.get("outage_capacity_mw"), 0), value(thermal.get("net_capacity_change_mw"), 0)],
            [thermal.get("revived_details") or "NIL", thermal.get("outage_details") or "NIL", ""],
        ], [72 * mm, 100 * mm, 55 * mm])

        add_section("TTC / ATC, Schedule and Actual")
        states = results.get("states") or {}
        rows = [["State", "TTC", "ATC", "Max schedule", "Time", "Max actual", "Time", "Violation"]]
        for state in STATES:
            item = states.get(state) or {}
            rows.append([
                state, value(item.get("ttc_limit_mw"), 0), value(item.get("atc_limit_mw"), 0),
                value(item.get("max_schedule_mw"), 0), item.get("max_schedule_time"),
                value(item.get("max_actual_mw"), 0), item.get("max_actual_time"),
                value(item.get("atc_violation_mw"), 0),
            ])
        add_table(rows, [30 * mm, 25 * mm, 25 * mm, 34 * mm, 24 * mm, 34 * mm, 24 * mm, 30 * mm])
        add_section("Major OD / UD")
        add_table([["OD statement", doc.get("major_od_text") or "NIL"], ["UD statement", doc.get("major_ud_text") or "NIL"]], [45 * mm, 220 * mm])

    add_section("Important Events (FTC / GD / GI / Load crash etc.)")
    add_table([["Details"], [doc.get("important_events") or "NIL"]], [265 * mm])
    story.extend([
        Spacer(1, 3),
        Paragraph(doc.get("signoff_regards") or "Regards", body),
        Paragraph(doc.get("signoff_name") or "Ashoke Kumar Basak, SIC ERLDC", body),
    ])
    pdf.build(story)
    output.seek(0)
    filename = f"DSO_{report_type.title()}_{report_date}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{report_type}/{report_date}/pdf")
async def download_report_pdf(report_type: str, report_date: str):
    if report_type not in {"morning", "evening"}:
        raise HTTPException(400, "Report type must be morning or evening.")
    doc = collection().find_one(
        {"doc_type": "processed_report", "report_type": report_type, "report_date": report_date},
        {"_id": 0},
    )
    if not doc:
        raise HTTPException(404, "Processed report was not found.")
    return report_pdf_response(doc, report_type, report_date)
