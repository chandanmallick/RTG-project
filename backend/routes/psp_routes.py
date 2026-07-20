from fastapi import APIRouter, BackgroundTasks
from pydantic import BaseModel
from datetime import date, timedelta, datetime
import io
from services.db_handler import MongoService
from services.psp_service import PSPService
from services.pipeline_config_service import PipelineConfigService
import traceback
import pandas as pd
import requests
import os
import threading
import queue
import subprocess
import sys
import json
import re
from collections import Counter, defaultdict
from services.psp_service import LegacySSLAdapter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string
from fastapi.responses import StreamingResponse
from routes.old_logbook_routes import COLLECTION_CONFIG, OLD_LOGBOOK_DB, clean_text, combine_fields, parse_logbook_date, to_jsonable

router = APIRouter(
    prefix="/api/psp",
    tags=["PSP"]
)

PSP_PEAK_BASELINES = {
    "BIHAR": {
        "max_demand": 8822.0, "max_demand_date": "2026-05-24", "max_demand_time": "22:01",
        "max_energy": 186.78, "max_energy_date": "2025-07-24"
    },
    "DVC": {
        "max_demand": 3674.0, "max_demand_date": "2024-06-14", "max_demand_time": "19:30",
        "max_energy": 81.22, "max_energy_date": "2022-04-22"
    },
    "JHARKHAND": {
        "max_demand": 2609.0, "max_demand_date": "2026-05-26", "max_demand_time": "20:00",
        "max_energy": 55.85, "max_energy_date": "2026-05-20"
    },
    "ODISHA": {
        "max_demand": 8482.0, "max_demand_date": "2026-05-22", "max_demand_time": "14:39",
        "max_energy": 162.51, "max_energy_date": "2026-05-22"
    },
    "SIKKIM": {
        "max_demand": 137.0, "max_demand_date": "2024-01-11", "max_demand_time": "18:59",
        "max_energy": 2.50, "max_energy_date": "2020-01-28"
    },
    "WEST BENGAL": {
        "max_demand": 13398.0, "max_demand_date": "2026-05-22", "max_demand_time": "23:26",
        "max_energy": 279.38, "max_energy_date": "2026-05-22"
    },
    "ER": {
        "max_demand": 34875.0, "max_demand_date": "2026-05-22", "max_demand_time": "23:23",
        "max_energy": 750.69, "max_energy_date": "2026-05-22"
    }
}

PSP_PEAK_STATES = [
    {"name": "BIHAR", "hist_prefix": "BIHAR", "daily_demand_name": "BIHAR", "daily_energy_name": "BIHAR"},
    {"name": "DVC", "hist_prefix": "DVC", "daily_demand_name": "DVC", "daily_energy_name": "DVC"},
    {"name": "JHARKHAND", "hist_prefix": "JHARKHAND", "daily_demand_name": "JHARKHAND", "daily_energy_name": "JHARKHAND"},
    {"name": "ODISHA", "hist_prefix": "ORISSA", "daily_demand_name": "ORISSA", "daily_energy_name": "ODISHA"},
    {"name": "SIKKIM", "hist_prefix": "SIKKIM", "daily_demand_name": "SIKKIM", "daily_energy_name": "SIKKIM"},
    {"name": "WEST BENGAL", "hist_prefix": "WEST_BENGAL", "daily_demand_name": "WEST BENGAL", "daily_energy_name": "WEST BENGAL"},
    {"name": "ER", "hist_prefix": "ER", "daily_demand_name": "REGION", "daily_energy_name": "ER"}
]

PSP_PORTFOLIO_SOURCE_VERSION = "wbes_mapping_no_mock_v4"

PSP_PORTFOLIO_STATE_CONFIGS = [
    {"name": "BIHAR", "psp": "BIHAR", "historical_prefix": "BIHAR", "wbes": "BIHAR_STATE",
     "scada": {"thermal": "BH_THERMAL", "hydro": "BIHAR HYDRO", "solar": "BIHAR SOLAR", "nuclear": None, "others": None},
     "scada_gen": "Bihar Total", "highlight": "Bihar", "curve_column": "V", "curve_header": "bseb",
     "mock_p": {"thermal": 0.42, "hydro": 0.05, "solar": 0.02, "isgs": 0.35, "px": 0.12}},
    {"name": "JHARKHAND", "psp": "JHARKHAND", "historical_prefix": "JHARKHAND", "wbes": "JHARKHAND_STATE",
     "scada": {"thermal": "JH_THERMAL", "hydro": "JSEB_HYDRO", "solar": None, "nuclear": None, "others": "JESB CPP+OTHERS"},
     "scada_gen": "JSEB_TOTAL", "highlight": "Jharkhand", "curve_column": "W", "curve_header": "jseb",
     "mock_p": {"thermal": 0.38, "hydro": 0.12, "solar": 0.0, "isgs": 0.30, "px": 0.10}},
    {"name": "DVC", "psp": "DVC", "historical_prefix": "DVC", "wbes": "DVC_STATE",
     "scada": {"thermal": "DVC_THERMAL", "hydro": "DVC _HYDRO", "solar": None, "nuclear": None, "others": "DVC CPP + OTHERS"},
     "scada_gen": "DVC _Total", "highlight": "DVC", "curve_column": "X", "curve_header": "dvc",
     "mock_p": {"thermal": 0.75, "hydro": 0.05, "solar": 0.0, "isgs": 0.10, "px": 0.05}},
    {"name": "ODISHA", "psp": "ORISSA", "historical_prefix": "ORISSA", "wbes": "ODISHA_STATE",
     "scada": {"thermal": "ODISHA_THERMAL", "hydro": "Odisha Hydro", "solar": "GRIDCO SOLAR", "nuclear": None, "others": "odisha cpp (CPPGR)"},
     "scada_gen": "Odisha_Total", "highlight": "Odisha", "curve_column": "Y", "curve_header": "gridco",
     "mock_p": {"thermal": 0.45, "hydro": 0.18, "solar": 0.03, "isgs": 0.22, "px": 0.08}},
    {"name": "WEST BENGAL", "psp": "WEST BENGAL", "historical_prefix": "WEST_BENGAL", "wbes": "WB_STATE",
     "scada": {"thermal": "WEST_BENGAL", "hydro": "WB_Hydro", "solar": "WB_RE_GEN", "nuclear": None, "others": "wb cpp (CPPWB)"},
     "scada_gen": "WB_Total", "highlight": "W. Bengal", "curve_column": "Z", "curve_header": "wbseb",
     "mock_p": {"thermal": 0.58, "hydro": 0.08, "solar": 0.02, "isgs": 0.20, "px": 0.08}},
    {"name": "SIKKIM", "psp": "SIKKIM", "historical_prefix": "SIKKIM", "wbes": "SIKKIM_STATE",
     "scada": {"thermal": None, "hydro": None, "solar": None, "nuclear": None, "others": None},
     "scada_gen": None, "highlight": "Sikkim", "curve_column": "AA", "curve_header": "sikkim",
     "mock_p": {"thermal": 0.0, "hydro": 0.0, "solar": 0.0, "isgs": 0.85, "px": 0.10}}
]

class DateRangeRequest(BaseModel):
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD

class PSPConfigRequest(BaseModel):
    psp_username: str
    psp_password: str
    psp_login_url: str
    psp_data_url: str
    nldc_demand_api_url: str = "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetMaxDemandMetTimeDataByDate/{date_text}"
    india_15_min_demand_api_url: str = "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1"
    all_state_demand_api_url: str = "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetPowerSupplyPositionStatesDataByDate/{date_text}"
    loadshed_api_url: str = "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1"
    outage_api_url: str = "https://report.erldc.in/POSOCO_API/api/Outage/GetQueryNpmcReportData/{date}"
    wbes_url: str = "https://gateway.grid-india.in/POSOCO/reports/1.0/WebAccessAPI/GetUtilityExternalSharedData"
    wbes_api_key: str = "6dc32d47-f46a-45c9-afaf-c6ce73c4ca71"
    wbes_username: str = "erldc_internal_prod"
    wbes_password: str = "ErldcPr0d@052024"
    curve_file_dir: str = r"\\10.3.95.200\HTTP-Access\Control_Room_Report\curve"
    curve_sheet_name: str = "30SEC"
    curve_time_column: str = "C"
    curve_state_columns: str = "V:AA"
    curve_er_column: str = "AE"
    curve_peak_hour_by_month: str = "1:19:00,2:19:00,3:20:00,4:20:00,5:20:00,6:20:00,7:20:00,8:20:00,9:20:00,10:19:00,11:19:00,12:19:00"
    curve_off_peak_hour: str = "03:00"

class PSPPortfolioMappingRow(BaseModel):
    name: str
    psp: str = ""
    wbes: str = ""
    scada_thermal: str = ""
    scada_hydro: str = ""
    scada_solar: str = ""
    scada_others: str = ""
    scada_nuclear: str = ""
    scada_gen: str = ""
    highlight: str = ""
    curve_column: str = ""
    curve_header: str = ""

class PSPPowerSystemBaseRow(BaseModel):
    state: str
    ists_inlet_points: float = 0.0
    per_capita_consumption: float = 0.0
    state_gna: float = 0.0

class PSPPowerSystemBaseRequest(BaseModel):
    effective_date: str
    rows: list[PSPPowerSystemBaseRow]

class DiurnalCurveRange(BaseModel):
    label: str = ""
    start_date: str
    end_date: str
    curve_type: str = "daily"
    selected_dates: list[str] = []

class DiurnalCurveRequest(BaseModel):
    states: list[str] = []
    date_ranges: list[DiurnalCurveRange] = []
    block_minutes: int = 15

class SnapshotOutputRange(BaseModel):
    label: str = ""
    start_date: str
    end_date: str

class SnapshotOutputRequest(BaseModel):
    ranges: list[SnapshotOutputRange] = []

class MisVoltageProfileRequest(BaseModel):
    master_point: str = ""
    start_date: str
    end_date: str
    station_names: list[str] = []
    time: int = 5
    voltage_bus: str = "voltageBus1"

class MisReactorSwitchingRequest(BaseModel):
    start_date: str
    end_date: str
    stations: list[str] = []

class MisOutageAnalysisRequest(BaseModel):
    start_date: str
    end_date: str
    element_names: list[str] = []
    entity_names: list[str] = []
    outage_types: list[str] = []
    excluded_outage_types: list[str] = []
    requesting_entities: list[str] = []
    owners: list[str] = []
    reason_query: str = ""

class MisPlannedOutageRequest(BaseModel):
    element_type: str = "GENERATING_UNIT"
    unit_name: str = ""
    planned_outage_date: str = ""
    planned_outage_from_date: str = ""
    planned_outage_to_date: str = ""
    reason: str = ""
    remarks: str = ""

MIS_OUTAGE_ELEMENT_TYPES = [
    "AC_TRANSMISSION_LINE_CIRCUIT",
    "TRANSFORMER",
    "BUS_REACTOR",
    "LINE_REACTOR",
    "BUS",
    "BAY",
    "GENERATING_UNIT",
    "AUTO_RECLOSER",
    "HVDC_POLE",
    "STATCOM",
]

def update_sync_progress(total: int, completed: int, current_date: str, status: str, error_msg: str = None):
    try:
        db = MongoService()
        db.db["psp_sync_status"].update_one(
            {"_id": "current_run"},
            {"$set": {
                "total": total,
                "completed": completed,
                "current_date": current_date,
                "status": status,
                "error": error_msg,
                "timestamp": datetime.utcnow().isoformat()
            }},
            upsert=True
        )
    except Exception as e:
        print(f"Error updating progress in MongoDB: {str(e)}")

def update_nldc_demand_progress(total: int, completed: int, current_date: str, status: str, error_msg: str = None):
    try:
        db = MongoService()
        db.db["nldc_psp_demand_sync_status"].update_one(
            {"_id": "current_run"},
            {"$set": {
                "total": total,
                "completed": completed,
                "current_date": current_date,
                "status": status,
                "error": error_msg,
                "timestamp": datetime.utcnow().isoformat()
            }},
            upsert=True
        )
    except Exception as e:
        print(f"Error updating NLDC PSP demand progress in MongoDB: {str(e)}")

def update_india_15_min_demand_progress(total: int, completed: int, current_date: str, status: str, error_msg: str = None):
    try:
        db = MongoService()
        db.db["india_15_min_demand_sync_status"].update_one(
            {"_id": "current_run"},
            {"$set": {
                "total": total,
                "completed": completed,
                "current_date": current_date,
                "status": status,
                "error": error_msg,
                "timestamp": datetime.utcnow().isoformat()
            }},
            upsert=True
        )
    except Exception as e:
        print(f"Error updating India 15 Min demand progress in MongoDB: {str(e)}")

def update_all_state_demand_progress(total: int, completed: int, current_date: str, status: str, error_msg: str = None):
    try:
        db = MongoService()
        db.db["all_state_demand_sync_status"].update_one(
            {"_id": "current_run"},
            {"$set": {
                "total": total,
                "completed": completed,
                "current_date": current_date,
                "status": status,
                "error": error_msg,
                "timestamp": datetime.utcnow().isoformat()
            }},
            upsert=True
        )
    except Exception as e:
        print(f"Error updating All State demand progress in MongoDB: {str(e)}")

def run_psp_range_task(start_dt: date, end_dt: date):
    """Background task: login once, then fetch all dates in range."""
    print(f"[PSP] Started sync range: {start_dt} to {end_dt}")
    total_days = (end_dt - start_dt).days + 1
    update_sync_progress(total_days, 0, start_dt.strftime("%Y-%m-%d"), "RUNNING")

    completed = 0
    curr_dt = start_dt
    session = None
    try:
        session = PSPService._create_session()
    except Exception as e:
        print(f"[PSP] Login failed for range task: {e}")

    while curr_dt <= end_dt:
        update_sync_progress(total_days, completed, curr_dt.strftime("%Y-%m-%d"), "RUNNING")
        try:
            PSPService.fetch_and_save_date(curr_dt, session=session)
            try:
                check_and_update_highest_portfolio(curr_dt)
            except Exception as ex:
                print(f"Error updating highest for {curr_dt}: {ex}")
            completed += 1
        except Exception as e:
            print(f"[PSP] Failed for {curr_dt}: {e}")
            completed += 1
        curr_dt += timedelta(days=1)

    update_sync_progress(total_days, completed, "", "COMPLETED")
    print(f"[PSP] Sync range completed. {completed} days processed.")

def run_nldc_demand_range_task(start_dt: date, end_dt: date):
    print(f"[NLDC PSP Demand] Started sync range: {start_dt} to {end_dt}")
    total_days = (end_dt - start_dt).days + 1
    update_nldc_demand_progress(
        total_days,
        0,
        start_dt.strftime("%Y-%m-%d"),
        "RUNNING"
    )

    completed = 0
    curr_dt = start_dt
    last_error = None

    while curr_dt <= end_dt:
        date_text = curr_dt.strftime("%Y-%m-%d")
        update_nldc_demand_progress(
            total_days,
            completed,
            date_text,
            "RUNNING"
        )

        try:
            PSPService.fetch_and_save_nldc_demand_date(curr_dt)
        except Exception as e:
            last_error = f"{date_text}: {str(e)}"
            print(f"[NLDC PSP Demand] Failed for {date_text}: {e}")
        finally:
            completed += 1
            curr_dt += timedelta(days=1)

    update_nldc_demand_progress(
        total_days,
        completed,
        "",
        "COMPLETED",
        last_error
    )
    print(f"[NLDC PSP Demand] Sync range completed. {completed} days processed.")

def run_india_15_min_demand_range_task(start_dt: date, end_dt: date):
    print(f"[India 15 Min Demand] Started sync range: {start_dt} to {end_dt}")
    total_days = (end_dt - start_dt).days + 1
    update_india_15_min_demand_progress(
        total_days,
        0,
        start_dt.strftime("%Y-%m-%d"),
        "RUNNING"
    )

    completed = 0
    curr_dt = start_dt
    last_error = None

    while curr_dt <= end_dt:
        date_text = curr_dt.strftime("%Y-%m-%d")
        update_india_15_min_demand_progress(
            total_days,
            completed,
            date_text,
            "RUNNING"
        )

        try:
            PSPService.fetch_and_save_india_15_min_demand_date(curr_dt)
        except Exception as e:
            last_error = f"{date_text}: {str(e)}"
            print(f"[India 15 Min Demand] Failed for {date_text}: {e}")
        finally:
            completed += 1
            curr_dt += timedelta(days=1)

    update_india_15_min_demand_progress(
        total_days,
        completed,
        "",
        "COMPLETED",
        last_error
    )
    print(f"[India 15 Min Demand] Sync range completed. {completed} days processed.")

def run_all_state_demand_range_task(start_dt: date, end_dt: date):
    print(f"[All State Demand] Started sync range: {start_dt} to {end_dt}")
    total_days = (end_dt - start_dt).days + 1
    update_all_state_demand_progress(
        total_days,
        0,
        start_dt.strftime("%Y-%m-%d"),
        "RUNNING"
    )

    completed = 0
    curr_dt = start_dt
    last_error = None

    while curr_dt <= end_dt:
        date_text = curr_dt.strftime("%Y-%m-%d")
        update_all_state_demand_progress(
            total_days,
            completed,
            date_text,
            "RUNNING"
        )

        try:
            PSPService.fetch_and_save_all_state_demand_date(curr_dt)
        except Exception as e:
            last_error = f"{date_text}: {str(e)}"
            print(f"[All State Demand] Failed for {date_text}: {e}")
        finally:
            completed += 1
            curr_dt += timedelta(days=1)

    update_all_state_demand_progress(
        total_days,
        completed,
        "",
        "COMPLETED",
        last_error
    )
    print(f"[All State Demand] Sync range completed. {completed} days processed.")

@router.get("/status")
async def get_psp_status(start_date: str = None, end_date: str = None):
    db = MongoService()
    
    if start_date and end_date:
        try:
            start_dt = date.fromisoformat(start_date)
            end_dt = date.fromisoformat(end_date)
            if start_dt > end_dt:
                return {"success": False, "message": "Start date must be less than or equal to end date."}
            total_days = (end_dt - start_dt).days + 1
            dates = [start_dt + timedelta(days=i) for i in range(total_days)]
            # Reverse to show most recent first
            dates.reverse()
        except ValueError:
            return {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}
    else:
        # Generate dates for the last 30 days starting from yesterday
        yesterday = date.today() - timedelta(days=1)
        dates = [yesterday - timedelta(days=i) for i in range(30)]
        
    date_strings = [d.strftime("%Y-%m-%d") for d in dates]
    
    # Query database for existing documents in this range
    existing_docs = list(db.psp_collection.find(
        {"date": {"$in": date_strings}},
        {"date": 1, "fetched_at": 1, "_id": 0}
    ))
    
    # Map date to fetched timestamp
    status_map = {doc["date"]: doc.get("fetched_at") for doc in existing_docs}
    
    status_list = []
    for d in dates:
        d_str = d.strftime("%Y-%m-%d")
        fetched_at = status_map.get(d_str)
        status_list.append({
            "date": d_str,
            "status": "SUCCESS" if fetched_at else "MISSING",
            "fetched_at": fetched_at
        })
        
    return {
        "success": True,
        "data": status_list
    }

@router.get("/sync-progress")
async def get_sync_progress():
    try:
        db = MongoService()
        progress = db.db["psp_sync_status"].find_one({"_id": "current_run"})
        if not progress:
            return {"success": True, "status": "IDLE"}
        return {"success": True, **progress}
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.post("/run-range")
async def run_psp_range(req: DateRangeRequest, background_tasks: BackgroundTasks):
    try:
        start_dt = date.fromisoformat(req.start_date)
        end_dt = date.fromisoformat(req.end_date)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }
        
    if start_dt > end_dt:
        return {
            "success": False,
            "message": "Start date must be less than or equal to end date."
        }
    
    # Reset/trigger the background task
    background_tasks.add_task(run_psp_range_task, start_dt, end_dt)
    
    return {
        "success": True,
        "message": f"Sync task triggered in background from {req.start_date} to {req.end_date}."
    }

@router.post("/sync-date/{date_str}")
async def sync_single_date(date_str: str):
    try:
        dt = date.fromisoformat(date_str)
        res = PSPService.fetch_and_save_date(dt)
        try:
            check_and_update_highest_portfolio(dt)
        except Exception as ex:
            print(f"Error checking highest records for single date sync: {ex}")
        return res
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }
    except Exception as e:
        return {
            "success": False,
            "message": str(e)
        }

@router.get("/nldc-demand/status")
async def get_nldc_demand_status(start_date: str = None, end_date: str = None):
    db = MongoService()

    if start_date and end_date:
        try:
            start_dt = date.fromisoformat(start_date)
            end_dt = date.fromisoformat(end_date)
            if start_dt > end_dt:
                return {
                    "success": False,
                    "message": "Start date must be less than or equal to end date."
                }
            total_days = (end_dt - start_dt).days + 1
            dates = [
                start_dt + timedelta(days=i)
                for i in range(total_days)
            ]
            dates.reverse()
        except ValueError:
            return {
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD."
            }
    else:
        yesterday = date.today() - timedelta(days=1)
        dates = [
            yesterday - timedelta(days=i)
            for i in range(30)
        ]

    date_strings = [
        d.strftime("%Y-%m-%d")
        for d in dates
    ]

    existing_docs = list(
        db.nldc_psp_demand_collection.find(
            {"date": {"$in": date_strings}},
            {
                "date": 1,
                "fetched_at": 1,
                "record_count": 1,
                "_id": 0
            }
        )
    )

    status_map = {
        doc["date"]: doc
        for doc in existing_docs
    }

    return {
        "success": True,
        "data": [
            {
                "date": d.strftime("%Y-%m-%d"),
                "status": "SUCCESS" if d.strftime("%Y-%m-%d") in status_map else "MISSING",
                "fetched_at": status_map.get(d.strftime("%Y-%m-%d"), {}).get("fetched_at"),
                "record_count": status_map.get(d.strftime("%Y-%m-%d"), {}).get("record_count", 0)
            }
            for d in dates
        ]
    }

@router.get("/nldc-demand/sync-progress")
async def get_nldc_demand_sync_progress():
    try:
        db = MongoService()
        progress = db.db["nldc_psp_demand_sync_status"].find_one(
            {"_id": "current_run"}
        )
        if not progress:
            return {"success": True, "status": "IDLE"}
        return {"success": True, **progress}
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.post("/nldc-demand/run-range")
async def run_nldc_demand_range(req: DateRangeRequest, background_tasks: BackgroundTasks):
    try:
        start_dt = date.fromisoformat(req.start_date)
        end_dt = date.fromisoformat(req.end_date)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }

    if start_dt > end_dt:
        return {
            "success": False,
            "message": "Start date must be less than or equal to end date."
        }

    background_tasks.add_task(
        run_nldc_demand_range_task,
        start_dt,
        end_dt
    )

    return {
        "success": True,
        "message": f"NLDC PSP demand sync started from {req.start_date} to {req.end_date}."
    }

@router.post("/nldc-demand/sync-date/{date_str}")
async def sync_nldc_demand_date(date_str: str):
    try:
        dt = date.fromisoformat(date_str)
        return PSPService.fetch_and_save_nldc_demand_date(dt)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }
    except Exception as e:
        traceback.print_exc()
        return {
            "success": False,
            "message": str(e)
        }

@router.get("/nldc-demand/trend")
async def get_nldc_demand_trend(start_date: str = None, end_date: str = None):
    try:
        if end_date:
            end_dt = date.fromisoformat(end_date)
        else:
            end_dt = date.today() - timedelta(days=1)

        if start_date:
            start_dt = date.fromisoformat(start_date)
        else:
            start_dt = end_dt - timedelta(days=30)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Use YYYY-MM-DD."
        }

    if start_dt > end_dt:
        return {
            "success": False,
            "message": "Start date must be less than or equal to end date."
        }

    region_keys = ["ALL INDIA", "NR", "WR", "SR", "NER", "ER"]
    region_labels = {
        "ALL INDIA": "All India",
        "NR": "NR",
        "WR": "WR",
        "SR": "SR",
        "NER": "NER",
        "ER": "ER"
    }

    docs = list(
        MongoService()
        .nldc_psp_demand_collection
        .find(
            {
                "date": {
                    "$gte": start_dt.isoformat(),
                    "$lte": end_dt.isoformat()
                }
            },
            {
                "_id": 0,
                "date": 1,
                "regions": 1
            }
        )
        .sort("date", 1)
    )

    yearly = (end_dt - start_dt).days > 366
    demand_metrics = {
        "overall": {
            "field": "max_demand_met",
            "time_field": "max_demand_met_time",
            "suffix": ""
        },
        "solar": {
            "field": "max_demand_met_solar",
            "time_field": "max_demand_met_solar_time",
            "suffix": "__solar"
        },
        "non_solar": {
            "field": "max_demand_met_non_solar",
            "time_field": "max_demand_met_non_solar_time",
            "suffix": "__non_solar"
        }
    }
    max_by_region = {
        metric_key: {}
        for metric_key in demand_metrics
    }

    def get_region_value(doc, region_key, metric_key="overall"):
        row = (doc.get("regions") or {}).get(region_key) or {}
        metric = demand_metrics.get(metric_key) or demand_metrics["overall"]
        value = row.get(metric["field"])
        try:
            value = float(value)
        except (TypeError, ValueError):
            return None, row
        return value, row

    def output_key(region_key, metric_key):
        return f"{region_key}{demand_metrics[metric_key]['suffix']}"

    def set_output_value(output, region_key, metric_key, value, date_value, row):
        key = output_key(region_key, metric_key)
        time_field = demand_metrics[metric_key]["time_field"]
        output[key] = value
        output[f"{key}_date"] = date_value
        output[f"{key}_time"] = row.get(time_field) or ""

    yearly_max_by_region = {
        metric_key: {
            region_key: {}
            for region_key in region_keys
        }
        for metric_key in demand_metrics
    }

    rows = []
    for doc in docs:
        output = {
            "period": doc.get("date"),
            "date": doc.get("date")
        }
        year = str(doc.get("date", ""))[:4]
        for region_key in region_keys:
            for metric_key in demand_metrics:
                value, row = get_region_value(doc, region_key, metric_key)
                if value is None:
                    continue
                set_output_value(output, region_key, metric_key, value, doc.get("date"), row)
                max_item = {
                    "value": value,
                    "date": doc.get("date"),
                    "time": row.get(demand_metrics[metric_key]["time_field"]) or "",
                    "year": year
                }
                existing = max_by_region[metric_key].get(region_key)
                if not existing or value > existing["value"]:
                    max_by_region[metric_key][region_key] = max_item
                yearly_existing = yearly_max_by_region[metric_key][region_key].get(year)
                if year and (not yearly_existing or value > yearly_existing["value"]):
                    yearly_max_by_region[metric_key][region_key][year] = max_item
        rows.append(output)

    values = [
        row.get(output_key(region_key, metric_key))
        for row in rows
        for region_key in region_keys
        for metric_key in demand_metrics
        if row.get(output_key(region_key, metric_key)) is not None
    ]
    all_india_values = [
        row.get(output_key("ALL INDIA", metric_key))
        for row in rows
        for metric_key in demand_metrics
        if row.get(output_key("ALL INDIA", metric_key)) is not None
    ]
    regional_values = [
        row.get(output_key(region_key, metric_key))
        for row in rows
        for region_key in region_keys
        for metric_key in demand_metrics
        if region_key != "ALL INDIA" and row.get(output_key(region_key, metric_key)) is not None
    ]

    use_secondary_axis = False
    if all_india_values and regional_values:
        max_all_india = max(all_india_values)
        max_regional = max(regional_values)
        min_regional = min(value for value in regional_values if value is not None)
        use_secondary_axis = (
            max_regional > 0 and max_all_india / max_regional >= 2.0
        ) or (
            min_regional > 0 and max_all_india / min_regional >= 4.0
        )

    return {
        "success": True,
        "start_date": start_dt.isoformat(),
        "end_date": end_dt.isoformat(),
        "mode": "yearly" if yearly else "daily",
        "regions": [
            {
                "key": key,
                "label": region_labels[key],
                "max": {
                    "value": max_by_region["overall"].get(key, {}).get("value"),
                    "date": max_by_region["overall"].get(key, {}).get("date"),
                    "time": max_by_region["overall"].get(key, {}).get("time")
                },
                "yearly_max": [
                    yearly_max_by_region["overall"].get(key, {}).get(year)
                    for year in sorted(yearly_max_by_region["overall"].get(key, {}))
                ] if yearly else [],
                "max_solar": {
                    "value": max_by_region["solar"].get(key, {}).get("value"),
                    "date": max_by_region["solar"].get(key, {}).get("date"),
                    "time": max_by_region["solar"].get(key, {}).get("time")
                },
                "yearly_max_solar": [
                    yearly_max_by_region["solar"].get(key, {}).get(year)
                    for year in sorted(yearly_max_by_region["solar"].get(key, {}))
                ] if yearly else [],
                "max_non_solar": {
                    "value": max_by_region["non_solar"].get(key, {}).get("value"),
                    "date": max_by_region["non_solar"].get(key, {}).get("date"),
                    "time": max_by_region["non_solar"].get(key, {}).get("time")
                },
                "yearly_max_non_solar": [
                    yearly_max_by_region["non_solar"].get(key, {}).get(year)
                    for year in sorted(yearly_max_by_region["non_solar"].get(key, {}))
                ] if yearly else []
            }
            for key in region_keys
        ],
        "use_secondary_axis": use_secondary_axis,
        "rows": rows,
        "record_count": len(rows),
        "max_value": max(values) if values else None
    }

@router.get("/india-15-min-demand/status")
async def get_india_15_min_demand_status(start_date: str = None, end_date: str = None):
    db = MongoService()

    if start_date and end_date:
        try:
            start_dt = date.fromisoformat(start_date)
            end_dt = date.fromisoformat(end_date)
            if start_dt > end_dt:
                return {
                    "success": False,
                    "message": "Start date must be less than or equal to end date."
                }
            total_days = (end_dt - start_dt).days + 1
            dates = [
                start_dt + timedelta(days=i)
                for i in range(total_days)
            ]
            dates.reverse()
        except ValueError:
            return {
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD."
            }
    else:
        yesterday = date.today() - timedelta(days=1)
        dates = [
            yesterday - timedelta(days=i)
            for i in range(30)
        ]

    date_strings = [
        d.strftime("%Y-%m-%d")
        for d in dates
    ]

    existing_docs = list(
        db.india_15_min_demand_collection.find(
            {"date": {"$in": date_strings}},
            {
                "date": 1,
                "fetched_at": 1,
                "record_count": 1,
                "_id": 0
            }
        )
    )

    status_map = {
        doc["date"]: doc
        for doc in existing_docs
    }

    return {
        "success": True,
        "data": [
            {
                "date": d.strftime("%Y-%m-%d"),
                "status": "SUCCESS" if d.strftime("%Y-%m-%d") in status_map else "MISSING",
                "fetched_at": status_map.get(d.strftime("%Y-%m-%d"), {}).get("fetched_at"),
                "record_count": status_map.get(d.strftime("%Y-%m-%d"), {}).get("record_count", 0)
            }
            for d in dates
        ]
    }

@router.get("/india-15-min-demand/sync-progress")
async def get_india_15_min_demand_sync_progress():
    try:
        db = MongoService()
        progress = db.db["india_15_min_demand_sync_status"].find_one(
            {"_id": "current_run"}
        )
        if not progress:
            return {"success": True, "status": "IDLE"}
        return {"success": True, **progress}
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.post("/india-15-min-demand/run-range")
async def run_india_15_min_demand_range(req: DateRangeRequest, background_tasks: BackgroundTasks):
    try:
        start_dt = date.fromisoformat(req.start_date)
        end_dt = date.fromisoformat(req.end_date)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }

    if start_dt > end_dt:
        return {
            "success": False,
            "message": "Start date must be less than or equal to end date."
        }

    background_tasks.add_task(
        run_india_15_min_demand_range_task,
        start_dt,
        end_dt
    )

    return {
        "success": True,
        "message": f"India 15 Min demand sync started from {req.start_date} to {req.end_date}."
    }

@router.post("/india-15-min-demand/sync-date/{date_str}")
async def sync_india_15_min_demand_date(date_str: str):
    try:
        dt = date.fromisoformat(date_str)
        return PSPService.fetch_and_save_india_15_min_demand_date(dt)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }
    except Exception as e:
        traceback.print_exc()
        return {
            "success": False,
            "message": str(e)
        }

@router.get("/india-15-min-demand/generation-breakup")
async def get_india_15_min_generation_breakup(date_str: str = None):
    components = [
        ("AI_GAS", "Gas"),
        ("AI_HYD", "Hydro"),
        ("AI_NUC", "Nuclear"),
        ("AI_TH", "Thermal"),
        ("AI_WIND", "Wind"),
        ("AI_SOLAR", "Solar"),
        ("AI_OTHERS", "Others"),
        ("AI_PSP", "Pump"),
        ("AI_BESS", "BESS"),
        ("NET_TRANSNATIONAL_EXCHANGE", "Transnational Exchange"),
    ]

    try:
        db = MongoService()
        query = {}
        if date_str:
            date.fromisoformat(date_str)
            query = {"date": date_str}

        doc = (
            db.india_15_min_demand_collection
            .find_one(query, {"_id": 0}, sort=[("date", -1)])
        )
        if not doc:
            return {
                "success": True,
                "date": date_str,
                "rows": [],
                "components": [
                    {"key": key, "label": label, "total": 0, "max": 0, "share": 0}
                    for key, label in components
                ],
                "record_count": 0,
                "message": "No India 15 Min demand document found."
            }

        demand_doc = db.nldc_psp_demand_collection.find_one(
            {"date": doc.get("date")},
            {"_id": 0, "regions.ALL INDIA": 1},
        ) or {}
        demand_row = ((demand_doc.get("regions") or {}).get("ALL INDIA") or {})
        all_india_demand = {
            "value": demand_row.get("max_demand_met"),
            "time": demand_row.get("max_demand_met_time") or "",
            "solar_period_value": demand_row.get("max_demand_met_solar"),
            "solar_period_time": demand_row.get("max_demand_met_solar_time") or "",
            "non_solar_period_value": demand_row.get("max_demand_met_non_solar"),
            "non_solar_period_time": demand_row.get("max_demand_met_non_solar_time") or "",
        }

        def as_float(value):
            try:
                return float(value or 0)
            except (TypeError, ValueError):
                return 0.0

        def parse_time_of_day(value):
            text = str(value or "").strip()
            if not text:
                return 9999, ""
            parts = text.split(":")
            try:
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
            except (TypeError, ValueError):
                return 9999, text
            if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                return 9999, text
            return hour * 60 + minute, f"{hour:02d}:{minute:02d}"

        raw_rows = doc.get("raw_rows") or []
        rows = []
        totals = {
            key: 0.0
            for key, _ in components
        }
        max_values = {
            key: 0.0
            for key, _ in components
        }

        for index, row in enumerate(raw_rows):
            timestamp = row.get("TIMESTAMP") or row.get("timestamp") or ""
            time_minutes, time_label = parse_time_of_day(timestamp)
            item = {
                "block": index + 1,
                "timestamp": time_label,
                "raw_timestamp": timestamp,
                "time_minutes": time_minutes,
                "date_key": row.get("DATE_KEY") or row.get("date_key"),
            }
            total_generation = 0.0
            for key, _ in components:
                value = as_float(row.get(key))
                item[key] = round(value, 3)
                totals[key] += value
                max_values[key] = max(max_values[key], value)
                total_generation += value
            item["total_generation"] = round(total_generation, 3)
            rows.append(item)

        rows.sort(key=lambda item: (item.get("time_minutes", 9999), item.get("block", 9999)))
        for index, item in enumerate(rows):
            item["block"] = index + 1
            item["solar_generation"] = round(item.get("AI_SOLAR") or 0, 3)
            item["non_solar_generation"] = round(
                (item.get("total_generation") or 0) - item["solar_generation"], 3
            )

        total_energy = sum(abs(value) for value in totals.values())
        component_summary = [
            {
                "key": key,
                "label": label,
                "total": round(totals[key], 3),
                "max": round(max_values[key], 3),
                "share": round((abs(totals[key]) / total_energy) * 100, 3) if total_energy else 0
            }
            for key, label in components
        ]

        def nearest_generation_row(time_value):
            target_minutes, _ = parse_time_of_day(time_value)
            if target_minutes == 9999 or not rows:
                return {}
            return min(rows, key=lambda row: abs((row.get("time_minutes") or 0) - target_minutes))

        def build_generation_mix(mode, label, generation_row, demand_value=None, demand_time=""):
            total = float(generation_row.get("total_generation") or 0)
            solar = float(generation_row.get("solar_generation") or 0)
            non_solar = float(generation_row.get("non_solar_generation") or 0)
            return {
                "mode": mode,
                "label": label,
                "block": generation_row.get("block"),
                "timestamp": generation_row.get("timestamp") or "",
                "total_generation": round(total, 3),
                "solar_generation": round(solar, 3),
                "non_solar_generation": round(non_solar, 3),
                "solar_share": round((solar / total) * 100, 3) if total else 0,
                "non_solar_share": round((non_solar / total) * 100, 3) if total else 0,
                "demand_value": as_float(demand_value) if demand_value not in [None, ""] else None,
                "demand_time": demand_time or "",
            }

        peak_row = max(rows, key=lambda row: row.get("total_generation") or 0, default={})
        generation_mix_modes = {
            "peak_generation": build_generation_mix(
                "peak_generation", "Peak Generation", peak_row,
                all_india_demand.get("value"), all_india_demand.get("time"),
            ),
            "solar_max_demand": build_generation_mix(
                "solar_max_demand", "Solar-period Maximum Demand",
                nearest_generation_row(all_india_demand.get("solar_period_time")),
                all_india_demand.get("solar_period_value"), all_india_demand.get("solar_period_time"),
            ),
            "non_solar_max_demand": build_generation_mix(
                "non_solar_max_demand", "Non-solar-period Maximum Demand",
                nearest_generation_row(all_india_demand.get("non_solar_period_time")),
                all_india_demand.get("non_solar_period_value"), all_india_demand.get("non_solar_period_time"),
            ),
        }
        maximum_generation_mix = generation_mix_modes["peak_generation"]

        return {
            "success": True,
            "date": doc.get("date"),
            "date_text": doc.get("date_text"),
            "source": "India_15_Min_Demand.raw_rows",
            "rows": rows,
            "components": component_summary,
            "record_count": len(rows),
            "all_india_demand": all_india_demand,
            "maximum_generation_mix": maximum_generation_mix,
            "generation_mix_modes": generation_mix_modes,
            "max_total_generation": max((row.get("total_generation") or 0 for row in rows), default=0),
            "min_total_generation": min((row.get("total_generation") or 0 for row in rows), default=0),
        }
    except ValueError:
        return {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "message": str(e)}


@router.get("/india-15-min-demand/generation-breakup/export")
async def export_india_15_min_generation_breakup(date_str: str = None):
    result = await get_india_15_min_generation_breakup(date_str)
    if not result.get("success"):
        return result

    rows = result.get("rows") or []
    components = result.get("components") or []
    mix = result.get("maximum_generation_mix") or {}
    mix_modes = result.get("generation_mix_modes") or {}
    demand = result.get("all_india_demand") or {}
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Maximum Generation Mix"

    title_fill = PatternFill("solid", fgColor="022726")
    header_fill = PatternFill("solid", fgColor="DDEFEA")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="03624C", bold=True)

    summary.merge_cells("A1:D1")
    summary["A1"] = "All India Demand Contribution"
    summary["A1"].fill = title_fill
    summary["A1"].font = title_font
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.append(["Date", result.get("date") or date_str or "", "Source", result.get("source") or ""])
    summary.append(["All India Maximum Demand", demand.get("value") or 0, "Time", demand.get("time") or "-"])
    summary.append([])
    summary.append(["Maximum Generation Mix", "MW", "Share (%)", "Time / Block"])
    summary.append(["Total Generation", mix.get("total_generation", 0), 100, f'{mix.get("timestamp") or "-"} / Block {mix.get("block") or "-"}'])
    summary.append(["Solar", mix.get("solar_generation", 0), mix.get("solar_share", 0), mix.get("timestamp") or "-"])
    summary.append(["Non-Solar", mix.get("non_solar_generation", 0), mix.get("non_solar_share", 0), mix.get("timestamp") or "-"])
    summary.append([])
    summary.append(["Component", "Maximum (MW)", "Daily Share (%)", "Block Total"])
    for component in components:
        summary.append([
            component.get("label") or component.get("key"),
            component.get("max", 0),
            component.get("share", 0),
            component.get("total", 0),
        ])

    for cell in summary[5]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in summary[10]:
        cell.fill = header_fill
        cell.font = header_font
    summary.freeze_panes = "A5"
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 24

    detail = workbook.create_sheet("15-Minute Data")
    component_keys = [component.get("key") for component in components if component.get("key")]
    headers = ["Block", "Timestamp", *component_keys, "Solar Generation", "Non-Solar Generation", "Total Generation"]
    detail.append(headers)
    for cell in detail[1]:
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        detail.append([
            row.get("block"),
            row.get("timestamp"),
            *[row.get(key, 0) for key in component_keys],
            row.get("solar_generation", 0),
            row.get("non_solar_generation", 0),
            row.get("total_generation", 0),
        ])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for column in detail.columns:
        letter = column[0].column_letter
        detail.column_dimensions[letter].width = max(13, min(24, max(len(str(cell.value or "")) for cell in column) + 2))

    demand_mix = workbook.create_sheet("Demand-Based Mix")
    demand_mix.append([
        "Basis", "Demand (MW)", "Demand Time", "Nearest Generation Time", "Block",
        "Total Generation (MW)", "Solar Generation (MW)", "Solar Share (%)",
        "Non-Solar Generation (MW)", "Non-Solar Share (%)",
    ])
    for cell in demand_mix[1]:
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for mode in ["peak_generation", "solar_max_demand", "non_solar_max_demand"]:
        item = mix_modes.get(mode) or {}
        demand_mix.append([
            item.get("label") or mode,
            item.get("demand_value"),
            item.get("demand_time") or "",
            item.get("timestamp") or "",
            item.get("block"),
            item.get("total_generation", 0),
            item.get("solar_generation", 0),
            item.get("solar_share", 0),
            item.get("non_solar_generation", 0),
            item.get("non_solar_share", 0),
        ])
    demand_mix.freeze_panes = "A2"
    demand_mix.auto_filter.ref = demand_mix.dimensions
    for column in demand_mix.columns:
        letter = column[0].column_letter
        demand_mix.column_dimensions[letter].width = max(15, min(28, max(len(str(cell.value or "")) for cell in column) + 2))

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    export_date = result.get("date") or date_str or "latest"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="all_india_demand_contribution_{export_date}.xlsx"'},
    )

@router.get("/all-state-demand/status")
async def get_all_state_demand_status(start_date: str = None, end_date: str = None):
    db = MongoService()

    if start_date and end_date:
        try:
            start_dt = date.fromisoformat(start_date)
            end_dt = date.fromisoformat(end_date)
            if start_dt > end_dt:
                return {
                    "success": False,
                    "message": "Start date must be less than or equal to end date."
                }
            total_days = (end_dt - start_dt).days + 1
            dates = [
                start_dt + timedelta(days=i)
                for i in range(total_days)
            ]
            dates.reverse()
        except ValueError:
            return {
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD."
            }
    else:
        yesterday = date.today() - timedelta(days=1)
        dates = [
            yesterday - timedelta(days=i)
            for i in range(30)
        ]

    date_strings = [
        d.strftime("%Y-%m-%d")
        for d in dates
    ]

    existing_docs = list(
        db.all_state_demand_collection.find(
            {"date": {"$in": date_strings}},
            {
                "date": 1,
                "fetched_at": 1,
                "record_count": 1,
                "_id": 0
            }
        )
    )

    status_map = {
        doc["date"]: doc
        for doc in existing_docs
    }

    return {
        "success": True,
        "data": [
            {
                "date": d.strftime("%Y-%m-%d"),
                "status": "SUCCESS" if d.strftime("%Y-%m-%d") in status_map else "MISSING",
                "fetched_at": status_map.get(d.strftime("%Y-%m-%d"), {}).get("fetched_at"),
                "record_count": status_map.get(d.strftime("%Y-%m-%d"), {}).get("record_count", 0)
            }
            for d in dates
        ]
    }

@router.get("/all-state-demand/sync-progress")
async def get_all_state_demand_sync_progress():
    try:
        db = MongoService()
        progress = db.db["all_state_demand_sync_status"].find_one(
            {"_id": "current_run"}
        )
        if not progress:
            return {"success": True, "status": "IDLE"}
        return {"success": True, **progress}
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.post("/all-state-demand/run-range")
async def run_all_state_demand_range(req: DateRangeRequest, background_tasks: BackgroundTasks):
    try:
        start_dt = date.fromisoformat(req.start_date)
        end_dt = date.fromisoformat(req.end_date)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }

    if start_dt > end_dt:
        return {
            "success": False,
            "message": "Start date must be less than or equal to end date."
        }

    background_tasks.add_task(
        run_all_state_demand_range_task,
        start_dt,
        end_dt
    )

    return {
        "success": True,
        "message": f"All State demand sync started from {req.start_date} to {req.end_date}."
    }

@router.post("/all-state-demand/sync-date/{date_str}")
async def sync_all_state_demand_date(date_str: str):
    try:
        dt = date.fromisoformat(date_str)
        return PSPService.fetch_and_save_all_state_demand_date(dt)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Please use YYYY-MM-DD."
        }
    except Exception as e:
        traceback.print_exc()
        return {
            "success": False,
            "message": str(e)
        }

@router.post("/refresh-sources")
async def refresh_psp_sources(date_str: str = None):
    try:
        target_date = date.fromisoformat(date_str) if date_str else date.today() - timedelta(days=1)
    except ValueError:
        return {"success": False, "message": "Invalid date format. Please use YYYY-MM-DD."}

    try:
        results = refresh_psp_operational_sources(target_date)
        failed = {
            key: value
            for key, value in results.items()
            if not value.get("success")
        }
        return {
            "success": not bool(failed),
            "date": target_date.strftime("%Y-%m-%d"),
            "message": "All PSP source data refreshed and cached." if not failed else "PSP source refresh completed with some failures.",
            "results": results,
            "failed": failed,
        }
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "message": str(e)}

@router.get("/config")
async def get_psp_config():
    config_service = PipelineConfigService()
    config = config_service.get_config("PSP")
    if not config:
        config = {
            "config_type": "PSP",
            "psp_login_url": "https://report.erldc.in/POSOCO/Account/Login",
            "psp_data_url": "https://report.erldc.in/POSOCO/PSP/GetPSPData",
            "psp_username": "erldc",
            "psp_password": "erldc1",
            "nldc_demand_api_url": "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetMaxDemandMetTimeDataByDate/{date_text}",
            "india_15_min_demand_api_url": "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1",
            "all_state_demand_api_url": "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetPowerSupplyPositionStatesDataByDate/{date_text}",
            "loadshed_api_url": "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1",
            "outage_api_url": "https://report.erldc.in/POSOCO_API/api/Outage/GetQueryNpmcReportData/{date}",
            "wbes_url": "https://gateway.grid-india.in/POSOCO/reports/1.0/WebAccessAPI/GetUtilityExternalSharedData",
            "wbes_api_key": "6dc32d47-f46a-45c9-afaf-c6ce73c4ca71",
            "wbes_username": "erldc_internal_prod",
            "wbes_password": "ErldcPr0d@052024",
            "curve_file_dir": r"\\10.3.95.200\HTTP-Access\Control_Room_Report\curve",
            "curve_sheet_name": "30SEC",
            "curve_time_column": "C",
            "curve_state_columns": "V:AA",
            "curve_er_column": "AE",
            "curve_peak_hour_by_month": "1:19:00,2:19:00,3:20:00,4:20:00,5:20:00,6:20:00,7:20:00,8:20:00,9:20:00,10:19:00,11:19:00,12:19:00",
            "curve_off_peak_hour": "03:00"
        }
    else:
        # Populate defaults for missing wbes keys if any
        if "wbes_url" not in config:
            config["wbes_url"] = "https://gateway.grid-india.in/POSOCO/reports/1.0/WebAccessAPI/GetUtilityExternalSharedData"
        if "wbes_api_key" not in config:
            config["wbes_api_key"] = "6dc32d47-f46a-45c9-afaf-c6ce73c4ca71"
        if "wbes_username" not in config:
            config["wbes_username"] = "erldc_internal_prod"
        if "wbes_password" not in config:
            config["wbes_password"] = "ErldcPr0d@052024"
        if "loadshed_api_url" not in config:
            config["loadshed_api_url"] = "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1"
        if "nldc_demand_api_url" not in config:
            config["nldc_demand_api_url"] = "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetMaxDemandMetTimeDataByDate/{date_text}"
        if "india_15_min_demand_api_url" not in config:
            config["india_15_min_demand_api_url"] = config.get("loadshed_api_url") or "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1"
        if "all_state_demand_api_url" not in config:
            config["all_state_demand_api_url"] = "https://reporting.nldc.in/Reporting_API/API/NLDCReport/GetPowerSupplyPositionStatesDataByDate/{date_text}"
        if "outage_api_url" not in config:
            config["outage_api_url"] = "https://report.erldc.in/POSOCO_API/api/Outage/GetQueryNpmcReportData/{date}"
        if "curve_file_dir" not in config:
            config["curve_file_dir"] = r"\\10.3.95.200\HTTP-Access\Control_Room_Report\curve"
        if "curve_sheet_name" not in config:
            config["curve_sheet_name"] = "30SEC"
        if "curve_time_column" not in config:
            config["curve_time_column"] = "C"
        if "curve_state_columns" not in config:
            config["curve_state_columns"] = "V:AA"
        if "curve_er_column" not in config:
            config["curve_er_column"] = "AE"
        if "curve_peak_hour_by_month" not in config:
            config["curve_peak_hour_by_month"] = "1:19:00,2:19:00,3:20:00,4:20:00,5:20:00,6:20:00,7:20:00,8:20:00,9:20:00,10:19:00,11:19:00,12:19:00"
        if "curve_off_peak_hour" not in config:
            config["curve_off_peak_hour"] = "03:00"
    return {
        "success": True,
        "config": config
    }

@router.post("/config")
async def save_psp_config(req: PSPConfigRequest):
    try:
        db = MongoService()
        db.db["pipeline_config"].update_one(
            {"config_type": "PSP"},
            {"$set": {
                "psp_username": req.psp_username,
                "psp_password": req.psp_password,
                "psp_login_url": req.psp_login_url,
                "psp_data_url": req.psp_data_url,
                "nldc_demand_api_url": req.nldc_demand_api_url,
                "india_15_min_demand_api_url": req.india_15_min_demand_api_url,
                "all_state_demand_api_url": req.all_state_demand_api_url,
                "loadshed_api_url": req.loadshed_api_url,
                "outage_api_url": req.outage_api_url,
                "wbes_url": req.wbes_url,
                "wbes_api_key": req.wbes_api_key,
                "wbes_username": req.wbes_username,
                "wbes_password": req.wbes_password,
                "curve_file_dir": req.curve_file_dir,
                "curve_sheet_name": req.curve_sheet_name,
                "curve_time_column": req.curve_time_column,
                "curve_state_columns": req.curve_state_columns,
                "curve_er_column": req.curve_er_column,
                "curve_peak_hour_by_month": req.curve_peak_hour_by_month,
                "curve_off_peak_hour": req.curve_off_peak_hour
            }},
            upsert=True
        )
        return {
            "success": True,
            "message": "PSP configuration updated successfully."
        }
    except Exception as e:
        return {
            "success": False,
            "message": str(e)
        }

@router.get("/portfolio-mapping")
async def get_psp_portfolio_mapping():
    try:
        db = MongoService()
        states = load_psp_portfolio_base_states(db)
        effective = build_psp_portfolio_states(db)
        candidate_map = {
            state["name"]: state.get("wbes_candidates", [])
            for state in effective
        }
        rows = []
        for state in states:
            row = serialize_psp_portfolio_mapping_state(state)
            row["wbes_candidates"] = candidate_map.get(row["name"], [])
            rows.append(row)
        return {"success": True, "data": rows}
    except Exception as e:
        return {"success": False, "message": str(e), "data": []}

@router.put("/portfolio-mapping")
async def save_psp_portfolio_mapping(rows: list[PSPPortfolioMappingRow]):
    try:
        db = MongoService()
        now_str = datetime.utcnow().isoformat()
        valid_names = {state["name"] for state in PSP_PORTFOLIO_STATE_CONFIGS}
        saved = 0

        for row in rows:
            if row.name not in valid_names:
                continue
            mapping = row.dict()
            db.db["psp_portfolio_mapping"].update_one(
                {"_id": row.name},
                {
                    "$set": {
                        "mapping": mapping,
                        "updated_at": now_str
                    }
                },
                upsert=True
            )
            saved += 1

        # Portfolio cache depends on these SCADA/WBES keys.
        db.db["psp_portfolio_breakdown"].delete_many({})
        return {
            "success": True,
            "message": f"Saved {saved} PSP portfolio mapping rows and cleared cached portfolio calculations."
        }
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.get("/portfolio-curve-headers")
async def get_psp_portfolio_curve_headers(date_str: str = None):
    return await get_psp_curve_headers(date_str)

def get_power_system_states():
    return [state for state in PSP_PEAK_STATES if state["name"] != "ER"]

def to_float(value):
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0

def normalize_check_key(value):
    text = str(value or "").strip().upper()
    text = text.replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    aliases = {
        "ORISSA": "ODISHA",
        "WB": "WEST BENGAL",
        "W BENGAL": "WEST BENGAL",
        "W. BENGAL": "WEST BENGAL",
        "WESTBENGAL": "WEST BENGAL",
        "REGION": "ER",
        "EASTERN REGION": "ER",
        "ER TOTAL": "ER",
    }
    return aliases.get(text, text)

def curve_file_name_for_date(date_str: str):
    dt = date.fromisoformat(date_str)
    return f"curve_{dt.strftime('%d%m%Y')}.xlsm"

def expand_excel_columns(column_range: str):
    text = str(column_range or "").strip().upper()
    if ":" not in text:
        return [text] if text else []
    start, end = [part.strip() for part in text.split(":", 1)]
    start_idx = column_index_from_string(start)
    end_idx = column_index_from_string(end)
    cols = []
    for idx in range(start_idx, end_idx + 1):
        col = ""
        value = idx
        while value:
            value, rem = divmod(value - 1, 26)
            col = chr(65 + rem) + col
        cols.append(col)
    return cols

def get_psp_config_with_curve_defaults():
    config_service = PipelineConfigService()
    config = config_service.get_config("PSP") or {}
    defaults = {
        "curve_file_dir": r"\\10.3.95.200\HTTP-Access\Control_Room_Report\curve",
        "curve_sheet_name": "30SEC",
        "curve_time_column": "C",
        "curve_state_columns": "V:AA",
        "curve_er_column": "AE",
        "curve_peak_hour_by_month": "1:19:00,2:19:00,3:20:00,4:20:00,5:20:00,6:20:00,7:20:00,8:20:00,9:20:00,10:19:00,11:19:00,12:19:00",
        "curve_off_peak_hour": "03:00",
    }
    for key, value in defaults.items():
        config.setdefault(key, value)
    return config

def parse_peak_hour_by_month(value):
    default_map = {month: ("20:00" if 3 <= month <= 9 else "19:00") for month in range(1, 13)}
    text = str(value or "").strip()
    if not text:
        return default_map
    for part in text.split(","):
        if ":" not in part:
            continue
        month_text, hour_text = part.split(":", 1)
        try:
            month = int(month_text.strip())
        except ValueError:
            continue
        if 1 <= month <= 12 and hour_text.strip():
            default_map[month] = hour_text.strip()
    return default_map

def get_configured_peak_hour(config, date_str):
    month = date.fromisoformat(date_str).month
    peak_map = parse_peak_hour_by_month(config.get("curve_peak_hour_by_month"))
    return peak_map.get(month) or ("20:00" if 3 <= month <= 9 else "19:00")

def get_curve_state_mapping(db):
    mapping = {}
    for state in load_psp_portfolio_base_states(db):
        name = state.get("name")
        col = str(state.get("curve_column") or "").strip().upper()
        if name and col:
            mapping[col] = {
                "state": name,
                "header": state.get("curve_header") or name,
            }
    return mapping

def path_exists_with_timeout(path: str, timeout_seconds: float = 3.0):
    result_queue = queue.Queue(maxsize=1)

    def worker():
        try:
            result_queue.put((os.path.exists(path), None))
        except Exception as exc:
            result_queue.put((False, str(exc)))

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    try:
        exists, error = result_queue.get(timeout=timeout_seconds)
        return exists, error
    except queue.Empty:
        return False, f"Timed out checking path after {timeout_seconds:.0f}s"

def get_curve_file_metrics(date_str: str, config: dict):
    curve_dir = str(config.get("curve_file_dir") or "").strip()
    if not curve_dir:
        return {}, {"available": False, "message": "Curve file directory is not configured."}

    file_name = curve_file_name_for_date(date_str)
    file_path = os.path.join(curve_dir, file_name)
    file_exists, path_error = path_exists_with_timeout(file_path)
    if not file_exists:
        suffix = f" ({path_error})" if path_error else ""
        return {}, {"available": False, "path": file_path, "message": f"Curve file not found or unreachable: {file_name}{suffix}"}

    try:
        payload = {
            "date_str": date_str,
            "file_path": file_path,
            "file_name": file_name,
            "sheet_name": config.get("curve_sheet_name") or "30SEC",
            "time_col": config.get("curve_time_column") or "C",
            "state_cols": expand_excel_columns(config.get("curve_state_columns") or "V:AA"),
            "er_col": config.get("curve_er_column") or "AE",
            "state_mapping": config.get("curve_state_mapping") or {},
            "peak_hour": get_configured_peak_hour(config, date_str),
            "off_peak_hour": config.get("curve_off_peak_hour") or "03:00",
        }
        code = r'''
import json
import sys
import atexit
import os
import shutil
import tempfile
import time
import zipfile
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def local_workbook_copy(src):
    last_error = None
    for _ in range(3):
        fd, tmp = tempfile.mkstemp(suffix=".xlsm")
        os.close(fd)
        try:
            shutil.copy2(src, tmp)
            with zipfile.ZipFile(tmp, "r") as archive:
                bad_member = archive.testzip()
            if bad_member:
                raise zipfile.BadZipFile(f"Bad workbook member: {bad_member}")
            atexit.register(lambda path=tmp: os.path.exists(path) and os.remove(path))
            return tmp
        except Exception as exc:
            last_error = exc
            try:
                os.remove(tmp)
            except OSError:
                pass
            time.sleep(1)
    raise RuntimeError(f"Curve workbook is unreadable or still being written: {last_error}")

def normalize(value):
    text = str(value or "").strip().upper().replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    aliases = {
        "ORISSA": "ODISHA",
        "WB": "WEST BENGAL",
        "W BENGAL": "WEST BENGAL",
        "W. BENGAL": "WEST BENGAL",
        "WESTBENGAL": "WEST BENGAL",
        "REGION": "ER",
        "EASTERN REGION": "ER",
        "ER TOTAL": "ER",
    }
    return aliases.get(text, text)

payload = json.loads(sys.stdin.read())
date_str = payload["date_str"]
file_path = payload["file_path"]
file_name = payload["file_name"]
sheet_name = payload["sheet_name"]
time_col = payload["time_col"]
columns = [*payload["state_cols"], payload["er_col"]]
er_col = payload["er_col"]
state_mapping = payload.get("state_mapping") or {}
configured_peak_hour = payload.get("peak_hour") or "20:00"
configured_off_peak_hour = payload.get("off_peak_hour") or "03:00"

def col_to_idx(col):
    idx = 0
    for char in str(col).upper():
        if "A" <= char <= "Z":
            idx = idx * 26 + ord(char) - ord("A") + 1
    return idx - 1

def time_text_from_value(value):
    if hasattr(value, "strftime"):
        if value.__class__.__name__ == "date":
            return "00:00:00"
        return value.strftime("%H:%M:%S")
    return str(value or "").strip()

def best_header(rows, col_idx):
    for row_idx in (4, 5, 3, 2):
        if row_idx >= len(rows) or col_idx >= len(rows[row_idx]):
            continue
        value = str(rows[row_idx][col_idx] or "").strip()
        if value and not value.isdigit() and value.upper() != "X":
            return value.replace(" demand", "").replace(" Demand", "")
    return ""

try:
    from python_calamine import load_workbook as load_calamine_workbook
    calamine_wb = load_calamine_workbook(file_path)
    if sheet_name not in calamine_wb.sheet_names:
        print(json.dumps({"metrics": {}, "meta": {"available": False, "path": file_path, "message": f"Sheet not found: {sheet_name}"}}))
        raise SystemExit(0)
    sheet = calamine_wb.get_sheet_by_name(sheet_name)
    rows = sheet.to_python()
    header_by_col = {}
    for col in columns:
        mapped = state_mapping.get(col) or state_mapping.get(str(col).upper())
        key = mapped.get("state") if isinstance(mapped, dict) else None
        if not key:
            key = normalize(best_header(rows, col_to_idx(col)))
        if col == er_col:
            key = "ER"
        if key:
            header_by_col[col] = key

    series_by_key = {key: [] for key in header_by_col.values()}
    time_idx = col_to_idx(time_col)
    col_indexes = {col: col_to_idx(col) for col in header_by_col.keys()}
    for row in rows[6:2886]:
        time_value = row[time_idx] if time_idx < len(row) else None
        time_text = time_text_from_value(time_value)
        for col, key in header_by_col.items():
            idx = col_indexes[col]
            value = row[idx] if idx < len(row) else None
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                numeric = None
            if numeric is not None:
                series_by_key.setdefault(key, []).append({"time": time_text, "value": numeric})

    peak_hour = configured_peak_hour
    metrics = {}
    for key, points in series_by_key.items():
        values = [point["value"] for point in points if point.get("value") is not None]
        if not values:
            continue
        def nearest_value(target_hhmm):
            target = target_hhmm.replace(":", "")
            for point in points:
                clean_time = str(point.get("time") or "").replace(":", "")[:4]
                if clean_time == target:
                    return point["value"]
            return None
        metrics[key] = {
            "energy": round(sum(values) / 120000.0, 3),
            "max_demand": round(max(values), 2),
            "min_demand": round(min(values), 2),
            "peak_demand": round(nearest_value(peak_hour) or 0.0, 2),
            "off_peak_demand": round(nearest_value(configured_off_peak_hour) or 0.0, 2),
            "peak_hour": peak_hour,
            "source": file_name,
        }
    calamine_wb.close()
    print(json.dumps({"metrics": metrics, "meta": {"available": True, "path": file_path, "sheet": sheet_name, "file": file_name, "reader": "python-calamine"}}))
    raise SystemExit(0)
except SystemExit:
    raise
except Exception:
    pass

try:
    local_path = local_workbook_copy(file_path)
    wb = load_workbook(local_path, read_only=True, data_only=True, keep_vba=False)
except Exception as exc:
    print(json.dumps({"metrics": {}, "meta": {"available": False, "path": file_path, "message": str(exc)}}))
    raise SystemExit(0)
if sheet_name not in wb.sheetnames:
    print(json.dumps({"metrics": {}, "meta": {"available": False, "path": file_path, "message": f"Sheet not found: {sheet_name}"}}))
    raise SystemExit(0)
ws = wb[sheet_name]
header_by_col = {}
for col in columns:
    header = ws[f"{col}6"].value
    mapped = state_mapping.get(col) or state_mapping.get(str(col).upper())
    key = mapped.get("state") if isinstance(mapped, dict) else None
    if not key:
        key = normalize(header)
    if col == er_col:
        key = "ER"
    if key:
        header_by_col[col] = key

series_by_key = {key: [] for key in header_by_col.values()}
min_col = min(column_index_from_string(col) for col in [time_col, *header_by_col.keys()])
max_col = max(column_index_from_string(col) for col in [time_col, *header_by_col.keys()])
time_offset = column_index_from_string(time_col) - min_col
col_offsets = {
    col: column_index_from_string(col) - min_col
    for col in header_by_col.keys()
}
for row in ws.iter_rows(min_row=8, max_row=2887, min_col=min_col, max_col=max_col, values_only=True):
    time_value = row[time_offset] if time_offset < len(row) else None
    time_text = str(time_value or "").strip()
    if hasattr(time_value, "strftime"):
        time_text = time_value.strftime("%H:%M:%S")
    for col, key in header_by_col.items():
        offset = col_offsets.get(col)
        value = row[offset] if offset is not None and offset < len(row) else None
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            numeric = None
        if numeric is not None:
            series_by_key.setdefault(key, []).append({"time": time_text, "value": numeric})

peak_hour = configured_peak_hour
metrics = {}
for key, points in series_by_key.items():
    values = [point["value"] for point in points if point.get("value") is not None]
    if not values:
        continue
    def nearest_value(target_hhmm):
        target = target_hhmm.replace(":", "")
        for point in points:
            clean_time = str(point.get("time") or "").replace(":", "")[:4]
            if clean_time == target:
                return point["value"]
        return None
    metrics[key] = {
        "energy": round(sum(values) / 120000.0, 3),
        "max_demand": round(max(values), 2),
        "min_demand": round(min(values), 2),
        "peak_demand": round(nearest_value(peak_hour) or 0.0, 2),
        "off_peak_demand": round(nearest_value(configured_off_peak_hour) or 0.0, 2),
        "peak_hour": peak_hour,
        "source": file_name,
    }
wb.close()
print(json.dumps({"metrics": metrics, "meta": {"available": True, "path": file_path, "sheet": sheet_name, "file": file_name}}))
'''
        completed = subprocess.run(
            [sys.executable, "-c", code],
            input=json.dumps(payload),
            text=True,
            capture_output=True,
            timeout=45,
        )
        if completed.returncode != 0:
            return {}, {"available": False, "path": file_path, "message": completed.stderr.strip() or "Curve read failed"}
        parsed = json.loads(completed.stdout or "{}")
        return parsed.get("metrics", {}), parsed.get("meta", {"available": False, "path": file_path})
    except subprocess.TimeoutExpired:
        return {}, {"available": False, "path": file_path, "message": f"Curve file read timed out: {file_name}"}
    except Exception as exc:
        return {}, {"available": False, "path": file_path, "message": f"Curve read failed: {exc}"}

def read_curve_file_metrics_sync(date_str: str, file_path: str, config: dict, file_name: str):
    sheet_name = config.get("curve_sheet_name") or "30SEC"
    time_col = config.get("curve_time_column") or "C"
    state_cols = expand_excel_columns(config.get("curve_state_columns") or "V:AA")
    er_col = config.get("curve_er_column") or "AE"
    configured_peak_hour = get_configured_peak_hour(config, date_str)
    configured_off_peak_hour = config.get("curve_off_peak_hour") or "03:00"
    columns = [*state_cols, er_col]

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_vba=False)
        if sheet_name not in wb.sheetnames:
            return {}, {"available": False, "path": file_path, "message": f"Sheet not found: {sheet_name}"}
        ws = wb[sheet_name]

        header_by_col = {}
        for col in columns:
            header = ws[f"{col}6"].value
            key = normalize_check_key(header)
            if not key and col == er_col:
                key = "ER"
            if key:
                header_by_col[col] = key

        series_by_key = {key: [] for key in header_by_col.values()}
        time_by_row = {}
        for row_idx in range(8, 2888):
            time_value = ws[f"{time_col}{row_idx}"].value
            time_text = str(time_value or "").strip()
            if hasattr(time_value, "strftime"):
                time_text = time_value.strftime("%H:%M:%S")
            time_by_row[row_idx] = time_text
            for col, key in header_by_col.items():
                value = ws[f"{col}{row_idx}"].value
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    numeric = None
                if numeric is not None:
                    series_by_key.setdefault(key, []).append({"time": time_text, "value": numeric})
    except Exception as exc:
        return {}, {"available": False, "path": file_path, "message": f"Curve read failed: {exc}"}

    peak_hour = configured_peak_hour
    metrics = {}
    for key, points in series_by_key.items():
        values = [point["value"] for point in points if point.get("value") is not None]
        if not values:
            continue

        def nearest_value(target_hhmm):
            target = target_hhmm.replace(":", "")
            for point in points:
                clean_time = str(point.get("time") or "").replace(":", "")[:4]
                if clean_time == target:
                    return point["value"]
            return None

        metrics[key] = {
            "energy": round(sum(values) / 120000.0, 3),
            "max_demand": round(max(values), 2),
            "min_demand": round(min(values), 2),
            "peak_demand": round(nearest_value(peak_hour) or 0.0, 2),
            "off_peak_demand": round(nearest_value(configured_off_peak_hour) or 0.0, 2),
            "peak_hour": peak_hour,
            "source": file_name,
        }

    return metrics, {"available": True, "path": file_path, "sheet": sheet_name, "file": file_name}

def save_curve_metrics_cache(db, date_str: str, metrics: dict, meta: dict):
    db.db["psp_curve_metrics"].update_one(
        {"_id": date_str},
        {"$set": {
            "_id": date_str,
            "date": date_str,
            "metrics": metrics or {},
            "meta": meta or {},
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )

def fetch_and_cache_curve_metrics(db, date_str: str):
    config = get_psp_config_with_curve_defaults()
    config["curve_state_mapping"] = get_curve_state_mapping(db)
    metrics, meta = get_curve_file_metrics(date_str, config)
    save_curve_metrics_cache(db, date_str, metrics, meta)
    return metrics, meta

def get_curve_metrics_from_cache(db, date_str: str, refresh: bool = False):
    if refresh:
        return fetch_and_cache_curve_metrics(db, date_str)

    doc = db.db["psp_curve_metrics"].find_one({"_id": date_str}, {"_id": 0})
    if doc:
        meta = doc.get("meta") or {}
        meta.setdefault("source", "mongo_cache")
        return doc.get("metrics") or {}, meta

    return {}, {
        "available": False,
        "source": "mongo_cache",
        "message": "Curve metrics are not cached. Use Fetch PSP Sources to read the curve file and save it in Mongo."
    }

def expand_diurnal_dates(range_req: DiurnalCurveRange):
    if range_req.selected_dates:
        dates = sorted({date.fromisoformat(item).strftime("%Y-%m-%d") for item in range_req.selected_dates})
        return dates
    start_dt = date.fromisoformat(range_req.start_date)
    end_dt = date.fromisoformat(range_req.end_date)
    if start_dt > end_dt:
        raise ValueError("Start date must be less than or equal to end date.")
    return [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end_dt - start_dt).days + 1)]

def normalize_time_block(time_text, block_minutes: int):
    text = str(time_text or "").strip()
    if not text:
        return None
    parts = text.split(":")
    try:
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        second = int(float(parts[2])) if len(parts) > 2 else 0
    except (TypeError, ValueError):
        return None
    total_seconds = hour * 3600 + minute * 60 + second
    block_seconds = max(1, int(block_minutes)) * 60
    block_start = (total_seconds // block_seconds) * block_seconds
    block_start = min(block_start, 23 * 3600 + 59 * 60)
    return f"{block_start // 3600:02d}:{(block_start % 3600) // 60:02d}"

def read_curve_file_series(date_str: str, states: list[str], config: dict):
    curve_dir = str(config.get("curve_file_dir") or "").strip()
    if not curve_dir:
        return {}, {"available": False, "message": "Curve file directory is not configured."}

    file_name = curve_file_name_for_date(date_str)
    file_path = os.path.join(curve_dir, file_name)
    file_exists, path_error = path_exists_with_timeout(file_path)
    if not file_exists:
        suffix = f" ({path_error})" if path_error else ""
        return {}, {"available": False, "path": file_path, "message": f"Curve file not found or unreachable: {file_name}{suffix}"}

    payload = {
        "file_path": file_path,
        "file_name": file_name,
        "sheet_name": config.get("curve_sheet_name") or "30SEC",
        "time_col": config.get("curve_time_column") or "C",
        "state_cols": expand_excel_columns(config.get("curve_state_columns") or "V:AA"),
        "er_col": config.get("curve_er_column") or "AE",
        "state_mapping": config.get("curve_state_mapping") or {},
        "states": states,
    }
    code = r'''
import json
import sys
import atexit
import os
import shutil
import tempfile
import time
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def normalize(value):
    text = str(value or "").strip().upper().replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    aliases = {
        "ORISSA": "ODISHA",
        "WB": "WEST BENGAL",
        "W BENGAL": "WEST BENGAL",
        "W. BENGAL": "WEST BENGAL",
        "WESTBENGAL": "WEST BENGAL",
        "REGION": "ER",
        "EASTERN REGION": "ER",
        "ER TOTAL": "ER",
    }
    return aliases.get(text, text)

def local_workbook_copy(src):
    last_error = None
    for _ in range(3):
        fd, tmp = tempfile.mkstemp(suffix=".xlsm")
        os.close(fd)
        try:
            shutil.copy2(src, tmp)
            with zipfile.ZipFile(tmp, "r") as archive:
                bad_member = archive.testzip()
            if bad_member:
                raise zipfile.BadZipFile(f"Bad workbook member: {bad_member}")
            atexit.register(lambda path=tmp: os.path.exists(path) and os.remove(path))
            return tmp
        except Exception as exc:
            last_error = exc
            try:
                os.remove(tmp)
            except OSError:
                pass
            time.sleep(1)
    raise RuntimeError(f"Curve workbook is unreadable or still being written: {last_error}")

def col_to_idx(col):
    idx = 0
    for char in str(col).upper():
        if "A" <= char <= "Z":
            idx = idx * 26 + ord(char) - ord("A") + 1
    return idx - 1

def time_text_from_value(value):
    if hasattr(value, "strftime"):
        if value.__class__.__name__ == "date":
            return "00:00:00"
        return value.strftime("%H:%M:%S")
    return str(value or "").strip()

def best_header(rows, col_idx):
    for row_idx in (4, 5, 3, 2):
        if row_idx >= len(rows) or col_idx >= len(rows[row_idx]):
            continue
        value = str(rows[row_idx][col_idx] or "").strip()
        if value and not value.isdigit() and value.upper() != "X":
            return value.replace(" demand", "").replace(" Demand", "")
    return ""

payload = json.loads(sys.stdin.read())
sheet_name = payload["sheet_name"]
time_col = payload["time_col"]
columns = [*payload["state_cols"], payload["er_col"]]
er_col = payload["er_col"]
state_mapping = payload.get("state_mapping") or {}
requested = {normalize(item) for item in payload.get("states") or []}
if not requested:
    requested = {"ER", "BIHAR", "DVC", "JHARKHAND", "ODISHA", "SIKKIM", "WEST BENGAL"}

def mapped_headers_from_rows(rows):
    header_by_col = {}
    for col in columns:
        mapped = state_mapping.get(col) or state_mapping.get(str(col).upper())
        key = mapped.get("state") if isinstance(mapped, dict) else None
        if not key:
            key = normalize(best_header(rows, col_to_idx(col)))
        if col == er_col:
            key = "ER"
        key = normalize(key)
        if key in requested:
            header_by_col[col] = key
    return header_by_col

try:
    from python_calamine import load_workbook as load_calamine_workbook
    wb = load_calamine_workbook(payload["file_path"])
    if sheet_name not in wb.sheet_names:
        print(json.dumps({"series": {}, "meta": {"available": False, "message": f"Sheet not found: {sheet_name}"}}))
        raise SystemExit(0)
    rows = wb.get_sheet_by_name(sheet_name).to_python()
    header_by_col = mapped_headers_from_rows(rows)
    series = {key: [] for key in header_by_col.values()}
    time_idx = col_to_idx(time_col)
    col_indexes = {col: col_to_idx(col) for col in header_by_col}
    for row in rows[6:2886]:
        time_value = row[time_idx] if time_idx < len(row) else None
        time_text = time_text_from_value(time_value)
        for col, key in header_by_col.items():
            idx = col_indexes[col]
            value = row[idx] if idx < len(row) else None
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            series.setdefault(key, []).append({"time": time_text, "value": numeric})
    wb.close()
    print(json.dumps({"series": series, "meta": {"available": True, "reader": "python-calamine", "file": payload["file_name"]}}))
    raise SystemExit(0)
except SystemExit:
    raise
except Exception:
    pass

try:
    local_path = local_workbook_copy(payload["file_path"])
    wb = load_workbook(local_path, read_only=True, data_only=True, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        print(json.dumps({"series": {}, "meta": {"available": False, "message": f"Sheet not found: {sheet_name}"}}))
        raise SystemExit(0)
    ws = wb[sheet_name]
    header_by_col = {}
    for col in columns:
        mapped = state_mapping.get(col) or state_mapping.get(str(col).upper())
        key = mapped.get("state") if isinstance(mapped, dict) else None
        if not key:
            key = normalize(ws[f"{col}6"].value)
        if col == er_col:
            key = "ER"
        key = normalize(key)
        if key in requested:
            header_by_col[col] = key
    series = {key: [] for key in header_by_col.values()}
    min_col = min(column_index_from_string(col) for col in [time_col, *header_by_col.keys()])
    max_col = max(column_index_from_string(col) for col in [time_col, *header_by_col.keys()])
    time_offset = column_index_from_string(time_col) - min_col
    col_offsets = {col: column_index_from_string(col) - min_col for col in header_by_col.keys()}
    for row in ws.iter_rows(min_row=8, max_row=2887, min_col=min_col, max_col=max_col, values_only=True):
        time_value = row[time_offset] if time_offset < len(row) else None
        time_text = time_text_from_value(time_value)
        for col, key in header_by_col.items():
            offset = col_offsets[col]
            value = row[offset] if offset < len(row) else None
            try:
                numeric = float(value)
            except (TypeError, ValueError):
                continue
            series.setdefault(key, []).append({"time": time_text, "value": numeric})
    wb.close()
    print(json.dumps({"series": series, "meta": {"available": True, "reader": "openpyxl", "file": payload["file_name"]}}))
except Exception as exc:
    print(json.dumps({"series": {}, "meta": {"available": False, "message": str(exc), "file": payload["file_name"]}}))
'''
    try:
        completed = subprocess.run(
            [sys.executable, "-c", code],
            input=json.dumps(payload),
            text=True,
            capture_output=True,
            timeout=60,
        )
        if completed.returncode != 0:
            return {}, {"available": False, "path": file_path, "message": completed.stderr.strip() or "Curve read failed"}
        parsed = json.loads(completed.stdout or "{}")
        return parsed.get("series") or {}, parsed.get("meta") or {"available": False}
    except subprocess.TimeoutExpired:
        return {}, {"available": False, "path": file_path, "message": f"Curve file read timed out: {file_name}"}
    except Exception as exc:
        return {}, {"available": False, "path": file_path, "message": f"Curve read failed: {exc}"}

def aggregate_diurnal_series(series_by_date_state: dict, req: DiurnalCurveRequest):
    block_minutes = max(1, int(req.block_minutes or 15))
    curves = {}
    meta = {}

    def add_point(curve_key, time_key, value):
        bucket = curves.setdefault(curve_key, {}).setdefault(time_key, [])
        bucket.append(float(value))

    for range_req in req.date_ranges:
        dates = expand_diurnal_dates(range_req)
        label_prefix = range_req.label.strip() or f"{range_req.start_date} to {range_req.end_date}"
        curve_type = (range_req.curve_type or "daily").lower()
        for date_str in dates:
            dt = date.fromisoformat(date_str)
            month_label = dt.strftime("%b-%Y")
            day_type = "Weekend" if dt.weekday() >= 5 else "Weekday"
            for state in req.states:
                points = series_by_date_state.get(date_str, {}).get(normalize_check_key(state), [])
                if not points:
                    continue
                if curve_type == "monthly_average":
                    curve_key = f"{state}|{label_prefix}|{month_label} Avg"
                    curve_label = f"{state} {month_label} Avg"
                elif curve_type == "weekday_weekend_average":
                    curve_key = f"{state}|{label_prefix}|{day_type} Avg"
                    curve_label = f"{state} {day_type} Avg"
                else:
                    curve_key = f"{state}|{date_str}"
                    curve_label = f"{state} {date_str}"
                meta[curve_key] = {
                    "key": curve_key,
                    "label": curve_label,
                    "state": state,
                    "curve_type": curve_type,
                    "range_label": label_prefix,
                }
                for point in points:
                    time_key = normalize_time_block(point.get("time"), block_minutes)
                    if time_key is not None:
                        add_point(curve_key, time_key, point.get("value"))

    chart_by_time = {}
    table_rows = []
    for curve_key, by_time in curves.items():
        for time_key, values in by_time.items():
            avg_value = round(sum(values) / len(values), 3) if values else None
            chart_by_time.setdefault(time_key, {"time": time_key})[curve_key] = avg_value
            table_rows.append({
                "time": time_key,
                "curve_key": curve_key,
                "curve_label": meta.get(curve_key, {}).get("label", curve_key),
                "value": avg_value,
            })

    chart_rows = [chart_by_time[key] for key in sorted(chart_by_time.keys())]
    table_rows.sort(key=lambda row: (row["curve_label"], row["time"]))
    return {
        "series": list(meta.values()),
        "chart_rows": chart_rows,
        "table_rows": table_rows,
    }

def refresh_psp_operational_sources(target_date: date):
    db = MongoService()
    date_str = target_date.strftime("%Y-%m-%d")
    results = {}

    try:
        psp_result = PSPService.fetch_and_save_date(target_date)
        results["psp_data"] = {"success": True, "result": psp_result}
    except Exception as exc:
        results["psp_data"] = {"success": False, "message": str(exc)}

    try:
        check_and_update_highest_portfolio(target_date)
        calculate_and_save_portfolio(target_date)
        results["portfolio"] = {"success": True}
    except Exception as exc:
        results["portfolio"] = {"success": False, "message": str(exc)}

    try:
        rows = fetch_and_cache_loadshedding(db, target_date)
        results["loadshedding_hourly"] = {"success": True, "records": len(rows or [])}
    except Exception as exc:
        results["loadshedding_hourly"] = {"success": False, "message": str(exc)}

    try:
        session = make_legacy_session()
        target_rows = fetch_outage_details_for_date(db, target_date, session=session)
        previous_rows = fetch_outage_details_for_date(db, target_date - timedelta(days=1), session=session)
        changes = build_generation_outage_changes(target_rows, previous_rows)
        summary = {
            "restored_mw": changes["restored_mw"],
            "tripped_mw": changes["tripped_mw"],
            "net_mw": changes["net_mw"],
            "restored_count": len(changes["restored"]),
            "tripped_count": len(changes["tripped"]),
        }
        db.db["psp_generation_outage_changes"].update_one(
            {"_id": date_str},
            {"$set": {
                "_id": date_str,
                "date": date_str,
                "previous_date": (target_date - timedelta(days=1)).strftime("%Y-%m-%d"),
                "summary": summary,
                "restored": changes["restored"],
                "tripped": changes["tripped"],
                "updated_at": datetime.utcnow().isoformat()
            }},
            upsert=True
        )
        results["generation_outage_changes"] = {"success": True, **summary}
    except Exception as exc:
        results["generation_outage_changes"] = {"success": False, "message": str(exc)}

    try:
        curve_metrics, curve_meta = fetch_and_cache_curve_metrics(db, date_str)
        results["curve_metrics"] = {
            "success": bool(curve_meta.get("available")),
            "records": len(curve_metrics or {}),
            "meta": curve_meta
        }
    except Exception as exc:
        results["curve_metrics"] = {"success": False, "message": str(exc)}

    return results

def extract_operational_check_values(doc):
    values = {}
    if not doc:
        return values

    def negative_value(raw_value):
        if raw_value is None or raw_value == "":
            return None
        try:
            return -abs(float(raw_value))
        except (TypeError, ValueError):
            return None

    demand_rows = doc.get("pspstatedemandrequirement", []) or []
    for row in demand_rows:
        key = normalize_check_key(row.get("STATE_NAME"))
        if not key:
            continue
        values.setdefault(key, {})
        values[key]["max_demand"] = to_float(row.get("MAX_DEMAND"))
        values[key]["max_demand_time"] = str(row.get("MAX_DEMAND_TIME") or "")
        values[key]["min_demand"] = (
            to_float(row.get("MIN_DEMAND_MET"))
            if row.get("MIN_DEMAND_MET") is not None
            else None
        )
        values[key]["max_demand_shortage"] = negative_value(row.get("MAX_DEMAND_SHORTAGE_RGN"))

    load_rows = doc.get("pspstateloaddetailsER", []) or []
    er_energy = 0.0
    for row in load_rows:
        key = normalize_check_key(row.get("STATE_NAME"))
        energy = to_float(row.get("CONSUMPTION"))
        er_energy += energy
        if key:
            values.setdefault(key, {})
            values[key]["energy"] = energy
            values[key]["day_shortage"] = negative_value(row.get("DAY_SHORT"))
    values.setdefault("ER", {})
    values["ER"]["energy"] = er_energy

    forecast_rows = doc.get("pspstatedemandforecast", []) or []
    for row in forecast_rows:
        key = normalize_check_key(row.get("STATE_NAME"))
        if not key:
            continue
        values.setdefault(key, {})
        values[key]["peak_demand"] = (
            to_float(row.get("PEAK_MAX_DEMAND_MET"))
            if row.get("PEAK_MAX_DEMAND_MET") is not None
            else None
        )
        values[key]["peak_demand_time"] = str(row.get("PEAK_TIME") or "")
        values[key]["off_peak_demand"] = (
            to_float(row.get("OFF_PEAK_MAX_DEMAND_MET"))
            if row.get("OFF_PEAK_MAX_DEMAND_MET") is not None
            else None
        )
        values[key]["off_peak_demand_time"] = str(row.get("OFF_PEAK_TIME") or "")
        values[key]["peak_shortage"] = negative_value(row.get("PK_SHORT"))
        values[key]["off_peak_shortage"] = negative_value(row.get("OFFPK_SHORT"))

    er_demand_row = next((row for row in demand_rows if normalize_check_key(row.get("STATE_NAME")) == "ER"), None)
    if er_demand_row:
        values["ER"]["max_demand"] = to_float(er_demand_row.get("MAX_DEMAND"))
        values["ER"]["min_demand"] = to_float(er_demand_row.get("MIN_DEMAND_MET"))
    else:
        values["ER"]["max_demand"] = sum(item.get("max_demand", 0.0) for key, item in values.items() if key != "ER")
        min_values = [item.get("min_demand") for key, item in values.items() if key != "ER" and item.get("min_demand") is not None]
        values["ER"]["min_demand"] = sum(min_values) if min_values else None

    for shortage_key in ("day_shortage", "max_demand_shortage", "peak_shortage", "off_peak_shortage"):
        if values["ER"].get(shortage_key) is None:
            shortage_values = [
                item.get(shortage_key)
                for state_key, item in values.items()
                if state_key != "ER" and item.get(shortage_key) is not None
            ]
            values["ER"][shortage_key] = round(sum(shortage_values), 3) if shortage_values else None

    for demand_key in ("peak_demand", "off_peak_demand"):
        if values["ER"].get(demand_key) is None:
            demand_values = [
                item.get(demand_key)
                for state_key, item in values.items()
                if state_key != "ER" and item.get(demand_key) is not None
            ]
            values["ER"][demand_key] = round(sum(demand_values), 3) if demand_values else None

    return values

def get_operational_check_values(db, date_str: str):
    doc = db.psp_collection.find_one(
        {"date": date_str},
        {"_id": 0, "date": 1, "pspstateloaddetailsER": 1, "pspstatedemandrequirement": 1, "pspstatedemandforecast": 1}
    )
    return extract_operational_check_values(doc)

def extract_frequency_check(doc):
    if not doc:
        return {
            "bands": [],
            "max_min": {},
            "total_band_percent": None,
        }
    profile = (doc.get("pspFrequencyProfile") or [{}])[0] or {}
    max_min = (doc.get("pspFrequencyProfileMaxMin") or [{}])[0] or {}

    def optional_float(value):
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    bands = [
        {
            "key": "below_band",
            "label": "Below The Band",
            "logic": profile.get("FREQ4") or "< 49.9",
            "source_key": "pspFrequencyProfile.FREQ4_VALUE",
            "value": optional_float(profile.get("FREQ4_VALUE")),
            "color": "#DC2626",
        },
        {
            "key": "within_band",
            "label": "Within The Band",
            "logic": profile.get("FREQ5") or ">= 49.9 - <= 50.05",
            "source_key": "pspFrequencyProfile.FREQ5_VALUE",
            "value": optional_float(profile.get("FREQ5_VALUE")),
            "color": "#059669",
        },
        {
            "key": "above_band",
            "label": "Above The Band",
            "logic": profile.get("FREQ7") or "> 50.05",
            "source_key": "pspFrequencyProfile.FREQ7_VALUE",
            "value": optional_float(profile.get("FREQ7_VALUE")),
            "color": "#D97706",
        },
    ]
    total_values = [band["value"] for band in bands if band["value"] is not None]
    return {
        "bands": bands,
        "total_band_percent": round(sum(total_values), 3) if total_values else None,
        "max_min": {
            "max_freq": optional_float(max_min.get("MAX_FREQ")),
            "max_time": max_min.get("MAX_TIME"),
            "min_freq": optional_float(max_min.get("MIN_FREQ")),
            "min_time": max_min.get("MIN_TIME"),
            "average_frequency": optional_float(max_min.get("AVERAGE_FREQUENCY")),
            "source_key": "pspFrequencyProfileMaxMin.MAX_FREQ / MIN_FREQ",
        },
        "logic": [
            "Below The Band: pspFrequencyProfile.FREQ4 logic and FREQ4_VALUE",
            "Within The Band: pspFrequencyProfile.FREQ5 logic and FREQ5_VALUE",
            "Above The Band: pspFrequencyProfile.FREQ7 logic and FREQ7_VALUE",
            "Band total check: FREQ4_VALUE + FREQ5_VALUE + FREQ7_VALUE should be 100",
            "Max/Min: pspFrequencyProfileMaxMin.MAX_FREQ and MIN_FREQ",
        ],
    }

def build_transnational_exchange_check(doc):
    def optional_float(value):
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def day_order_status(day_peak, day_avg, day_min):
        values = [optional_float(day_peak), optional_float(day_avg), optional_float(day_min)]
        if any(value is None for value in values):
            return {"ok": None, "message": "Missing peak/avg/min"}
        peak, avg, min_val = [abs(value) for value in values]
        if peak == 0 and avg == 0 and min_val == 0:
            return {"ok": True, "message": "All zero"}
        ok = peak >= avg >= min_val
        return {
            "ok": ok,
            "message": "abs(Day_Peak) >= abs(DAY_AVG) >= abs(DAY_MIN)" if ok else "Day peak/avg/min order mismatch"
        }

    def schedule_actual_status(schedule, actual):
        schedule_val = optional_float(schedule)
        actual_val = optional_float(actual)
        if schedule_val is None or actual_val is None:
            return {"ok": None, "diff_percent": None, "message": "Schedule/actual not available"}
        denominator = abs(schedule_val)
        diff_percent = 0.0 if denominator == 0 and abs(actual_val) == 0 else None
        if diff_percent is None and denominator > 0:
            diff_percent = abs(actual_val - schedule_val) / denominator * 100.0
        ok = diff_percent is not None and diff_percent <= 2.0
        return {
            "ok": ok,
            "diff_percent": round(diff_percent, 3) if diff_percent is not None else None,
            "message": "Within 2%" if ok else "Schedule/actual difference exceeds 2%"
        }

    rows = []
    for row in (doc or {}).get("pspTransnationalExchangeState", []) or []:
        day_status = day_order_status(row.get("Day_Peak"), row.get("DAY_AVG"), row.get("DAY_MIN"))
        schedule_status = schedule_actual_status(row.get("Scheduled_EX"), row.get("ACTUAL_EX"))
        rows.append({
            "type": "State",
            "name": row.get("STATE_NAME") or "-",
            "day_peak": optional_float(row.get("Day_Peak")),
            "day_avg": optional_float(row.get("DAY_AVG")),
            "day_min": optional_float(row.get("DAY_MIN")),
            "scheduled_ex": optional_float(row.get("Scheduled_EX")),
            "actual_ex": optional_float(row.get("ACTUAL_EX")),
            "schedule_actual_diff_percent": schedule_status.get("diff_percent"),
            "day_order_ok": day_status.get("ok"),
            "schedule_actual_ok": schedule_status.get("ok"),
            "status": "OK" if day_status.get("ok") is True and schedule_status.get("ok") is True else "CHECK",
            "remarks": f"{day_status.get('message')}; {schedule_status.get('message')}",
        })

    for row in (doc or {}).get("pspTransnationalExchangeLine", []) or []:
        day_status = day_order_status(row.get("Day_Peak"), row.get("DAY_AVG"), row.get("DAY_MIN"))
        rows.append({
            "type": "Line",
            "name": row.get("LINE_NAME") or "-",
            "day_peak": optional_float(row.get("Day_Peak")),
            "day_avg": optional_float(row.get("DAY_AVG")),
            "day_min": optional_float(row.get("DAY_MIN")),
            "scheduled_ex": None,
            "actual_ex": optional_float(row.get("ENERGY_EXCHANGE")),
            "schedule_actual_diff_percent": None,
            "day_order_ok": day_status.get("ok"),
            "schedule_actual_ok": None,
            "status": "OK" if day_status.get("ok") is True else "CHECK",
            "remarks": day_status.get("message"),
        })

    return {
        "key": "transnational_exchange",
        "title": "Transnational Exchange Check",
        "unit": "MW / MU",
        "table_type": "transnational",
        "include_curve_columns": False,
        "source_key": "State: pspTransnationalExchangeState. Line: pspTransnationalExchangeLine. Day order uses absolute values to handle import/export sign convention; schedule vs actual must be within 2%.",
        "logic": [
            "Logic 1: abs(Day_Peak) >= abs(DAY_AVG) >= abs(DAY_MIN); all-zero line values pass.",
            "Logic 2: abs(ACTUAL_EX - Scheduled_EX) / abs(Scheduled_EX) <= 2% for state rows.",
        ],
        "rows": rows,
    }

def average_metric_over_dates(db, date_list, metric_key, state_key):
    vals = []
    for day in date_list:
        values = get_operational_check_values(db, day.strftime("%Y-%m-%d"))
        val = values.get(state_key, {}).get(metric_key)
        if val is not None:
            vals.append(to_float(val))
    return round(sum(vals) / len(vals), 3) if vals else None

def hourly_boundary_hour(time_value):
    text = str(time_value or "").strip()
    if not text:
        return None
    parts = text.split(":")
    if len(parts) < 2:
        return None
    try:
        hour = int(parts[0])
        minute = int(parts[1])
        second = int(parts[2]) if len(parts) > 2 else 0
    except ValueError:
        return None
    if minute == 0 and second == 0 and 0 <= hour <= 23:
        return hour
    return None

def build_hourly_demand_lookup(rows):
    state_order = ["BIHAR", "JHARKHAND", "DVC", "ODISHA", "WEST BENGAL", "SIKKIM"]
    lookup = {}
    er_by_hour = {}
    values_by_state = {}
    for item in rows or []:
        state = normalize_check_key(item.get("STATE_NAME"))
        try:
            hour = int(item.get("HOUR_ID"))
        except (TypeError, ValueError):
            continue
        demand = item.get("HOUR_DEMAND")
        if demand is None:
            continue
        demand = to_float(demand)
        lookup[(state, hour)] = demand
        values_by_state.setdefault(state, []).append(demand)
        if state in state_order:
            er_by_hour[hour] = er_by_hour.get(hour, 0.0) + demand
    for hour, demand in er_by_hour.items():
        lookup[("ER", hour)] = round(demand, 3)
    if er_by_hour:
        values_by_state["ER"] = list(er_by_hour.values())
    min_by_state = {
        state: round(min(values), 3)
        for state, values in values_by_state.items()
        if values
    }
    return {"hourly": lookup, "min": min_by_state}

def get_hourly_demand_lookup(db, target_date_str):
    doc = db.db["psp_loadshedding"].find_one({"_id": target_date_str}, {"_id": 0, "rows": 1})
    rows = doc.get("rows", []) if doc else fetch_and_cache_loadshedding(db, date.fromisoformat(target_date_str))
    return build_hourly_demand_lookup(rows)

def build_report_checking_rows(db, target_date_str: str, curve_metrics: dict):
    target_dt = date.fromisoformat(target_date_str)
    target_values = get_operational_check_values(db, target_date_str)
    last_7_dates = [target_dt - timedelta(days=i) for i in range(1, 8)]
    last_week_str = (target_dt - timedelta(days=7)).strftime("%Y-%m-%d")
    last_year_str = target_dt.replace(year=target_dt.year - 1).strftime("%Y-%m-%d")
    last_week_values = get_operational_check_values(db, last_week_str)
    last_year_values = get_operational_check_values(db, last_year_str)

    states = ["BIHAR", "DVC", "JHARKHAND", "ODISHA", "SIKKIM", "WEST BENGAL", "ER"]
    metric_defs = [
        ("energy", "Energy Consumption", "MU", True, "Operational: pspstateloaddetailsER.CONSUMPTION. Curve: sum of 30-sec MW samples / 120000."),
    ]

    tables = []
    for metric_key, title, unit, include_curve_columns, source_key in metric_defs:
        rows = []
        for state in states:
            curve_value = curve_metrics.get(state, {}).get(metric_key)
            rows.append({
                "state": state,
                "operational_day": target_values.get(state, {}).get(metric_key),
                "average_last_7_days": average_metric_over_dates(db, last_7_dates, metric_key, state),
                "same_day_last_week": last_week_values.get(state, {}).get(metric_key),
                "same_day_last_year": last_year_values.get(state, {}).get(metric_key),
                "curve_file": curve_value,
                "difference_operational_vs_curve": (
                    round(to_float(target_values.get(state, {}).get(metric_key)) - to_float(curve_value), 3)
                    if target_values.get(state, {}).get(metric_key) is not None and curve_value is not None
                    else None
                ),
            })
        tables.append({
            "key": metric_key,
            "title": title,
            "unit": unit,
            "table_type": "comparison",
            "include_curve_columns": include_curve_columns,
            "source_key": source_key,
            "rows": rows
        })
    demand_metrics = [
        ("max_demand", "Max Demand", "pspstatedemandrequirement.MAX_DEMAND", "maximum value from mapped 30-sec curve column", "max_demand_time"),
        ("peak_demand", "Peak Demand", "pspstatedemandforecast.PEAK_MAX_DEMAND_MET", "configured month-wise peak hour from curve file", "peak_demand_time"),
        ("off_peak_demand", "Off Peak Demand", "pspstatedemandforecast.OFF_PEAK_MAX_DEMAND_MET", "configured off-peak hour from curve file", "off_peak_demand_time"),
        ("min_demand", "Min Demand", "pspstatedemandrequirement.MIN_DEMAND_MET", "minimum value from mapped 30-sec curve column", None),
    ]
    try:
        hourly_demand_data = get_hourly_demand_lookup(db, target_date_str)
        hourly_demand_lookup = hourly_demand_data.get("hourly", {})
        hourly_min_lookup = hourly_demand_data.get("min", {})
        hourly_demand_error = None
    except Exception as exc:
        hourly_demand_lookup = {}
        hourly_min_lookup = {}
        hourly_demand_error = str(exc)
    demand_rows = []
    for state in states:
        state_values = target_values.get(state, {})
        row = {"state": state}
        for metric_key, _, _, _, time_key in demand_metrics:
            operational_value = state_values.get(metric_key)
            curve_value = curve_metrics.get(state, {}).get(metric_key)
            time_value = state_values.get(time_key) if time_key else None
            hour = hourly_boundary_hour(time_value)
            hourly_value = hourly_demand_lookup.get((state, hour)) if hour is not None else None
            if metric_key == "min_demand":
                hourly_value = hourly_min_lookup.get(state)
            row[f"{metric_key}_operational"] = operational_value
            row[f"{metric_key}_time"] = time_value
            row[f"{metric_key}_average_last_7_days"] = average_metric_over_dates(db, last_7_dates, metric_key, state)
            row[f"{metric_key}_same_day_last_week"] = last_week_values.get(state, {}).get(metric_key)
            row[f"{metric_key}_same_day_last_year"] = last_year_values.get(state, {}).get(metric_key)
            row[f"{metric_key}_hourly_api"] = hourly_value
            row[f"{metric_key}_hourly_diff"] = (
                round(to_float(operational_value) - to_float(hourly_value), 3)
                if operational_value is not None and hourly_value is not None
                else None
            )
            if metric_key == "max_demand":
                tolerance_mw = 1.0
                if hour is None:
                    row["max_demand_hourly_status"] = None
                    row["max_demand_hourly_reason"] = f"PSP max demand time {time_value or 'N/A'} is not an hourly boundary, so HOUR_DEMAND cross-check is skipped."
                elif hourly_value is None:
                    row["max_demand_hourly_status"] = False
                    row["max_demand_hourly_reason"] = f"No HOUR_DEMAND found for {state} at {hour:02d}:00."
                else:
                    diff = abs(to_float(operational_value) - to_float(hourly_value))
                    row["max_demand_hourly_status"] = diff <= tolerance_mw
                    row["max_demand_hourly_reason"] = (
                        f"Matched HOUR_DEMAND at {hour:02d}:00 within {tolerance_mw} MW tolerance. PSP {operational_value}, hourly {hourly_value}, diff {round(diff, 3)}."
                        if diff <= tolerance_mw
                        else f"Mismatch at {hour:02d}:00 beyond {tolerance_mw} MW tolerance. PSP {operational_value}, hourly {hourly_value}, diff {round(diff, 3)}."
                    )
            row[f"{metric_key}_curve"] = curve_value
            row[f"{metric_key}_diff"] = (
                round(to_float(operational_value) - to_float(curve_value), 3)
                if operational_value is not None and curve_value is not None
                else None
            )
        demand_rows.append(row)
    tables.append({
        "key": "demand_check",
        "title": "Demand Check",
        "unit": "MW",
        "table_type": "demand",
        "include_curve_columns": True,
        "source_key": "Single demand section. PSP operational sources: MAX_DEMAND, PEAK_MAX_DEMAND_MET, OFF_PEAK_MAX_DEMAND_MET, MIN_DEMAND_MET. Hourly cross-check uses Loadshed API HOUR_DEMAND when PSP time is exactly HH:00. Curve uses configured peak/off-peak hours and mapped 30-sec columns.",
        "hourly_demand_source": "loadshed_api_url / HOUR_DEMAND",
        "hourly_demand_error": hourly_demand_error,
        "logic": [
            {"type": label, "psp_source": psp_source, "curve_source": curve_source}
            for _, label, psp_source, curve_source, _ in demand_metrics
        ],
        "rows": demand_rows,
    })
    shortage_rows = []
    for state in states:
        state_values = target_values.get(state, {})
        shortage_rows.append({
            "state": state,
            "day_shortage": state_values.get("day_shortage"),
            "max_demand_shortage": state_values.get("max_demand_shortage"),
            "peak_shortage": state_values.get("peak_shortage"),
            "off_peak_shortage": state_values.get("off_peak_shortage"),
        })
    tables.append({
        "key": "shortage_check",
        "title": "Shortage Check",
        "unit": "MW / MU",
        "table_type": "shortage",
        "include_curve_columns": False,
        "source_key": "Current day only: DAY_SHORT, MAX_DEMAND_SHORTAGE_RGN, PK_SHORT, OFFPK_SHORT. Values normalized as negative.",
        "rows": shortage_rows,
        "columns": [
            {"key": "state", "label": "Constituent"},
            {"key": "day_shortage", "label": "Day Shortage (MU)", "source": "pspstateloaddetailsER.DAY_SHORT"},
            {"key": "max_demand_shortage", "label": "Max Demand Shortage (MW)", "source": "pspstatedemandrequirement.MAX_DEMAND_SHORTAGE_RGN"},
            {"key": "peak_shortage", "label": "Peak Shortage (MW)", "source": "pspstatedemandforecast.PK_SHORT"},
            {"key": "off_peak_shortage", "label": "Off Peak Shortage (MW)", "source": "pspstatedemandforecast.OFFPK_SHORT"},
        ]
    })
    exchange_doc = db.psp_collection.find_one(
        {"date": target_date_str},
        {"_id": 0, "date": 1, "pspTransnationalExchangeState": 1, "pspTransnationalExchangeLine": 1}
    )
    tables.append(build_transnational_exchange_check(exchange_doc))
    return tables

def get_power_system_base_rows(db, target_date_str: str):
    states = get_power_system_states()
    docs = list(db.db["psp_power_system_base"].find(
        {"effective_date": {"$lte": target_date_str}},
        {"_id": 0}
    ).sort("effective_date", -1))

    latest_by_state = {}
    for doc in docs:
        state = doc.get("state")
        if state and state not in latest_by_state:
            latest_by_state[state] = doc

    rows = []
    for state in states:
        name = state["name"]
        doc = latest_by_state.get(name, {})
        rows.append({
            "state": name,
            "effective_date": doc.get("effective_date", ""),
            "ists_inlet_points": to_float(doc.get("ists_inlet_points")),
            "per_capita_consumption": to_float(doc.get("per_capita_consumption")),
            "state_gna": to_float(doc.get("state_gna")),
        })
    return rows

def get_power_system_installed_capacity(doc):
    capacity = {}
    for item in (doc or {}).get("pspstateentitiesgeneration", []) or []:
        if str(item.get("STATION_TYPE_NAME") or "").strip().upper() == "CPP_IMPORT":
            continue
        state_name = item.get("STATE_NAME")
        if not state_name:
            continue
        capacity[state_name] = capacity.get(state_name, 0.0) + to_float(item.get("INSTALLED_CAPACITY"))
    return capacity

def get_power_system_historical_peaks(db, target_date_str: str):
    result = {}
    for state in get_power_system_states():
        name = state["name"]
        prefix = state["hist_prefix"]
        demand_key = f"{prefix}_MAX_DEMAND"
        demand_time_key = f"{prefix}_MAX_DEMAND_TIME"
        energy_key = f"{prefix}_ENERGY" if prefix != "ORISSA" else "ODISHA_ENERGY"

        demand_doc = db.db["psp_historical"].find_one(
            {"date": {"$lte": target_date_str}, demand_key: {"$exists": True, "$ne": None}},
            {"date": 1, demand_key: 1, demand_time_key: 1},
            sort=[(demand_key, -1)]
        )
        energy_doc = db.db["psp_historical"].find_one(
            {"date": {"$lte": target_date_str}, energy_key: {"$exists": True, "$ne": None}},
            {"date": 1, energy_key: 1},
            sort=[(energy_key, -1)]
        )

        result[name] = {
            "maximum_demand_met": to_float((demand_doc or {}).get(demand_key)),
            "maximum_demand_met_date": (demand_doc or {}).get("date", ""),
            "maximum_demand_met_time": str((demand_doc or {}).get(demand_time_key) or ""),
            "maximum_energy_consumption": to_float((energy_doc or {}).get(energy_key)),
            "maximum_energy_met_date": (energy_doc or {}).get("date", ""),
        }
    return result

def format_loadshed_hour(hour_id):
    try:
        hour = int(hour_id)
    except (TypeError, ValueError):
        return "N/A"
    return f"{hour:02d}:00"

def build_loadshedding_summary(rows):
    state_order = ["BIHAR", "JHARKHAND", "DVC", "ODISHA", "WEST BENGAL", "SIKKIM"]
    by_state = {state: {"state": state, "max_load_shedding": 0.0, "time": "N/A"} for state in state_order}
    hourly_er = {}

    for item in rows or []:
        state = str(item.get("STATE_NAME") or "").strip().upper()
        if state not in by_state:
            continue
        hour_id = item.get("HOUR_ID")
        load_shed = to_float(item.get("HOUR_LOAD_SHEDDING"))

        hourly_er[hour_id] = hourly_er.get(hour_id, 0.0) + load_shed
        if load_shed > by_state[state]["max_load_shedding"]:
            by_state[state]["max_load_shedding"] = load_shed
            by_state[state]["time"] = format_loadshed_hour(hour_id)

    er_max = 0.0
    er_time = "N/A"
    for hour_id, value in hourly_er.items():
        if value > er_max:
            er_max = value
            er_time = format_loadshed_hour(hour_id)

    result = []
    for state in state_order:
        row = by_state[state]
        if row["max_load_shedding"] <= 0:
            row["time"] = "N/A"
        row["max_load_shedding"] = round(row["max_load_shedding"], 2)
        result.append(row)

    result.append({
        "state": "ER",
        "max_load_shedding": round(er_max, 2),
        "time": er_time if er_max > 0 else "N/A"
    })
    return result

def fetch_and_cache_loadshedding(db, target_date: date):
    config = PipelineConfigService().get_config("PSP") or {}
    template = config.get(
        "loadshed_api_url",
        "https://report.erldc.in/posoco_api/api/StgHourlyStateData/GetStgHourlyStateDataNRByMonthNLDC/{date_from}/{date_to}/1"
    )
    date_api = target_date.strftime("%d-%m-%Y")
    url = template.format(date_from=date_api, date_to=date_api)

    session = PSPService._create_session()
    res = session.get(url, verify=False, timeout=30)
    res.raise_for_status()
    rows = res.json()

    date_str = target_date.strftime("%Y-%m-%d")
    db.db["psp_loadshedding"].update_one(
        {"_id": date_str},
        {"$set": {
            "_id": date_str,
            "date": date_str,
            "source_url": url,
            "rows": rows,
            "fetched_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    return rows

def make_legacy_session():
    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    return session

def normalize_outage_row(item):
    element_name = str(item.get("ELEMENT_NAME") or "").strip()
    unit_number = item.get("UNIT_NUMBER")
    installed_capacity = to_float(item.get("INSTALLED_CAPACITY"))
    return {
        "id": item.get("Id"),
        "identity": f"{element_name} / U-{unit_number} / {installed_capacity:g} MW",
        "element_name": element_name,
        "unit_number": item.get("UNIT_NUMBER"),
        "installed_capacity": installed_capacity,
        "sector": item.get("SECTOR", ""),
        "location": item.get("LOCATION", ""),
        "owner_name": item.get("OWNER_NAME", ""),
        "fuel": item.get("FUEL", ""),
        "outage_type": item.get("OUTAGE_TYPE", ""),
        "outage_date": item.get("OUTAGE_DATE", ""),
        "outage_time": item.get("OUTAGE_TIME", ""),
        "revival_date": item.get("REVIVAL_DATE", ""),
        "revival_time": item.get("REVIVAL_TIME", ""),
        "expected_revival_date": item.get("EXPECTED_REVIVAL_DATE", ""),
        "expected_revival_time": item.get("EXPECTED_REVIVAL_TIME", ""),
        "reason": item.get("OUT_REASON", ""),
    }

def outage_compare_key(item):
    return (
        str(item.get("ELEMENT_NAME") or "").strip().upper(),
        str(item.get("UNIT_NUMBER") or "").strip(),
        f"{to_float(item.get('INSTALLED_CAPACITY')):g}",
    )

def fetch_outage_details_for_date(db, target_date: date, session=None):
    date_str = target_date.strftime("%Y-%m-%d")
    cached = db.db["psp_generation_outage_daily"].find_one({"_id": date_str})
    if cached:
        return cached.get("rows", [])

    config = PipelineConfigService().get_config("PSP") or {}
    template = config.get(
        "outage_api_url",
        "https://report.erldc.in/POSOCO_API/api/Outage/GetQueryNpmcReportData/{date}"
    )
    date_api = target_date.strftime("%d-%m-%Y")
    url = template.format(date=date_api)
    session = session or make_legacy_session()
    res = session.get(url, verify=False, timeout=30)
    res.raise_for_status()
    rows = res.json()

    db.db["psp_generation_outage_daily"].update_one(
        {"_id": date_str},
        {"$set": {
            "_id": date_str,
            "date": date_str,
            "source_url": url,
            "rows": rows,
            "fetched_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    return rows

def build_generation_outage_changes(yesterday_rows, day_before_rows):
    yesterday_by_id = {outage_compare_key(item): item for item in yesterday_rows or []}
    day_before_by_id = {outage_compare_key(item): item for item in day_before_rows or []}

    yesterday_set = set(yesterday_by_id)
    day_before_set = set(day_before_by_id)

    tripped = [
        normalize_outage_row(yesterday_by_id[item_id])
        for item_id in sorted(yesterday_set - day_before_set)
    ]
    restored = [
        normalize_outage_row(day_before_by_id[item_id])
        for item_id in sorted(day_before_set - yesterday_set)
    ]

    restored_mw = round(sum(item["installed_capacity"] for item in restored), 2)
    tripped_mw = round(sum(item["installed_capacity"] for item in tripped), 2)
    return {
        "restored": restored,
        "tripped": tripped,
        "restored_mw": restored_mw,
        "tripped_mw": tripped_mw,
        "net_mw": round(restored_mw - tripped_mw, 2),
    }

@router.get("/power-system-base")
async def get_power_system_base(date_str: str = None):
    db = MongoService()
    target_date_str = date_str or date.today().strftime("%Y-%m-%d")
    return {
        "success": True,
        "date": target_date_str,
        "rows": get_power_system_base_rows(db, target_date_str)
    }

@router.put("/power-system-base")
async def save_power_system_base(req: PSPPowerSystemBaseRequest):
    try:
        date.fromisoformat(req.effective_date)
    except ValueError:
        return {"success": False, "message": "Invalid effective_date. Use YYYY-MM-DD."}

    db = MongoService()
    now_str = datetime.utcnow().isoformat()
    valid_states = {state["name"] for state in get_power_system_states()}
    saved = 0

    for row in req.rows:
        if row.state not in valid_states:
            continue
        db.db["psp_power_system_base"].update_one(
            {"_id": f"{req.effective_date}:{row.state}"},
            {"$set": {
                "effective_date": req.effective_date,
                "state": row.state,
                "ists_inlet_points": row.ists_inlet_points,
                "per_capita_consumption": row.per_capita_consumption,
                "state_gna": row.state_gna,
                "updated_at": now_str,
            }},
            upsert=True
        )
        saved += 1

    return {
        "success": True,
        "message": f"Saved {saved} power system base rows for {req.effective_date}."
    }

@router.get("/report-checking")
async def get_psp_report_checking(date_str: str = None, include_curve: bool = False):
    try:
        db = MongoService()
        if not date_str:
            latest_doc = db.psp_collection.find_one(
                {"pspstateloaddetailsER": {"$exists": True, "$ne": []}},
                {"date": 1},
                sort=[("date", -1)]
            )
            date_str = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")
        date.fromisoformat(date_str)

        curve_metrics, curve_meta = get_curve_metrics_from_cache(db, date_str, refresh=include_curve)
        return {
            "success": True,
            "date": date_str,
            "curve": curve_meta,
            "frequency_check": extract_frequency_check(db.psp_collection.find_one(
                {"date": date_str},
                {"_id": 0, "date": 1, "pspFrequencyProfile": 1, "pspFrequencyProfileMaxMin": 1}
            )),
            "tables": build_report_checking_rows(db, date_str, curve_metrics),
            "columns": [
                {"key": "state", "label": "Constituent"},
                {"key": "operational_day", "label": "Operational Day"},
                {"key": "average_last_7_days", "label": "Avg Last 7 Days"},
                {"key": "same_day_last_week", "label": "Same Day Last Week"},
                {"key": "same_day_last_year", "label": "Same Day Last Year"},
                {"key": "curve_file", "label": "Curve File"},
                {"key": "difference_operational_vs_curve", "label": "Operational - Curve"},
            ]
        }
    except ValueError:
        return {"success": False, "message": "Invalid date_str. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.get("/report-checking/frequency-trend")
async def get_psp_frequency_trend(start_date: str = None, end_date: str = None):
    try:
        db = MongoService()
        if not end_date:
            latest_doc = db.psp_collection.find_one(
                {"pspFrequencyProfileMaxMin": {"$exists": True, "$ne": []}},
                {"date": 1},
                sort=[("date", -1)]
            )
            end_date = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (date.fromisoformat(end_date) - timedelta(days=14)).strftime("%Y-%m-%d")
        date.fromisoformat(start_date)
        date.fromisoformat(end_date)
        docs = list(db.psp_collection.find(
            {"date": {"$gte": start_date, "$lte": end_date}},
            {"_id": 0, "date": 1, "pspFrequencyProfileMaxMin": 1}
        ).sort("date", 1))

        rows = []
        for doc in docs:
            freq = extract_frequency_check(doc).get("max_min", {})
            rows.append({
                "date": doc.get("date"),
                "max_freq": freq.get("max_freq"),
                "min_freq": freq.get("min_freq"),
                "average_frequency": freq.get("average_frequency"),
                "max_time": freq.get("max_time"),
                "min_time": freq.get("min_time"),
            })
        return {
            "success": True,
            "start_date": start_date,
            "end_date": end_date,
            "data": rows,
            "series": [
                {"key": "max_freq", "label": "Max Freq", "color": "#DC2626"},
                {"key": "min_freq", "label": "Min Freq", "color": "#2563EB"},
                {"key": "average_frequency", "label": "Average Freq", "color": "#059669"},
            ]
        }
    except ValueError:
        return {"success": False, "message": "Invalid date. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.get("/report-checking/shortage-trend")
async def get_psp_shortage_trend(start_date: str = None, end_date: str = None, state: str = "ER"):
    try:
        db = MongoService()
        if not end_date:
            latest_doc = db.psp_collection.find_one(
                {"$or": [
                    {"pspstateloaddetailsER": {"$exists": True, "$ne": []}},
                    {"pspstatedemandrequirement": {"$exists": True, "$ne": []}},
                    {"pspstatedemandforecast": {"$exists": True, "$ne": []}},
                ]},
                {"date": 1},
                sort=[("date", -1)]
            )
            end_date = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (date.fromisoformat(end_date) - timedelta(days=14)).strftime("%Y-%m-%d")
        date.fromisoformat(start_date)
        date.fromisoformat(end_date)

        state_key = normalize_check_key(state or "ER")
        docs = list(db.psp_collection.find(
            {"date": {"$gte": start_date, "$lte": end_date}},
            {"_id": 0, "date": 1, "pspstateloaddetailsER": 1, "pspstatedemandrequirement": 1, "pspstatedemandforecast": 1}
        ).sort("date", 1))

        rows = []
        for doc in docs:
            values = extract_operational_check_values(doc)
            item = values.get(state_key, {})
            rows.append({
                "date": doc.get("date"),
                "state": state_key,
                "day_shortage": item.get("day_shortage"),
                "max_demand_shortage": item.get("max_demand_shortage"),
                "peak_shortage": item.get("peak_shortage"),
                "off_peak_shortage": item.get("off_peak_shortage"),
            })

        return {
            "success": True,
            "start_date": start_date,
            "end_date": end_date,
            "state": state_key,
            "data": rows,
            "series": [
                {"key": "day_shortage", "label": "Day Shortage (MU)", "color": "#DC2626"},
                {"key": "max_demand_shortage", "label": "Max Demand Shortage (MW)", "color": "#EA580C"},
                {"key": "peak_shortage", "label": "Peak Shortage (MW)", "color": "#7C3AED"},
                {"key": "off_peak_shortage", "label": "Off Peak Shortage (MW)", "color": "#0369A1"},
            ],
            "states": ["BIHAR", "DVC", "JHARKHAND", "ODISHA", "SIKKIM", "WEST BENGAL", "ER"],
        }
    except ValueError:
        return {"success": False, "message": "Invalid date. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.get("/curve-headers")
async def get_psp_curve_headers(date_str: str = None):
    try:
        if not date_str:
            date_str = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        date.fromisoformat(date_str)
        config = get_psp_config_with_curve_defaults()
        curve_dir = str(config.get("curve_file_dir") or "").strip()
        file_name = curve_file_name_for_date(date_str)
        file_path = os.path.join(curve_dir, file_name)
        file_exists, path_error = path_exists_with_timeout(file_path)
        if not file_exists:
            suffix = f" ({path_error})" if path_error else ""
            return {"success": False, "message": f"Curve file not found or unreachable: {file_name}{suffix}", "headers": []}

        payload = {
            "file_path": file_path,
            "file_name": file_name,
            "sheet_name": config.get("curve_sheet_name") or "30SEC",
            "state_cols": expand_excel_columns(config.get("curve_state_columns") or "V:AA"),
            "er_col": config.get("curve_er_column") or "AE",
        }
        code = r'''
import json
import sys
import atexit
import os
import shutil
import tempfile
import time
import zipfile
from openpyxl import load_workbook

def local_workbook_copy(src):
    last_error = None
    for _ in range(3):
        fd, tmp = tempfile.mkstemp(suffix=".xlsm")
        os.close(fd)
        try:
            shutil.copy2(src, tmp)
            with zipfile.ZipFile(tmp, "r") as archive:
                bad_member = archive.testzip()
            if bad_member:
                raise zipfile.BadZipFile(f"Bad workbook member: {bad_member}")
            atexit.register(lambda path=tmp: os.path.exists(path) and os.remove(path))
            return tmp
        except Exception as exc:
            last_error = exc
            try:
                os.remove(tmp)
            except OSError:
                pass
            time.sleep(1)
    raise RuntimeError(f"Curve workbook is unreadable or still being written: {last_error}")

payload = json.loads(sys.stdin.read())
try:
    from python_calamine import load_workbook as load_calamine_workbook
    calamine_wb = load_calamine_workbook(payload["file_path"])
    sheet_name = payload["sheet_name"]
    if sheet_name not in calamine_wb.sheet_names:
        print(json.dumps({"success": False, "message": f"Sheet not found: {sheet_name}", "headers": []}))
        raise SystemExit(0)
    rows = calamine_wb.get_sheet_by_name(sheet_name).to_python()
    def col_to_idx(col):
        idx = 0
        for char in str(col).upper():
            if "A" <= char <= "Z":
                idx = idx * 26 + ord(char) - ord("A") + 1
        return idx - 1
    def best_header(col_idx):
        for row_idx in (4, 5, 3, 2):
            if row_idx >= len(rows) or col_idx >= len(rows[row_idx]):
                continue
            value = str(rows[row_idx][col_idx] or "").strip()
            if value and not value.isdigit() and value.upper() != "X":
                return value.replace(" demand", "").replace(" Demand", "")
        return ""
    headers = []
    for col in [*payload["state_cols"], payload["er_col"]]:
        headers.append({"column": col, "header": best_header(col_to_idx(col)), "is_er": col == payload["er_col"]})
    calamine_wb.close()
    print(json.dumps({"success": True, "headers": headers, "file": payload["file_name"], "sheet": sheet_name, "reader": "python-calamine"}))
    raise SystemExit(0)
except SystemExit:
    raise
except Exception:
    pass

try:
    local_path = local_workbook_copy(payload["file_path"])
    wb = load_workbook(local_path, read_only=True, data_only=True, keep_vba=False)
except Exception as exc:
    print(json.dumps({"success": False, "message": str(exc), "headers": []}))
    raise SystemExit(0)
sheet_name = payload["sheet_name"]
if sheet_name not in wb.sheetnames:
    print(json.dumps({"success": False, "message": f"Sheet not found: {sheet_name}", "headers": []}))
    raise SystemExit(0)
ws = wb[sheet_name]
headers = []
for col in [*payload["state_cols"], payload["er_col"]]:
    headers.append({"column": col, "header": str(ws[f"{col}6"].value or "").strip(), "is_er": col == payload["er_col"]})
wb.close()
print(json.dumps({"success": True, "headers": headers, "file": payload["file_name"], "sheet": sheet_name}))
'''
        completed = subprocess.run(
            [sys.executable, "-c", code],
            input=json.dumps(payload),
            text=True,
            capture_output=True,
            timeout=10,
        )
        if completed.returncode != 0:
            return {"success": False, "message": completed.stderr.strip() or "Could not read curve headers", "headers": []}
        parsed = json.loads(completed.stdout or "{}")
        parsed.update({"path": file_path})
        return parsed
    except subprocess.TimeoutExpired:
        return {"success": False, "message": "Curve header read timed out.", "headers": []}
    except Exception as exc:
        return {"success": False, "message": str(exc), "headers": []}

@router.get("/power-system-data")
async def get_power_system_data(date_str: str = None):
    db = MongoService()
    target_date_str = date_str
    if not target_date_str:
        latest_doc = db.psp_collection.find_one(
            {"pspstateentitiesgeneration": {"$exists": True, "$ne": []}},
            {"date": 1},
            sort=[("date", -1)]
        )
        target_date_str = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")

    doc = db.psp_collection.find_one({"date": target_date_str}, {"_id": 0, "date": 1, "pspstateentitiesgeneration": 1})
    capacities = get_power_system_installed_capacity(doc)
    peaks = get_power_system_historical_peaks(db, target_date_str)
    base_rows = get_power_system_base_rows(db, target_date_str)
    base_by_state = {row["state"]: row for row in base_rows}

    states = get_power_system_states()
    columns = [{"key": state["name"], "label": "Odisha" if state["name"] == "ODISHA" else state["name"].title()} for state in states]
    state_values = {}
    for state in states:
        name = state["name"]
        base = base_by_state.get(name, {})
        state_values[name] = {
            "installed_capacity": round(capacities.get(state["daily_demand_name"], capacities.get(name, 0.0)), 2),
            **peaks.get(name, {}),
            "ists_inlet_points": base.get("ists_inlet_points", 0.0),
            "per_capita_consumption": base.get("per_capita_consumption", 0.0),
            "state_gna": base.get("state_gna", 0.0),
            "base_effective_date": base.get("effective_date", ""),
        }

    metric_rows = [
        {"key": "installed_capacity", "label": "Installed Capacity (MW)", "format": "number0"},
        {"key": "maximum_demand_met", "label": "Maximum demand met (MW)", "format": "number0"},
        {"key": "maximum_demand_met_date", "label": "Maximum demand met Date", "format": "date"},
        {"key": "maximum_energy_consumption", "label": "Maximum Energy Consumption (MW)", "format": "number2"},
        {"key": "maximum_energy_met_date", "label": "Maximum Energy Met Date", "format": "date"},
        {"key": "ists_inlet_points", "label": "ISTS inlet points (Nos)", "format": "number0"},
        {"key": "per_capita_consumption", "label": "Per Capita consumption (kWh)", "format": "number0"},
        {"key": "state_gna", "label": "State GNA", "format": "number0"},
    ]

    return {
        "success": True,
        "has_data": bool(doc),
        "date": target_date_str,
        "columns": columns,
        "rows": metric_rows,
        "values": state_values,
    }

@router.get("/power-system-generating-stations")
async def get_power_system_generating_stations(state: str, date_str: str = None):
    db = MongoService()
    target_date_str = date_str
    if not target_date_str:
        latest_doc = db.psp_collection.find_one(
            {"pspstateentitiesgeneration": {"$exists": True, "$ne": []}},
            {"date": 1},
            sort=[("date", -1)]
        )
        target_date_str = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")

    state_key = str(state or "").strip().upper()
    state_config = next(
        (
            item for item in get_power_system_states()
            if item["name"] == state_key or item["daily_demand_name"] == state_key
        ),
        None
    )
    if not state_config:
        return {"success": False, "message": "Invalid state."}

    doc = db.psp_collection.find_one(
        {"date": target_date_str},
        {"_id": 0, "date": 1, "pspstateentitiesgeneration": 1}
    )
    rows = []
    for item in (doc or {}).get("pspstateentitiesgeneration", []) or []:
        if item.get("STATE_NAME") != state_config["daily_demand_name"]:
            continue
        if str(item.get("STATION_TYPE_NAME") or "").strip().upper() == "CPP_IMPORT":
            continue
        installed_capacity = to_float(item.get("INSTALLED_CAPACITY"))
        rows.append({
            "constituent_name": item.get("CONSTITUENT_NAME", ""),
            "installed_capacity": installed_capacity,
            "station_type": item.get("STATION_TYPE_NAME", ""),
            "classification": item.get("CLASSIFICATION_NAME", ""),
        })

    rows.sort(key=lambda row: row["installed_capacity"], reverse=True)
    return {
        "success": True,
        "date": target_date_str,
        "state": state_config["name"],
        "rows": rows,
        "total_installed_capacity": round(sum(row["installed_capacity"] for row in rows), 2),
    }

@router.get("/loadshedding")
async def get_loadshedding(date_str: str = None, refresh: bool = False):
    db = MongoService()
    if date_str:
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            return {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}
    else:
        latest_doc = db.psp_collection.find_one({"date": {"$exists": True}}, {"date": 1}, sort=[("date", -1)])
        target_date = date.fromisoformat(latest_doc["date"]) if latest_doc else date.today() - timedelta(days=1)

    date_str_db = target_date.strftime("%Y-%m-%d")
    doc = None if refresh else db.db["psp_loadshedding"].find_one({"_id": date_str_db})
    try:
        rows = doc.get("rows", []) if doc else fetch_and_cache_loadshedding(db, target_date)
    except Exception as e:
        if doc:
            rows = doc.get("rows", [])
        else:
            return {"success": False, "message": f"Failed to fetch loadshedding data: {str(e)}"}

    return {
        "success": True,
        "date": date_str_db,
        "rows": build_loadshedding_summary(rows),
        "source": "mongo_cache" if doc and not refresh else "api"
    }

@router.get("/generation-outage-changes")
async def get_generation_outage_changes(date_str: str = None, refresh: bool = False):
    db = MongoService()
    if date_str:
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            return {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}
    else:
        target_date = date.today() - timedelta(days=1)

    day_before = target_date - timedelta(days=1)
    target_str = target_date.strftime("%Y-%m-%d")
    cached = None if refresh else db.db["psp_generation_outage_changes"].find_one({"_id": target_str})
    if cached:
        return {
            "success": True,
            "date": target_str,
            "previous_date": day_before.strftime("%Y-%m-%d"),
            "summary": cached.get("summary", {}),
            "restored": cached.get("restored", []),
            "tripped": cached.get("tripped", []),
            "source": "mongo_cache"
        }

    try:
        session = make_legacy_session()
        target_rows = fetch_outage_details_for_date(db, target_date, session=session)
        previous_rows = fetch_outage_details_for_date(db, day_before, session=session)
        changes = build_generation_outage_changes(target_rows, previous_rows)
    except Exception as e:
        return {"success": False, "message": f"Failed to fetch outage changes: {str(e)}"}

    summary = {
        "restored_mw": changes["restored_mw"],
        "tripped_mw": changes["tripped_mw"],
        "net_mw": changes["net_mw"],
        "restored_count": len(changes["restored"]),
        "tripped_count": len(changes["tripped"]),
    }
    db.db["psp_generation_outage_changes"].update_one(
        {"_id": target_str},
        {"$set": {
            "_id": target_str,
            "date": target_str,
            "previous_date": day_before.strftime("%Y-%m-%d"),
            "summary": summary,
            "restored": changes["restored"],
            "tripped": changes["tripped"],
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    return {
        "success": True,
        "date": target_str,
        "previous_date": day_before.strftime("%Y-%m-%d"),
        "summary": summary,
        "restored": changes["restored"],
        "tripped": changes["tripped"],
        "source": "api"
    }

@router.get("/analytics")
async def get_psp_analytics():
    db = MongoService()
    docs = list(db.psp_collection.find({}, {"_id": 0}).sort("date", 1))
    
    if not docs:
        return {
            "success": True,
            "has_data": False,
            "latest_date": None,
            "state_data": [],
            "trend_data": []
        }
        
    latest_doc = docs[-1]
    latest_date = latest_doc.get("date")
    
    # 1. State-wise Requirement & Availability for the latest date
    state_data = []
    avail_demand_list = latest_doc.get("pspregionalavailibilitydemand", [])
    
    states_list = ['DVC', 'BIHAR', 'ORISSA', 'SIKKIM', 'JHARKHAND', 'WEST BENGAL']
    for item in avail_demand_list:
        state_name = item.get("StateName", "").upper()
        if state_name in states_list:
            req = item.get("RegionalRequirement") or item.get("Requirement") or 0.0
            avail = item.get("RegionalAvailability") or item.get("Availability") or item.get("RegionalAvailibility") or item.get("Availibility") or 0.0
            
            state_data.append({
                "name": state_name,
                "requirement": float(req),
                "availability": float(avail)
            })
            
    # 2. Daily trend of total Region Requirement & Availability (past 15 days)
    trend_data = []
    for doc in docs[-15:]:  # limit to last 15 days
        doc_date = doc.get("date", "")
        region_req = 0.0
        region_avail = 0.0
        
        avail_demand_list_doc = doc.get("pspregionalavailibilitydemand", [])
        
        region_item = next((x for x in avail_demand_list_doc if x.get("StateName", "").upper() == "REGION"), None)
        if region_item:
            region_req = region_item.get("RegionalRequirement") or region_item.get("Requirement") or 0.0
            region_avail = region_item.get("RegionalAvailability") or region_item.get("Availability") or region_item.get("RegionalAvailibility") or region_item.get("Availibility") or 0.0
        else:
            for item in avail_demand_list_doc:
                if item.get("StateName", "").upper() != "REGION":
                    req = item.get("RegionalRequirement") or item.get("Requirement") or 0.0
                    avail = item.get("RegionalAvailability") or item.get("Availability") or item.get("RegionalAvailibility") or item.get("Availibility") or 0.0
                    region_req += float(req)
                    region_avail += float(avail)
                    
        trend_data.append({
          "date": doc_date,
          "requirement": float(region_req),
          "availability": float(region_avail)
        })
        
    return {
        "success": True,
        "has_data": True,
        "latest_date": latest_date,
        "state_data": state_data,
        "trend_data": trend_data
    }

@router.get("/energy-trend")
async def get_energy_trend(start_date: str = None, end_date: str = None):
    """Get date-range ER and state energy trend from historical PSP Mongo data."""
    db = MongoService()

    try:
        if end_date:
            end_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            latest_doc = db.db["psp_historical"].find_one(
                {"ER_ENERGY": {"$exists": True, "$ne": None}},
                {"date": 1},
                sort=[("date", -1)]
            )
            end_obj = datetime.strptime(latest_doc["date"], "%Y-%m-%d").date() if latest_doc else date.today()

        if start_date:
            start_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        else:
            start_obj = end_obj - timedelta(days=6)
    except ValueError:
        return {
            "success": False,
            "message": "Invalid date format. Use YYYY-MM-DD.",
            "rows": [],
            "total": 0
        }

    if start_obj > end_obj:
        start_obj, end_obj = end_obj, start_obj

    query = {
        "date": {
            "$gte": start_obj.strftime("%Y-%m-%d"),
            "$lte": end_obj.strftime("%Y-%m-%d")
        }
    }
    projection = {
        "_id": 0,
        "date": 1,
        "BIHAR_ENERGY": 1,
        "DVC_ENERGY": 1,
        "JHARKHAND_ENERGY": 1,
        "ODISHA_ENERGY": 1,
        "SIKKIM_ENERGY": 1,
        "WEST_BENGAL_ENERGY": 1,
        "ER_ENERGY": 1
    }
    rows_by_date = {}
    for doc in db.db["psp_historical"].find(query, projection).sort("date", 1):
        state_total = sum(
            float(doc.get(key) or 0)
            for key in [
                "BIHAR_ENERGY",
                "DVC_ENERGY",
                "JHARKHAND_ENERGY",
                "ODISHA_ENERGY",
                "SIKKIM_ENERGY",
                "WEST_BENGAL_ENERGY"
            ]
        )
        rows_by_date[doc.get("date")] = {
            "date": doc.get("date"),
            "bihar": round(float(doc.get("BIHAR_ENERGY") or 0), 2),
            "dvc": round(float(doc.get("DVC_ENERGY") or 0), 2),
            "jharkhand": round(float(doc.get("JHARKHAND_ENERGY") or 0), 2),
            "odisha": round(float(doc.get("ODISHA_ENERGY") or 0), 2),
            "sikkim": round(float(doc.get("SIKKIM_ENERGY") or 0), 2),
            "west_bengal": round(float(doc.get("WEST_BENGAL_ENERGY") or 0), 2),
            "er": round(float(doc.get("ER_ENERGY") or state_total), 2),
            "source": "psp_historical"
        }

    daily_projection = {"_id": 0, "date": 1, "pspstateloaddetailsER": 1}
    daily_docs = list(db.psp_collection.find(query, daily_projection).sort("date", 1))
    daily_name_map = {
        "BIHAR": "bihar",
        "DVC": "dvc",
        "JHARKHAND": "jharkhand",
        "ODISHA": "odisha",
        "ORISSA": "odisha",
        "SIKKIM": "sikkim",
        "WEST BENGAL": "west_bengal"
    }
    for doc in daily_docs:
        doc_date = doc.get("date")
        if doc_date in rows_by_date:
            continue
        load_rows = doc.get("pspstateloaddetailsER", []) or []
        if not load_rows:
            continue

        row = {
            "date": doc_date,
            "bihar": 0.0,
            "dvc": 0.0,
            "jharkhand": 0.0,
            "odisha": 0.0,
            "sikkim": 0.0,
            "west_bengal": 0.0,
            "er": 0.0,
            "source": "psp_collection.pspstateloaddetailsER"
        }
        for item in load_rows:
            state_key = daily_name_map.get(str(item.get("STATE_NAME") or "").strip().upper())
            energy = float(item.get("CONSUMPTION") or 0)
            row["er"] += energy
            if state_key:
                row[state_key] += energy

        row.update({key: round(value, 2) for key, value in row.items() if isinstance(value, float)})
        rows_by_date[doc_date] = row

    rows = [rows_by_date[key] for key in sorted(rows_by_date)]

    return {
        "success": True,
        "has_data": len(rows) > 0,
        "source": "psp_historical + psp_collection.pspstateloaddetailsER",
        "start_date": start_obj.strftime("%Y-%m-%d"),
        "end_date": end_obj.strftime("%Y-%m-%d"),
        "rows": rows,
        "total": len(rows)
    }

@router.get("/energy-consumption")
async def get_energy_consumption(date_str: str = None):
    """Get state-wise energy consumption from pspstateloaddetailsER for donut chart."""
    db = MongoService()
    
    if date_str:
        doc = db.psp_collection.find_one({"date": date_str}, {"_id": 0})
    else:
        # Get latest date that has actual state load data
        doc = db.psp_collection.find_one(
            {"pspstateloaddetailsER": {"$exists": True, "$ne": []}},
            {"_id": 0},
            sort=[("date", -1)]
        )
    
    if not doc:
        return {"success": True, "has_data": False, "states": [], "total": 0, "date": None}
    
    state_load_list = doc.get("pspstateloaddetailsER", [])
    
    states = []
    total_consumption = 0.0
    
    for item in state_load_list:
        state_name = item.get("STATE_NAME", "")
        consumption = float(item.get("CONSUMPTION", 0) or 0)
        total_consumption += consumption
        
        states.append({
            "name": state_name,
            "consumption": round(consumption, 2),
            "thermal": float(item.get("THERMAL", 0) or 0),
            "hydro": float(item.get("HYDRO", 0) or 0),
            "solar": float(item.get("SOLAR", 0) or 0),
            "renewable": float(item.get("RENEWABLE", 0) or 0),
        })
    
    # Sort by consumption descending
    states.sort(key=lambda x: x["consumption"], reverse=True)
    
    return {
        "success": True,
        "has_data": True,
        "date": doc.get("date"),
        "states": states,
        "total": round(total_consumption, 2)
    }

@router.get("/energy-breakdown")
async def get_energy_breakdown(date_str: str = None):
    """Get state-wise energy breakdown of the 11 components for the stacked bar chart."""
    db = MongoService()
    
    if date_str:
        doc = db.psp_collection.find_one({"date": date_str}, {"_id": 0})
    else:
        # Get latest date that has both keys
        doc = db.psp_collection.find_one(
            {"pspstateloaddetailsER": {"$exists": True, "$ne": []},
             "pspSTOADetails1": {"$exists": True, "$ne": []}},
            {"_id": 0},
            sort=[("date", -1)]
        )
        
    if not doc:
        return {"success": True, "has_data": False, "states": [], "er": {}, "date": None}
        
    load_list = doc.get("pspstateloaddetailsER", [])
    stoa_list = doc.get("pspSTOADetails1", [])
    
    # Map by state name
    load_map = {item.get("STATE_NAME", "").upper(): item for item in load_list if item.get("STATE_NAME")}
    stoa_map = {item.get("STATE_NAME", "").upper(): item for item in stoa_list if item.get("STATE_NAME")}
    
    # All unique state names that exist in both load and stoa lists
    all_states = sorted(list(set(load_map.keys()) & set(stoa_map.keys())))
    
    states_data = []
    er_fields = {
        "gna": 0.0, "tgna": 0.0, "rtm": 0.0, "dam": 0.0,
        "thermal": 0.0, "hydro": 0.0, "solar": 0.0, "wind": 0.0,
        "small_hydro": 0.0, "others": 0.0, "ui": 0.0, "consumption": 0.0
    }
    
    for state_name in all_states:
        ld = load_map[state_name]
        st = stoa_map[state_name]
        
        # Parse tags
        gna = float(st.get("DE_ISGS") or 0.0)
        tgna = float(st.get("DE_BILT") or 0.0)
        rtm = float(st.get("DE_RTM_IEX") or 0.0) + float(st.get("DE_RTM_PXI") or 0.0)
        dam = float(st.get("DE_GDAM_PXI") or 0.0) + float(st.get("DE_HPDAM_PXI") or 0.0) + float(st.get("DE_PX") or 0.0)
        
        thermal = float(ld.get("THERMAL") or 0.0)
        hydro = float(ld.get("HYDRO") or 0.0)
        solar = float(ld.get("SOLAR") or 0.0)
        wind = float(ld.get("WIND") or 0.0)
        small_hydro = float(ld.get("SMALL_HYDRO") or 0.0)
        others = float(ld.get("OTHERS") or 0.0)
        ui = float(ld.get("UI") or 0.0)
        
        consumption = float(ld.get("CONSUMPTION") or 0.0)
        
        state_breakdown = {
            "state": state_name,
            "gna": round(gna, 3),
            "tgna": round(tgna, 3),
            "rtm": round(rtm, 3),
            "dam": round(dam, 3),
            "thermal": round(thermal, 3),
            "hydro": round(hydro, 3),
            "solar": round(solar, 3),
            "wind": round(wind, 3),
            "small_hydro": round(small_hydro, 3),
            "others": round(others, 3),
            "ui": round(ui, 3),
            "consumption": round(consumption, 3)
        }
        states_data.append(state_breakdown)
        
        # Add to ER
        er_fields["gna"] += gna
        er_fields["tgna"] += tgna
        er_fields["rtm"] += rtm
        er_fields["dam"] += dam
        er_fields["thermal"] += thermal
        er_fields["hydro"] += hydro
        er_fields["solar"] += solar
        er_fields["wind"] += wind
        er_fields["small_hydro"] += small_hydro
        er_fields["others"] += others
        er_fields["ui"] += ui
        er_fields["consumption"] += consumption
        
    er_data = {
        "state": "ER",
        "gna": round(er_fields["gna"], 3),
        "tgna": round(er_fields["tgna"], 3),
        "rtm": round(er_fields["rtm"], 3),
        "dam": round(er_fields["dam"], 3),
        "thermal": round(er_fields["thermal"], 3),
        "hydro": round(er_fields["hydro"], 3),
        "solar": round(er_fields["solar"], 3),
        "wind": round(er_fields["wind"], 3),
        "small_hydro": round(er_fields["small_hydro"], 3),
        "others": round(er_fields["others"], 3),
        "ui": round(er_fields["ui"], 3),
        "consumption": round(er_fields["consumption"], 3)
    }
    
    return {
        "success": True,
        "has_data": True,
        "date": doc.get("date"),
        "states": states_data,
        "er": er_data
    }

@router.get("/state-generation-sources")
async def get_state_generation_sources(date_str: str = None):
    """Get state-wise PSP generation source mix from pspstateloaddetailsER."""
    db = MongoService()

    projection = {"_id": 0, "date": 1, "pspstateloaddetailsER": 1}
    if date_str:
        doc = db.psp_collection.find_one({"date": date_str}, projection)
    else:
        doc = db.psp_collection.find_one(
            {"pspstateloaddetailsER": {"$exists": True, "$ne": []}},
            projection,
            sort=[("date", -1)]
        )

    if not doc:
        return {
            "success": True,
            "has_data": False,
            "date": None,
            "states": [],
            "composition": [],
            "totals": {},
        }

    def to_float(value):
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    source_fields = [
        ("thermal", "THERMAL", "Thermal"),
        ("hydro", "HYDRO", "Hydro"),
        ("wind", "WIND", "Wind"),
        ("solar", "SOLAR", "Solar"),
        ("small_hydro", "SMALL_HYDRO", "Small Hydro"),
        ("others", "OTHERS", "Others"),
    ]

    states = []
    totals = {
        "generation_total": 0.0,
        "drawing_schedule": 0.0,
        "ui": 0.0,
        "consumption": 0.0,
    }
    source_totals = {key: 0.0 for key, _, _ in source_fields}

    for item in doc.get("pspstateloaddetailsER", []) or []:
        state_name = str(item.get("STATE_NAME") or "").strip()
        if not state_name:
            continue

        row = {"state": state_name}
        generation_total = 0.0
        generation_abs_total = 0.0

        for key, raw_key, _ in source_fields:
            value = to_float(item.get(raw_key))
            row[key] = round(value, 3)
            source_totals[key] += value
            generation_total += value
            generation_abs_total += abs(value)

        row["generation_total"] = round(generation_total, 3)
        row["generation_abs_total"] = round(generation_abs_total, 3)
        row["drawing_schedule"] = round(to_float(item.get("DRAWAL_SCHDULE")), 3)
        row["ui"] = round(to_float(item.get("UI")), 3)
        row["consumption"] = round(to_float(item.get("CONSUMPTION")), 3)
        states.append(row)

        totals["generation_total"] += generation_total
        totals["drawing_schedule"] += row["drawing_schedule"]
        totals["ui"] += row["ui"]
        totals["consumption"] += row["consumption"]

    states.sort(key=lambda item: abs(item.get("generation_total", 0.0)), reverse=True)

    composition = []
    composition_denominator = sum(abs(value) for value in source_totals.values())
    for key, _, label in source_fields:
        value = source_totals[key]
        composition.append({
            "key": key,
            "name": label,
            "value": round(value, 3),
            "abs_value": round(abs(value), 3),
            "percent": round((abs(value) / composition_denominator * 100), 2) if composition_denominator > 0 else 0.0,
        })

    return {
        "success": True,
        "has_data": True,
        "date": doc.get("date"),
        "states": states,
        "composition": composition,
        "totals": {key: round(value, 3) for key, value in totals.items()},
    }

def compute_state_portfolio(state: dict, target_date: date, query_time: str, max_demand: float, wbes_data: list, scada_df: pd.DataFrame = None):
    # Parse hour/minute
    try:
        hour = int(query_time.split(":")[0])
        minute = int(query_time.split(":")[1])
    except Exception:
        hour = 19
        minute = 0
        
    # SCADA Gen
    thermal = 0.0
    hydro = 0.0
    solar = 0.0
    biogas = 0.0  # formerly others
    nuclear = 0.0
    total_scada_gen = 0.0
    own_gen_source = "SCADA_FILE"
    
    if scada_df is not None:
        row_idx = hour * 60 + minute + 1
        if row_idx < len(scada_df):
            cols = state["scada"]
            # 1. Thermal
            if cols.get("thermal") and cols["thermal"] in scada_df.columns:
                thermal = float(scada_df[cols["thermal"]][row_idx] or 0)
            
            # 2. Hydro
            if cols.get("hydro") and cols["hydro"] in scada_df.columns:
                hydro = float(scada_df[cols["hydro"]][row_idx] or 0)
            
            # 3. Solar (with fallback check for suffix Int Gen "Solar" or Int Gen Solar)
            solar_keys = []
            if cols.get("solar"):
                solar_keys.append(cols["solar"])
            highlight_name = state.get("highlight", state["name"])
            solar_keys.append(f'{highlight_name} Int Gen "Solar"')
            solar_keys.append(f'{highlight_name} Int Gen Solar')
            solar_keys.append(f'{state["name"]} SOLAR')
            solar_keys.append(f'{state["name"]} Int Gen Solar')
            solar_keys.append(f'{state["name"]} Int Gen "Solar"')
            
            for sk in solar_keys:
                if sk in scada_df.columns:
                    solar = float(scada_df[sk][row_idx] or 0)
                    break
            
            # 4. Others (Biogas)
            if cols.get("others") and cols["others"] in scada_df.columns:
                biogas = float(scada_df[cols["others"]][row_idx] or 0)
            
            # 5. Nuclear
            if cols.get("nuclear") and cols["nuclear"] in scada_df.columns:
                nuclear = float(scada_df[cols["nuclear"]][row_idx] or 0)
                
            # 6. Total SCADA Gen
            scada_gen_col = state.get("scada_gen")
            if scada_gen_col and scada_gen_col in scada_df.columns:
                total_scada_gen = float(scada_df[scada_gen_col][row_idx] or 0)
            else:
                total_scada_gen = thermal + hydro + solar + biogas + nuclear
                own_gen_source = "SCADA_COMPONENT_SUM"
    else:
        # Fallback using mock proportions
        mp = state["mock_p"]
        thermal = max_demand * mp["thermal"]
        hydro = max_demand * mp["hydro"]
        solar = max_demand * mp["solar"]
        biogas = 0.0
        nuclear = 0.0
        total_scada_gen = thermal + hydro + solar + biogas + nuclear
        own_gen_source = "MOCK_FALLBACK_SCADA_UNAVAILABLE"
        
    # WBES Schedules
    isgs = 0.0
    gna = 0.0
    tgna = 0.0
    idam = 0.0
    rtm = 0.0
    
    wbes_candidates = state.get("wbes_candidates") or [state.get("wbes")]
    wbes_candidates = [str(x).strip().upper() for x in wbes_candidates if x]
    matched_wbes_acronym = ""
    state_schedule = next(
        (
            x for x in wbes_data
            if str(x.get("Acronym") or "").strip().upper() in wbes_candidates
        ),
        None
    )
    if state_schedule:
        matched_wbes_acronym = str(state_schedule.get("Acronym") or "")
        net_schd_list = state_schedule.get("NetScheduleSummary", {}).get("NetSchdDataList", [])
        block_idx = hour * 4 + minute // 15
        
        # Helper to parse wbes list safely
        def get_wbes_val(filter_fn):
            for item in net_schd_list:
                if filter_fn(item):
                    amounts = item.get('NetSchdAmount', [])
                    if len(amounts) > block_idx:
                        return float(amounts[block_idx] or 0)
            return 0.0
            
        isgs = (
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'ISGS') +
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'URS')
        )
        gna = (
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_GNA') +
            get_wbes_val(lambda x: x.get('EnergyScheduleSubTypeName') == 'OA_GNA')
        )
        tgna = (
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_TGNA') +
            get_wbes_val(lambda x: x.get('EnergyScheduleSubTypeName') == 'OA_TGNA')
        )
        idam = (
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_PX' and x.get('PXTransactionTypeName') == 'DAM') +
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_PX' and x.get('PXTransactionTypeName') == 'GDAM') +
            get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_PX' and x.get('PXTransactionTypeName') == 'HPDAM')
        )
        rtm = get_wbes_val(lambda x: x.get('EnergyScheduleTypeName') == 'OA_PX' and x.get('PXTransactionTypeName') == 'RTM')
        
    own_gen_for_dsm = total_scada_gen if total_scada_gen else (thermal + hydro + solar + biogas + nuclear)

    # DSM calculation (deviation)
    # Deviation = Demand - (Own SCADA Gen + ISGS + GNA + TGNA + iDAM + RTM)
    dsm = max_demand - (own_gen_for_dsm + isgs + gna + tgna + idam + rtm)
    
    return {
        "thermal": round(thermal, 1),
        "hydro": round(hydro, 1),
        "solar": round(solar, 1),
        "biogas": round(biogas, 1),
        "nuclear": round(nuclear, 1),
        "isgs": round(isgs, 1),
        "gna": round(gna, 1),
        "tgna": round(tgna, 1),
        "idam": round(idam, 1),
        "rtm": round(rtm, 1),
        "own_gen": round(own_gen_for_dsm, 1),
        "own_gen_source": own_gen_source,
        "wbes_acronym": matched_wbes_acronym,
        "wbes_source": "WBES_API" if state_schedule else "WBES_NOT_FOUND",
        "dsm": round(dsm, 1)
    }

def serialize_psp_portfolio_mapping_state(state):
    scada = state.get("scada", {}) or {}
    return {
        "name": state.get("name", ""),
        "psp": state.get("psp", ""),
        "wbes": state.get("wbes", ""),
        "scada_thermal": scada.get("thermal") or "",
        "scada_hydro": scada.get("hydro") or "",
        "scada_solar": scada.get("solar") or "",
        "scada_others": scada.get("others") or "",
        "scada_nuclear": scada.get("nuclear") or "",
        "scada_gen": state.get("scada_gen") or "",
        "highlight": state.get("highlight") or "",
        "curve_column": state.get("curve_column") or "",
        "curve_header": state.get("curve_header") or "",
    }

def load_psp_portfolio_base_states(db):
    overrides = {
        doc.get("_id"): doc
        for doc in db.db["psp_portfolio_mapping"].find({}, {"_id": 1, "mapping": 1})
        if doc.get("_id")
    }

    states = []
    for base in PSP_PORTFOLIO_STATE_CONFIGS:
        state = {
            **base,
            "scada": {**(base.get("scada") or {})},
            "mock_p": {**(base.get("mock_p") or {})},
        }
        override = (overrides.get(state["name"]) or {}).get("mapping") or {}
        if override:
            state["psp"] = override.get("psp") or state["psp"]
            state["wbes"] = override.get("wbes") or state["wbes"]
            state["scada"]["thermal"] = override.get("scada_thermal") or None
            state["scada"]["hydro"] = override.get("scada_hydro") or None
            state["scada"]["solar"] = override.get("scada_solar") or None
            state["scada"]["others"] = override.get("scada_others") or None
            state["scada"]["nuclear"] = override.get("scada_nuclear") or None
            state["scada_gen"] = override.get("scada_gen") or None
            state["highlight"] = override.get("highlight") or state.get("highlight")
            state["curve_column"] = override.get("curve_column") or state.get("curve_column")
            state["curve_header"] = override.get("curve_header") or state.get("curve_header")
        states.append(state)

    return states

def build_psp_portfolio_states(db):
    mapping_rows = list(db.db["frequency_mapping"].find(
        {
            "$or": [
                {"type": "State"},
                {"plant_id": {"$regex": "^STATE_", "$options": "i"}},
                {"rtg_plant_id": {"$regex": "^STATE_", "$options": "i"}},
            ]
        },
        {
            "_id": 0,
            "plant_id": 1,
            "plant_name": 1,
            "STAGE_NAME": 1,
            "rtg_plant_id": 1,
            "wbes_name": 1,
            "wbes_acronym": 1,
        }
    ))

    def norm(value):
        return str(value or "").strip().upper().replace("_", " ")

    states = []
    for base in load_psp_portfolio_base_states(db):
        state = {**base}
        candidates = {state["wbes"]}
        aliases = {
            norm(state["name"]),
            norm(state["psp"]),
            f"STATE {norm(state['name'])}",
            f"STATE {norm(state['psp'])}",
        }

        for row in mapping_rows:
            row_tokens = {
                norm(row.get("plant_id")),
                norm(row.get("rtg_plant_id")),
                norm(row.get("plant_name")),
                norm(row.get("STAGE_NAME")),
            }
            if aliases.intersection(row_tokens):
                for key in ("wbes_acronym", "wbes_name"):
                    value = str(row.get(key) or "").strip()
                    if value:
                        candidates.add(value)

        state["wbes_candidates"] = sorted(candidates)
        states.append(state)

    return states

def calculate_and_save_portfolio(target_date: date):
    db = MongoService()
    date_str_db = target_date.strftime("%Y-%m-%d")
    
    # 1. Load data sources (psp_data, psp_historical, WBES, SCADA)
    # psp_data
    psp_doc = db.psp_collection.find_one({"date": date_str_db})
    max_demand_list = psp_doc.get("pspstatedemandrequirement", []) if psp_doc else []
    
    # psp_historical
    hist_doc = db.db["psp_historical"].find_one({"date": date_str_db})
    
    STATES = build_psp_portfolio_states(db)

    # wbes
    config_service = PipelineConfigService()
    config = config_service.get_config("PSP") or {}
    wbes_url = config.get("wbes_url", "https://gateway.grid-india.in/POSOCO/reports/1.0/WebAccessAPI/GetUtilityExternalSharedData")
    wbes_api_key = config.get("wbes_api_key", "6dc32d47-f46a-45c9-afaf-c6ce73c4ca71")
    wbes_username = config.get("wbes_username", "erldc_internal_prod")
    wbes_password = config.get("wbes_password", "ErldcPr0d@052024")
    
    wbes_data = []
    try:
        url = f"{wbes_url}?apikey={wbes_api_key}"
        util_acronyms = sorted({
            acronym
            for state in STATES
            for acronym in state.get("wbes_candidates", [])
            if acronym
        })
        payload = {
            "Date": target_date.strftime("%d-%m-%Y"),
            "SchdRevNo": -1,
            "UserName": wbes_username,
            "UtilAcronymList": util_acronyms,
            "UtilRegionIdList": [1]
        }
        res = requests.post(url, json=payload, auth=(wbes_username, wbes_password), timeout=10)
        if res.status_code == 200:
            wbes_data = res.json().get('ResponseBody', {}).get('GroupWiseDataList', [])
    except Exception as e:
        print(f"Error fetching WBES data dynamically in save portfolio: {e}")
        
    # scada
    is_smb_online = False
    try:
        import socket
        socket.setdefaulttimeout(1.0)
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.connect(("10.3.95.200", 445))
        is_smb_online = True
    except Exception:
        pass
        
    scada_df = None
    if is_smb_online:
        scada_path = '//10.3.95.200/HTTP-Access/ScadaData/er_web/ER_THERMAL_GEN_{}.xlsx'.format(target_date.strftime("%d%m%Y"))
        if os.path.exists(scada_path):
            try:
                scada_df = pd.read_excel(scada_path, sheet_name="DATA", header=[1])
            except Exception as e:
                print(f"Error reading SCADA excel: {e}")
                
    # 2. Compute each state's peak portfolio
    states_data = {}
    sum_state_max_demand = 0.0
    
    for state in STATES:
        state_label = state["name"]
        psp_name = state["psp"]
        hist_prefix = state["historical_prefix"]
        
        max_demand = 0.0
        max_demand_time = "19:00"
        
        if hist_doc:
            max_demand = hist_doc.get(f"{hist_prefix}_MAX_DEMAND") or 0.0
            max_demand_time = hist_doc.get(f"{hist_prefix}_MAX_DEMAND_TIME") or "19:00"
            
        if not max_demand or max_demand == 0.0:
            match = next((x for x in max_demand_list if x.get("STATE_NAME") == psp_name), None)
            if match:
                max_demand = float(match.get("MAX_DEMAND") or 0.0)
                max_demand_time = match.get("MAX_DEMAND_TIME") or "19:00"
                
        if not max_demand or max_demand == 0.0:
            mock_scales = {
                "BIHAR": 6500.0, "JHARKHAND": 1500.0, "DVC": 3000.0,
                "ORISSA": 4800.0, "WEST BENGAL": 6800.0, "SIKKIM": 120.0
            }
            max_demand = mock_scales.get(psp_name, 1000.0)
            max_demand_time = "19:30"
            
        sum_state_max_demand += max_demand
        
        # Loadshed
        loadshed = 0.0
        match = next((x for x in max_demand_list if x.get("STATE_NAME") == psp_name), None)
        if match:
            loadshed = abs(float(match.get("MAX_DEMAND_SHORTAGE") or 0.0))
            
        portfolio = compute_state_portfolio(state, target_date, max_demand_time, max_demand, wbes_data, scada_df)
        
        states_data[state_label] = {
            "max_demand": round(max_demand, 1),
            "time": max_demand_time,
            "loadshed": round(loadshed, 1),
            "portfolio": portfolio
        }
        
    # 3. Compute ER Regional portfolio at regional peak time
    er_max_demand = 0.0
    er_max_demand_time = "19:00"
    if hist_doc:
        er_max_demand = hist_doc.get("ER_MAX_DEMAND") or 0.0
        er_max_demand_time = hist_doc.get("ER_MAX_DEMAND_TIME") or "19:00"
        
    if not er_max_demand or er_max_demand == 0.0:
        match = next((x for x in max_demand_list if x.get("STATE_NAME").upper() == "REGION"), None)
        if match:
            er_max_demand = float(match.get("MAX_DEMAND") or 0.0)
            er_max_demand_time = match.get("MAX_DEMAND_TIME") or "19:00"
            
    if not er_max_demand or er_max_demand == 0.0:
        er_max_demand = sum_state_max_demand if sum_state_max_demand > 0 else 20000.0
        er_max_demand_time = "19:30"
        
    scale_factor = er_max_demand / (sum_state_max_demand if sum_state_max_demand > 0 else er_max_demand)
    
    er_portfolio = {
        "thermal": 0.0, "hydro": 0.0, "solar": 0.0, "biogas": 0.0, "nuclear": 0.0,
        "isgs": 0.0, "gna": 0.0, "tgna": 0.0, "idam": 0.0, "rtm": 0.0,
        "own_gen": 0.0, "dsm": 0.0
    }
    
    for state in STATES:
        state_demand_at_er = states_data[state["name"]]["max_demand"] * scale_factor
        p = compute_state_portfolio(state, target_date, er_max_demand_time, state_demand_at_er, wbes_data, scada_df)
        for k in er_portfolio:
            er_portfolio[k] += p[k]
            
    # Round all values in er_portfolio
    for k in er_portfolio:
        er_portfolio[k] = round(er_portfolio[k], 1)
        
    # Loadshed for ER
    er_loadshed = 0.0
    match = next((x for x in max_demand_list if x.get("STATE_NAME").upper() == "REGION"), None)
    if match:
        er_loadshed = abs(float(match.get("MAX_DEMAND_SHORTAGE") or 0.0))
        
    er_data = {
        "max_demand": round(er_max_demand, 1),
        "time": er_max_demand_time,
        "loadshed": round(er_loadshed, 1),
        "portfolio": er_portfolio
    }
    
    # Save to psp_portfolio_breakdown collection
    result_doc = {
        "_id": date_str_db,
        "date": date_str_db,
        "states": states_data,
        "ER": er_data,
        "source_version": PSP_PORTFOLIO_SOURCE_VERSION,
        "updated_at": datetime.utcnow().isoformat()
    }
    
    db.db["psp_portfolio_breakdown"].update_one(
        {"_id": date_str_db},
        {"$set": result_doc},
        upsert=True
    )
    
    return result_doc

def extract_psp_daily_values(doc_hist, doc_daily):
    if doc_hist:
        res = {}
        for state in PSP_PEAK_STATES:
            name = state["name"]
            prefix = state["hist_prefix"]
            demand = doc_hist.get(f"{prefix}_MAX_DEMAND")
            demand_time = doc_hist.get(f"{prefix}_MAX_DEMAND_TIME") or "19:00"
            energy_key = f"{prefix}_ENERGY" if prefix != "ORISSA" else "ODISHA_ENERGY"
            energy = doc_hist.get(energy_key)

            res[name] = {
                "demand": float(demand) if demand is not None else 0.0,
                "demand_time": str(demand_time),
                "energy": float(energy) if energy is not None else 0.0
            }
        return res

    if doc_daily:
        demand_list = doc_daily.get("pspstatedemandrequirement", [])
        demand_map = {item.get("STATE_NAME"): item for item in demand_list if item.get("STATE_NAME")}

        load_list = doc_daily.get("pspstateloaddetailsER", [])
        load_map = {item.get("STATE_NAME"): item for item in load_list if item.get("STATE_NAME")}
        er_energy = sum(float(item.get("CONSUMPTION", 0) or 0) for item in load_list)

        res = {}
        for state in PSP_PEAK_STATES:
            name = state["name"]
            match_dem = demand_map.get(state["daily_demand_name"])
            demand = float(match_dem.get("MAX_DEMAND") or 0.0) if match_dem else 0.0
            demand_time = match_dem.get("MAX_DEMAND_TIME") or "19:00" if match_dem else "19:00"

            if name == "ER":
                energy = er_energy
            else:
                match_en = load_map.get(state["daily_energy_name"])
                energy = float(match_en.get("CONSUMPTION") or 0.0) if match_en else 0.0

            res[name] = {
                "demand": demand,
                "demand_time": demand_time,
                "energy": energy
            }
        return res

    return None

def get_psp_daily_values(db, date_str: str):
    hist_doc = db.db["psp_historical"].find_one({"date": date_str})
    daily_doc = db.psp_collection.find_one({"date": date_str})
    return extract_psp_daily_values(hist_doc, daily_doc)

def portfolio_cache_is_stale(db, date_str: str, portfolio_doc: dict) -> bool:
    if portfolio_doc.get("source_version") != PSP_PORTFOLIO_SOURCE_VERSION:
        return True

    portfolio_checks = [(
        portfolio_doc.get("ER", {}).get("portfolio", {}),
        portfolio_doc.get("ER", {}).get("max_demand"),
    )]
    portfolio_checks.extend(
        ((state_doc or {}).get("portfolio", {}), (state_doc or {}).get("max_demand"))
        for state_doc in (portfolio_doc.get("states", {}) or {}).values()
    )
    for portfolio, demand in portfolio_checks:
        if not portfolio:
            return True
        if float(demand or 0) > 0 and float(portfolio.get("own_gen") or 0) <= 0:
            return True

    daily_values = get_psp_daily_values(db, date_str)
    if not daily_values:
        return False

    checks = []
    er_demand = portfolio_doc.get("ER", {}).get("max_demand")
    checks.append(("ER", er_demand, daily_values.get("ER", {}).get("demand")))

    state_docs = portfolio_doc.get("states", {})
    for state in PSP_PEAK_STATES:
        name = state["name"]
        if name == "ER":
            continue
        cached = state_docs.get(name, {}).get("max_demand")
        raw = daily_values.get(name, {}).get("demand")
        checks.append((name, cached, raw))

    for _, cached, raw in checks:
        if raw is None or raw == 0:
            continue
        if cached is None:
            return True
        if abs(float(cached) - float(raw)) > 1.0:
            return True

    return False

def get_valid_portfolio_doc(db, target_date: date):
    date_str = target_date.strftime("%Y-%m-%d")
    portfolio_doc = db.db["psp_portfolio_breakdown"].find_one({"_id": date_str})
    if not portfolio_doc or portfolio_cache_is_stale(db, date_str, portfolio_doc):
        portfolio_doc = calculate_and_save_portfolio(target_date)
    return portfolio_doc

def get_highest_portfolio_doc(db, peak_date: str):
    portfolio_doc = db.db["psp_portfolio_breakdown"].find_one({"_id": peak_date})
    if portfolio_doc and not portfolio_cache_is_stale(db, peak_date, portfolio_doc):
        return portfolio_doc

    try:
        return calculate_and_save_portfolio(date.fromisoformat(peak_date))
    except Exception as exc:
        print(f"Unable to rebuild highest portfolio for {peak_date}: {exc}")

    return portfolio_doc

def rebuild_highest_records_from_psp_historical():
    db = MongoService()

    records = []
    now_str = datetime.utcnow().isoformat()
    for state in PSP_PEAK_STATES:
        name = state["name"]
        prefix = state["hist_prefix"]
        demand_key = f"{prefix}_MAX_DEMAND"
        time_key = f"{prefix}_MAX_DEMAND_TIME"

        hist_docs = list(db.db["psp_historical"].find(
            {demand_key: {"$exists": True, "$ne": None}},
            {"date": 1, demand_key: 1, time_key: 1}
        ).sort(demand_key, -1).limit(2))

        if not hist_docs:
            continue

        peak = hist_docs[0]
        peak_date = peak.get("date")
        breakdown_doc = get_highest_portfolio_doc(db, peak_date) or {}
        if name == "ER":
            portfolio = breakdown_doc.get("ER", {}).get("portfolio", {})
        else:
            portfolio = breakdown_doc.get("states", {}).get(name, {}).get("portfolio", {})

        previous = None
        if len(hist_docs) > 1:
            prev = hist_docs[1]
            prev_date = prev.get("date")
            prev_doc = get_highest_portfolio_doc(db, prev_date) or {}
            previous = {
                "date": prev_date,
                "max_demand": float(prev.get(demand_key) or 0),
                "max_demand_time": str(prev.get(time_key) or "19:00"),
                "portfolio": (
                    prev_doc.get("ER", {}).get("portfolio", {})
                    if name == "ER"
                    else prev_doc.get("states", {}).get(name, {}).get("portfolio", {})
                ),
                "source": "psp_historical"
            }

        record = {
            "_id": name,
            "state": name,
            "date": peak_date,
            "max_demand": float(peak.get(demand_key) or 0),
            "max_demand_time": str(peak.get(time_key) or "19:00"),
            "portfolio": portfolio,
            "previous_highest": previous,
            "updated_at": now_str,
            "source": "psp_historical_mongo_only"
        }
        db.db["psp_highest_records"].update_one(
            {"_id": name},
            {"$set": record},
            upsert=True
        )
        record.pop("_id", None)
        records.append(record)

    return records


def rebuild_highest_records_from_psp_daily():
    return rebuild_highest_records_from_psp_historical()

def check_and_update_highest_portfolio(target_date: date):
    rebuild_highest_records_from_psp_historical()

def initialize_highest_records():
    db = MongoService()
    hist_docs = list(db.db["psp_historical"].find().sort("date", 1))
    print(f"Initializing highest records. Found {len(hist_docs)} historical documents.")
    
    highest = {}
    STATES = [
        {"name": "BIHAR", "historical_prefix": "BIHAR"},
        {"name": "JHARKHAND", "historical_prefix": "JHARKHAND"},
        {"name": "DVC", "historical_prefix": "DVC"},
        {"name": "ODISHA", "historical_prefix": "ORISSA"},
        {"name": "WEST BENGAL", "historical_prefix": "WEST_BENGAL"},
        {"name": "SIKKIM", "historical_prefix": "SIKKIM"},
        {"name": "ER", "historical_prefix": "ER"}
    ]
    
    for doc in hist_docs:
        date_str = doc.get("date")
        if not date_str:
            continue
            
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            continue
            
        for state in STATES:
            state_label = state["name"]
            prefix = state["historical_prefix"]
            
            val = doc.get(f"{prefix}_MAX_DEMAND")
            if val is not None:
                val = float(val)
                time_str = doc.get(f"{prefix}_MAX_DEMAND_TIME") or "19:00"
                
                curr_highest = highest.get(state_label)
                if curr_highest is None or val > curr_highest["max_demand"]:
                    prev = None
                    if curr_highest:
                        prev = {
                            "date": curr_highest["date"],
                            "max_demand": curr_highest["max_demand"],
                            "max_demand_time": curr_highest["max_demand_time"],
                            "portfolio": curr_highest["portfolio"]
                        }
                        
                    breakdown_doc = db.db["psp_portfolio_breakdown"].find_one({"_id": date_str})
                    if not breakdown_doc:
                        try:
                            breakdown_doc = calculate_and_save_portfolio(target_date)
                        except Exception as e:
                            print(f"Error calculating portfolio for {date_str} during init: {e}")
                            continue
                            
                    if state_label == "ER":
                        portfolio = breakdown_doc.get("ER", {}).get("portfolio", {})
                    else:
                        portfolio = breakdown_doc.get("states", {}).get(state_label, {}).get("portfolio", {})
                        
                    highest[state_label] = {
                        "_id": state_label,
                        "state": state_label,
                        "date": date_str,
                        "max_demand": val,
                        "max_demand_time": time_str,
                        "portfolio": portfolio,
                        "previous_highest": prev,
                        "updated_at": datetime.utcnow().isoformat()
                    }
                    
    for state_label, record in highest.items():
        db.db["psp_highest_records"].update_one(
            {"_id": state_label},
            {"$set": record},
            upsert=True
        )
    print("Finished initializing highest records.")

@router.get("/portfolio-demand-breakdown")
async def get_portfolio_demand_breakdown(date_str: str = None):
    db = MongoService()
    
    if date_str:
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            return {"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}
    else:
        latest_doc = db.psp_collection.find_one(
            {"pspstatedemandrequirement": {"$exists": True, "$ne": []}},
            sort=[("date", -1)]
        )
        if latest_doc:
            target_date = date.fromisoformat(latest_doc["date"])
        else:
            target_date = date.today() - timedelta(days=1)
            
    date_str_db = target_date.strftime("%Y-%m-%d")
    
    try:
        portfolio_doc = get_valid_portfolio_doc(db, target_date)
    except Exception as e:
        return {"success": False, "message": f"Failed to calculate portfolio: {str(e)}"}
            
    breakdown_data = []
    for state_name, val in portfolio_doc.get("states", {}).items():
        breakdown_data.append({
            "state": state_name,
            "max_demand": val.get("max_demand"),
            "time": val.get("time"),
            "loadshed": val.get("loadshed"),
            "internal_gen": {
                "thermal": val.get("portfolio", {}).get("thermal", 0.0),
                "hydro": val.get("portfolio", {}).get("hydro", 0.0),
                "solar": val.get("portfolio", {}).get("solar", 0.0),
                "biogas": val.get("portfolio", {}).get("biogas", 0.0),
                "nuclear": val.get("portfolio", {}).get("nuclear", 0.0)
            },
            "portfolio": {
                "isgs": val.get("portfolio", {}).get("isgs", 0.0),
                "gna": val.get("portfolio", {}).get("gna", 0.0),
                "tgna": val.get("portfolio", {}).get("tgna", 0.0),
                "idam": val.get("portfolio", {}).get("idam", 0.0),
                "rtm": val.get("portfolio", {}).get("rtm", 0.0),
                "own_gen": val.get("portfolio", {}).get("own_gen", 0.0),
                "own_gen_source": val.get("portfolio", {}).get("own_gen_source", ""),
                "wbes_acronym": val.get("portfolio", {}).get("wbes_acronym", ""),
                "wbes_source": val.get("portfolio", {}).get("wbes_source", ""),
                "dsm": val.get("portfolio", {}).get("dsm", 0.0)
            }
        })
        
    er_val = portfolio_doc.get("ER", {})
    er_breakdown = {
        "state": "ER",
        "max_demand": er_val.get("max_demand"),
        "time": er_val.get("time"),
        "loadshed": er_val.get("loadshed"),
        "internal_gen": {
            "thermal": er_val.get("portfolio", {}).get("thermal", 0.0),
            "hydro": er_val.get("portfolio", {}).get("hydro", 0.0),
            "solar": er_val.get("portfolio", {}).get("solar", 0.0),
            "biogas": er_val.get("portfolio", {}).get("biogas", 0.0),
            "nuclear": er_val.get("portfolio", {}).get("nuclear", 0.0)
        },
        "portfolio": {
            "isgs": er_val.get("portfolio", {}).get("isgs", 0.0),
            "gna": er_val.get("portfolio", {}).get("gna", 0.0),
            "tgna": er_val.get("portfolio", {}).get("tgna", 0.0),
            "idam": er_val.get("portfolio", {}).get("idam", 0.0),
            "rtm": er_val.get("portfolio", {}).get("rtm", 0.0),
            "own_gen": er_val.get("portfolio", {}).get("own_gen", 0.0),
            "own_gen_source": er_val.get("portfolio", {}).get("own_gen_source", ""),
            "wbes_acronym": er_val.get("portfolio", {}).get("wbes_acronym", ""),
            "wbes_source": er_val.get("portfolio", {}).get("wbes_source", ""),
            "dsm": er_val.get("portfolio", {}).get("dsm", 0.0)
        }
    }
    
    return {
        "success": True,
        "date": date_str_db,
        "data": breakdown_data,
        "er_data": er_breakdown
    }

@router.get("/highest-records")
async def get_highest_records():
    try:
        db = MongoService()
        records = list(db.db["psp_highest_records"].find({}, {"_id": 0}).sort("state", 1))
        missing_portfolio = any(
            not record.get("portfolio")
            or record.get("portfolio", {}).get("own_gen") in (None, "")
            or (
                float(record.get("max_demand") or 0) > 0
                and float(record.get("portfolio", {}).get("own_gen") or 0) <= 0
            )
            for record in records
        )
        if len(records) < len(PSP_PEAK_STATES) or missing_portfolio:
            records = rebuild_highest_records_from_psp_historical()
    except Exception as e:
        return {"success": False, "message": f"Failed to load highest records: {str(e)}"}
    return {
        "success": True,
        "data": records
    }

@router.get("/initialize-highest")
async def trigger_initialize_highest():
    try:
        initialize_highest_records()
        return {"success": True, "message": "Highest records initialized successfully."}
    except Exception as e:
        return {"success": False, "message": str(e)}

@router.get("/power-position")
async def get_power_position(date_str: str = None):
    db = MongoService()
    
    # 1. Determine target date
    if not date_str:
        latest_doc = db.psp_collection.find_one(
            {"pspstatedemandrequirement": {"$exists": True, "$ne": []}},
            sort=[("date", -1)]
        )
        if latest_doc:
            target_date_str = latest_doc["date"]
        else:
            target_date_str = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    else:
        target_date_str = date_str
        
    # Spreadsheet baseline peaks strictly before 2026-06-02.
    BASELINE_PEAKS = PSP_PEAK_BASELINES
    STATES = PSP_PEAK_STATES

    # helper to get daily data
    def get_daily_values_inline(doc_hist, doc_daily):
        if doc_hist:
            res = {}
            for s in STATES:
                name = s["name"]
                prefix = s["hist_prefix"]
                demand = doc_hist.get(f"{prefix}_MAX_DEMAND")
                demand_time = doc_hist.get(f"{prefix}_MAX_DEMAND_TIME") or "19:00"
                energy_key = f"{prefix}_ENERGY" if prefix != "ORISSA" else "ODISHA_ENERGY"
                energy = doc_hist.get(energy_key)
                
                res[name] = {
                    "demand": float(demand) if demand is not None else 0.0,
                    "demand_time": str(demand_time),
                    "energy": float(energy) if energy is not None else 0.0
                }
            return res
        elif doc_daily:
            demand_list = doc_daily.get("pspstatedemandrequirement", [])
            demand_map = {item.get("STATE_NAME"): item for item in demand_list if item.get("STATE_NAME")}
            
            load_list = doc_daily.get("pspstateloaddetailsER", [])
            load_map = {item.get("STATE_NAME"): item for item in load_list if item.get("STATE_NAME")}
            
            er_energy = sum(float(item.get("CONSUMPTION", 0) or 0) for item in load_list)
            
            res = {}
            for s in STATES:
                name = s["name"]
                daily_dem_name = s["daily_demand_name"]
                daily_en_name = s["daily_energy_name"]
                
                # Demand
                match_dem = demand_map.get(daily_dem_name)
                demand = float(match_dem.get("MAX_DEMAND") or 0.0) if match_dem else 0.0
                demand_time = match_dem.get("MAX_DEMAND_TIME") or "19:00" if match_dem else "19:00"
                
                # Energy
                if name == "ER":
                    energy = er_energy
                else:
                    match_en = load_map.get(daily_en_name)
                    energy = float(match_en.get("CONSUMPTION") or 0.0) if match_en else 0.0
                    
                res[name] = {
                    "demand": demand,
                    "demand_time": demand_time,
                    "energy": energy
                }
            return res
        return None

    # Get target date's data
    t_hist = db.db["psp_historical"].find_one({"date": target_date_str})
    t_daily = db.psp_collection.find_one({"date": target_date_str})
    
    target_data = get_daily_values_inline(t_hist, t_daily)
    if not target_data:
        return {"success": True, "has_data": False, "date": target_date_str, "rows": []}

    # Chronologically compute running peaks strictly before target_date_str
    import json
    running_peaks = json.loads(json.dumps(BASELINE_PEAKS)) # deep copy
    
    start_dt = date.fromisoformat("2026-06-02")
    try:
        target_dt = date.fromisoformat(target_date_str)
    except Exception:
        target_dt = start_dt

    if target_dt > start_dt:
        # Bulk query for intermediate dates
        hist_docs = list(db.db["psp_historical"].find({
            "date": {"$gte": "2026-06-02", "$lt": target_date_str}
        }))
        hist_map = {d["date"]: d for d in hist_docs if "date" in d}
        
        daily_docs = list(db.psp_collection.find({
            "date": {"$gte": "2026-06-02", "$lt": target_date_str}
        }))
        daily_map = {d["date"]: d for d in daily_docs if "date" in d}
        
        curr_dt = start_dt
        while curr_dt < target_dt:
            curr_str = curr_dt.strftime("%Y-%m-%d")
            curr_vals = get_daily_values_inline(hist_map.get(curr_str), daily_map.get(curr_str))
            if curr_vals:
                for s in STATES:
                    name = s["name"]
                    # Demand
                    if curr_vals[name]["demand"] > running_peaks[name]["max_demand"]:
                        running_peaks[name]["max_demand"] = curr_vals[name]["demand"]
                        running_peaks[name]["max_demand_date"] = curr_str
                        running_peaks[name]["max_demand_time"] = curr_vals[name]["demand_time"]
                    # Energy
                    if curr_vals[name]["energy"] > running_peaks[name]["max_energy"]:
                        running_peaks[name]["max_energy"] = curr_vals[name]["energy"]
                        running_peaks[name]["max_energy_date"] = curr_str
            curr_dt += timedelta(days=1)

    # Assemble rows
    rows = []
    for s in STATES:
        name = s["name"]
        t_val = target_data[name]
        p_val = running_peaks[name]
        
        demand_break = t_val["demand"] > p_val["max_demand"]
        energy_break = t_val["energy"] > p_val["max_energy"]
        
        demand_diff_mw = t_val["demand"] - p_val["max_demand"]
        demand_diff_pct = (demand_diff_mw / p_val["max_demand"] * 100) if p_val["max_demand"] > 0 else 0.0
        
        energy_diff_mu = t_val["energy"] - p_val["max_energy"]
        energy_diff_pct = (energy_diff_mu / p_val["max_energy"] * 100) if p_val["max_energy"] > 0 else 0.0
        
        rows.append({
            "constituent": name,
            "daily_date": target_date_str,
            "daily_demand": t_val["demand"],
            "daily_demand_time": t_val["demand_time"],
            "daily_energy": round(t_val["energy"], 2),
            "daily_energy_date": target_date_str,
            
            "all_time_demand": p_val["max_demand"],
            "all_time_demand_date": p_val["max_demand_date"],
            "all_time_demand_time": p_val["max_demand_time"],
            
            "all_time_energy": round(p_val["max_energy"], 2),
            "all_time_energy_date": p_val["max_energy_date"],
            
            "demand_break": demand_break,
            "demand_diff_mw": round(demand_diff_mw, 1),
            "demand_diff_pct": round(demand_diff_pct, 2),
            
            "energy_break": energy_break,
            "energy_diff_mu": round(energy_diff_mu, 2),
            "energy_diff_pct": round(energy_diff_pct, 2)
        })
        
    return {
        "success": True,
        "has_data": True,
        "date": target_date_str,
        "rows": rows
    }

@router.get("/voltage-profile")
async def get_voltage_profile(date_str: str = None):
    """Get voltage profile for 400kV and 765kV substations for map visualization."""
    db = MongoService()

    if date_str:
        doc = db.psp_collection.find_one({"date": date_str})
    else:
        doc = db.psp_collection.find_one(
            {
                "pspVoltageProfile_400kv": {"$exists": True, "$not": {"$size": 0}}
            },
            sort=[("date", -1)]
        )

    if not doc:
        return {"success": True, "has_data": False, "date": None, "kv400": [], "kv765": []}

    def classify_station_400(station):
        min_v = float(station.get("min_voltage") or 0.0)
        max_v = float(station.get("max_voltage") or 0.0)
        if min_v == 0.0 and max_v == 0.0:
            return "offline"
        elif min_v < 380:
            return "low"
        elif max_v >= 420:
            return "high"
        else:
            return "normal"

    def classify_station_765(station):
        min_v = float(station.get("min_voltage") or 0.0)
        max_v = float(station.get("max_voltage") or 0.0)
        if min_v == 0.0 and max_v == 0.0:
            return "offline"
        elif min_v < 728:
            return "low"
        elif max_v >= 800:
            return "high"
        else:
            return "normal"

    kv400_raw = doc.get("pspVoltageProfile_400kv", [])
    kv765_raw = doc.get("pspVoltageProfile_765kv", [])

    er_substation_coords = {
        "ALIPURDUAR": {"lat": 26.49, "lng": 89.53, "state": "West Bengal"},
        "ANGUL": {"lat": 20.84, "lng": 85.10, "state": "Odisha"},
        "BANKA": {"lat": 24.88, "lng": 86.92, "state": "Bihar"},
        "BARIPADA": {"lat": 21.94, "lng": 86.72, "state": "Odisha"},
        "BEHRAMPORE": {"lat": 24.10, "lng": 88.25, "state": "West Bengal"},
        "BERHAMPORE": {"lat": 24.10, "lng": 88.25, "state": "West Bengal"},
        "BIHAR SHARIF": {"lat": 25.20, "lng": 85.52, "state": "Bihar"},
        "BODHGAYA": {"lat": 24.70, "lng": 84.99, "state": "Bihar"},
        "BOKARO": {"lat": 23.67, "lng": 86.15, "state": "Jharkhand"},
        "CHAIBASA": {"lat": 22.56, "lng": 85.80, "state": "Jharkhand"},
        "DARLIPALI": {"lat": 21.50, "lng": 83.40, "state": "Odisha"},
        "DURGAPUR": {"lat": 23.50, "lng": 87.30, "state": "West Bengal"},
        "FARAKKA": {"lat": 24.82, "lng": 87.92, "state": "West Bengal"},
        "GAYA": {"lat": 24.80, "lng": 85.00, "state": "Bihar"},
        "GODDA": {"lat": 24.83, "lng": 87.21, "state": "Jharkhand"},
        "GOKARNA": {"lat": 24.04, "lng": 88.18, "state": "West Bengal"},
        "JAMSHEDPUR": {"lat": 22.80, "lng": 86.20, "state": "Jharkhand"},
        "JEYPORE": {"lat": 18.90, "lng": 82.60, "state": "Odisha"},
        "JHARSUGUDA": {"lat": 21.90, "lng": 84.10, "state": "Odisha"},
        "KAHALGAON": {"lat": 25.25, "lng": 87.26, "state": "Bihar"},
        "KHARAGPUR": {"lat": 22.35, "lng": 87.32, "state": "West Bengal"},
        "KODERMA": {"lat": 24.50, "lng": 85.60, "state": "Jharkhand"},
        "MAITHON": {"lat": 23.78, "lng": 86.82, "state": "Jharkhand"},
        "MALDA": {"lat": 25.02, "lng": 88.14, "state": "West Bengal"},
        "MIDNAPORE": {"lat": 22.42, "lng": 87.32, "state": "West Bengal"},
        "MUZAFFARPUR": {"lat": 26.12, "lng": 85.40, "state": "Bihar"},
        "NEW PURNEA": {"lat": 25.78, "lng": 87.48, "state": "Bihar"},
        "PATNA": {"lat": 25.60, "lng": 85.10, "state": "Bihar"},
        "PURNIA": {"lat": 25.78, "lng": 87.48, "state": "Bihar"},
        "PURNEA": {"lat": 25.78, "lng": 87.48, "state": "Bihar"},
        "PURULIA": {"lat": 23.33, "lng": 86.36, "state": "West Bengal"},
        "RANCHI": {"lat": 23.40, "lng": 85.30, "state": "Jharkhand"},
        "RANGPO": {"lat": 27.18, "lng": 88.53, "state": "Sikkim"},
        "ROURKELA": {"lat": 22.22, "lng": 84.90, "state": "Odisha"},
        "SASARAM": {"lat": 24.95, "lng": 84.02, "state": "Bihar"},
        "STERLITE": {"lat": 21.90, "lng": 84.05, "state": "Odisha"},
        "TALCHER": {"lat": 20.95, "lng": 85.22, "state": "Odisha"},
        "TEESTA": {"lat": 26.75, "lng": 88.58, "state": "West Bengal"},
    }

    def clean_station_name(value: str):
        return (
            str(value or "")
            .upper()
            .replace("_", " ")
            .replace("-", " ")
            .replace("400KV", "")
            .replace("765KV", "")
            .replace("400 KV", "")
            .replace("765 KV", "")
            .strip()
        )

    def find_master_location(station_name: str):
        clean_name = clean_station_name(station_name)

        for collection_name in ("Substation_Master", "substation_master", "station_mapping"):
            if collection_name not in db.db.list_collection_names():
                continue

            doc_match = db.db[collection_name].find_one(
                {
                    "$or": [
                        {"name": {"$regex": clean_name, "$options": "i"}},
                        {"station_name": {"$regex": clean_name, "$options": "i"}},
                        {"STATION_NAME": {"$regex": clean_name, "$options": "i"}},
                    ]
                },
                {"_id": 0}
            )
            if not doc_match:
                continue

            lat = doc_match.get("lat") or doc_match.get("latitude")
            lng = doc_match.get("lng") or doc_match.get("lon") or doc_match.get("longitude")
            if lat is not None and lng is not None:
                return {
                    "lat": float(lat),
                    "lng": float(lng),
                    "state": doc_match.get("state") or doc_match.get("state_name") or "",
                    "location_source": collection_name,
                }

        for key, meta in er_substation_coords.items():
            if key in clean_name:
                return {**meta, "location_source": "ER_STATIC_MASTER"}

        return {"lat": None, "lng": None, "state": "", "location_source": "UNRESOLVED"}

    kv400 = []
    for s in kv400_raw:
        location = find_master_location(s.get("STATION_NAME", ""))
        kv400.append({
            "station_key": s.get("station_key"),
            "name": s.get("STATION_NAME", ""),
            **location,
            "min_voltage": float(s.get("min_voltage") or 0.0),
            "max_voltage": float(s.get("max_voltage") or 0.0),
            "min_time": s.get("MIN_TIME", ""),
            "max_time": s.get("MAX_TIME", ""),
            "volt1": s.get("volt1", ""),
            "volt1_value": float(s.get("volt1_value") or 0.0),
            "volt2": s.get("volt2", ""),
            "volt2_value": float(s.get("volt2_value") or 0.0),
            "volt3": s.get("volt3", ""),
            "volt3_value": float(s.get("volt3_value") or 0.0),
            "volt4": s.get("volt4", ""),
            "volt4_value": float(s.get("volt4_value") or 0.0) if s.get("volt4_value") is not None else 0.0,
            "deviation_index": float(s.get("Voltage_Dev_Index") or 0.0),
            "status": classify_station_400(s),
            "level": "400kV"
        })

    kv765 = []
    for s in kv765_raw:
        location = find_master_location(s.get("STATION_NAME", ""))
        kv765.append({
            "station_key": s.get("station_key"),
            "name": s.get("STATION_NAME", ""),
            **location,
            "min_voltage": float(s.get("min_voltage") or 0.0),
            "max_voltage": float(s.get("max_voltage") or 0.0),
            "min_time": s.get("MIN_TIME", ""),
            "max_time": s.get("MAX_TIME", ""),
            "volt1": s.get("volt1", ""),
            "volt1_value": float(s.get("volt1_value") or 0.0),
            "volt2": s.get("volt2", ""),
            "volt2_value": float(s.get("volt2_value") or 0.0),
            "volt3": s.get("volt3", ""),
            "volt3_value": float(s.get("volt3_value") or 0.0),
            "volt4": s.get("volt4") if s.get("volt4") != "null" else None,
            "volt4_value": float(s.get("volt4_value") or 0.0) if s.get("volt4_value") is not None else None,
            "deviation_index": float(s.get("Voltage_Dev_Index") or 0.0),
            "status": classify_station_765(s),
            "level": "765kV"
        })

    return {
        "success": True,
        "has_data": True,
        "date": doc.get("date"),
        "kv400": kv400,
        "kv765": kv765
    }

@router.get("/voltage-profile-trend")
async def get_voltage_profile_trend(start_date: str = None, end_date: str = None, stations: str = None):
    """Historical max/min voltage trend for selected substations."""
    try:
        db = MongoService()
        if not end_date:
            latest_doc = db.psp_collection.find_one(
                {"$or": [
                    {"pspVoltageProfile_400kv": {"$exists": True, "$ne": []}},
                    {"pspVoltageProfile_765kv": {"$exists": True, "$ne": []}},
                ]},
                {"date": 1},
                sort=[("date", -1)]
            )
            end_date = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (date.fromisoformat(end_date) - timedelta(days=14)).strftime("%Y-%m-%d")
        date.fromisoformat(start_date)
        date.fromisoformat(end_date)

        selected = {
            item.strip()
            for item in str(stations or "").split(",")
            if item.strip()
        }

        def station_key(raw: dict, level: str):
            key = str(raw.get("station_key") or "").strip()
            if key:
                return f"{level}:{key}"
            name = str(raw.get("STATION_NAME") or "").strip().upper()
            return f"{level}:{name}"

        def station_name(raw: dict):
            return str(raw.get("STATION_NAME") or raw.get("station_key") or "").strip()

        docs = list(db.psp_collection.find(
            {"date": {"$gte": start_date, "$lte": end_date}},
            {"_id": 0, "date": 1, "pspVoltageProfile_400kv": 1, "pspVoltageProfile_765kv": 1}
        ).sort("date", 1))

        available = {}
        rows = []
        for doc in docs:
            for level, source_key in (("400kV", "pspVoltageProfile_400kv"), ("765kV", "pspVoltageProfile_765kv")):
                for raw in doc.get(source_key, []) or []:
                    key = station_key(raw, level)
                    name = station_name(raw)
                    available.setdefault(key, {
                        "key": key,
                        "name": name,
                        "level": level,
                    })
                    if selected and key not in selected:
                        continue
                    rows.append({
                        "date": doc.get("date"),
                        "station_key": key,
                        "name": name,
                        "level": level,
                        "min_voltage": float(raw.get("min_voltage") or 0.0),
                        "max_voltage": float(raw.get("max_voltage") or 0.0),
                        "min_time": raw.get("MIN_TIME", ""),
                        "max_time": raw.get("MAX_TIME", ""),
                    })

        if not selected:
            default_keys = list(available.keys())[:5]
            rows = [row for row in rows if row["station_key"] in default_keys]
            selected = set(default_keys)

        return {
            "success": True,
            "start_date": start_date,
            "end_date": end_date,
            "stations": sorted(available.values(), key=lambda item: (item["level"], item["name"])),
            "selected": list(selected),
            "rows": rows,
        }
    except ValueError:
        return {"success": False, "message": "Invalid date. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}


@router.get("/power-exchange")
async def get_power_exchange(date_str: str = None):
    """Get ER inter-regional and transnational power exchange from PSP data."""
    db = MongoService()

    projection = {
        "_id": 0,
        "date": 1,
        "pspTransnationalExchangeState": 1,
        "pspinterregionalscheduleactual": 1,
    }

    if date_str:
        doc = db.psp_collection.find_one({"date": date_str}, projection)
    else:
        doc = db.psp_collection.find_one(
            {
                "$or": [
                    {"pspTransnationalExchangeState": {"$exists": True, "$ne": []}},
                    {"pspinterregionalscheduleactual": {"$exists": True, "$ne": []}},
                ]
            },
            projection,
            sort=[("date", -1)]
        )

    if not doc:
        return {
            "success": True,
            "has_data": False,
            "date": None,
            "interregional": [],
            "transnational": [],
            "totals": {
                "interregional_schedule": 0.0,
                "interregional_actual": 0.0,
                "transnational_schedule": 0.0,
                "transnational_actual": 0.0,
            },
        }

    def to_float(value):
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    region_labels = {
        "NR": "Northern Region",
        "WR": "Western Region",
        "SR": "Southern Region",
        "NER": "North Eastern Region",
    }

    interregional = []
    for row in doc.get("pspinterregionalscheduleactual", []) or []:
        to_region = str(row.get("TO_REGION_NAME") or "").upper()
        from_region = str(row.get("FROM_REGION_NAME") or "").upper()
        other_region = to_region if from_region == "ER" else from_region
        if other_region == "ER" or not other_region:
            continue

        interregional.append({
            "code": other_region,
            "name": region_labels.get(other_region, other_region),
            "from_region": from_region,
            "to_region": to_region,
            "schedule": round(to_float(row.get("TOTAL_IR_SCHEDULE")), 3),
            "actual": round(to_float(row.get("TOTAL_IR_ACTUAL")), 3),
            "net_ui": round(to_float(row.get("NET_IR_UI")), 3),
            "isgs_schedule": round(to_float(row.get("ISGS_SCHEDULE")), 3),
            "bilt_schedule": round(to_float(row.get("BILT_SCHEDULE")), 3),
            "px_schedule": round(to_float(row.get("PX_SCHEDULE")), 3),
        })

    country_labels = {
        "BANGLADESH": "Bangladesh",
        "BHUTAN": "Bhutan",
        "NEPAL": "Nepal",
    }

    transnational = []
    for row in doc.get("pspTransnationalExchangeState", []) or []:
        country = str(row.get("STATE_NAME") or "").upper()
        transnational.append({
            "code": country,
            "name": country_labels.get(country, country.title()),
            "schedule": round(to_float(row.get("Scheduled_EX")), 3),
            "actual": round(to_float(row.get("ACTUAL_EX")), 3),
            "day_peak": round(to_float(row.get("Day_Peak")), 3),
            "day_min": round(to_float(row.get("DAY_MIN")), 3),
            "day_avg": round(to_float(row.get("DAY_AVG")), 3),
        })

    return {
        "success": True,
        "has_data": True,
        "date": doc.get("date"),
        "interregional": interregional,
        "transnational": transnational,
        "totals": {
            "interregional_schedule": round(sum(item["schedule"] for item in interregional), 3),
            "interregional_actual": round(sum(item["actual"] for item in interregional), 3),
            "transnational_schedule": round(sum(item["schedule"] for item in transnational), 3),
            "transnational_actual": round(sum(item["actual"] for item in transnational), 3),
        },
    }

@router.get("/power-exchange-range")
async def get_power_exchange_range(start_date: str = None, end_date: str = None):
    """Summed ER exchange schedule/actual over a date range."""
    try:
        db = MongoService()
        if not end_date:
            latest_doc = db.psp_collection.find_one(
                {
                    "$or": [
                        {"pspTransnationalExchangeState": {"$exists": True, "$ne": []}},
                        {"pspinterregionalscheduleactual": {"$exists": True, "$ne": []}},
                    ]
                },
                {"date": 1},
                sort=[("date", -1)]
            )
            end_date = latest_doc["date"] if latest_doc else date.today().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (date.fromisoformat(end_date) - timedelta(days=6)).strftime("%Y-%m-%d")
        start_dt = date.fromisoformat(start_date)
        end_dt = date.fromisoformat(end_date)
        if start_dt > end_dt:
            return {"success": False, "message": "Start date must be less than or equal to end date."}

        projection = {
            "_id": 0,
            "date": 1,
            "pspTransnationalExchangeState": 1,
            "pspinterregionalscheduleactual": 1,
        }
        docs = list(db.psp_collection.find(
            {"date": {"$gte": start_date, "$lte": end_date}},
            projection
        ).sort("date", 1))

        def as_float(value):
            try:
                return float(value or 0.0)
            except (TypeError, ValueError):
                return 0.0

        region_labels = {
            "NR": "Northern Region",
            "WR": "Western Region",
            "SR": "Southern Region",
            "NER": "North Eastern Region",
        }
        country_labels = {
            "BANGLADESH": "Bangladesh",
            "BHUTAN": "Bhutan",
            "NEPAL": "Nepal",
        }
        aggregate = {}
        daily_rows = []

        def ensure_row(group, code, name):
            key = f"{group}:{code}"
            aggregate.setdefault(key, {
                "group": group,
                "code": code,
                "name": name,
                "schedule": 0.0,
                "actual": 0.0,
                "net_ui": 0.0,
                "days": 0,
            })
            return aggregate[key]

        for doc in docs:
            date_value = doc.get("date")
            for row in doc.get("pspinterregionalscheduleactual", []) or []:
                to_region = str(row.get("TO_REGION_NAME") or "").upper()
                from_region = str(row.get("FROM_REGION_NAME") or "").upper()
                other_region = to_region if from_region == "ER" else from_region
                if other_region == "ER" or not other_region:
                    continue
                schedule = as_float(row.get("TOTAL_IR_SCHEDULE"))
                actual = as_float(row.get("TOTAL_IR_ACTUAL"))
                net_ui = as_float(row.get("NET_IR_UI"))
                target = ensure_row("Inter Regional", other_region, region_labels.get(other_region, other_region))
                target["schedule"] += schedule
                target["actual"] += actual
                target["net_ui"] += net_ui
                target["days"] += 1
                daily_rows.append({
                    "date": date_value,
                    "group": "Inter Regional",
                    "code": other_region,
                    "name": region_labels.get(other_region, other_region),
                    "schedule": round(schedule, 3),
                    "actual": round(actual, 3),
                    "net_ui": round(net_ui, 3),
                })

            for row in doc.get("pspTransnationalExchangeState", []) or []:
                country = str(row.get("STATE_NAME") or "").upper()
                if not country:
                    continue
                schedule = as_float(row.get("Scheduled_EX"))
                actual = as_float(row.get("ACTUAL_EX"))
                target = ensure_row("Transnational", country, country_labels.get(country, country.title()))
                target["schedule"] += schedule
                target["actual"] += actual
                target["days"] += 1
                daily_rows.append({
                    "date": date_value,
                    "group": "Transnational",
                    "code": country,
                    "name": country_labels.get(country, country.title()),
                    "schedule": round(schedule, 3),
                    "actual": round(actual, 3),
                    "net_ui": None,
                })

        rows = []
        for item in aggregate.values():
            rows.append({
                **item,
                "schedule": round(item["schedule"], 3),
                "actual": round(item["actual"], 3),
                "net_ui": round(item["net_ui"], 3),
                "difference": round(item["actual"] - item["schedule"], 3),
            })
        rows.sort(key=lambda item: (item["group"], item["code"]))

        return {
            "success": True,
            "start_date": start_date,
            "end_date": end_date,
            "date_count": len(docs),
            "rows": rows,
            "daily_rows": daily_rows,
            "totals": {
                "interregional_schedule": round(sum(item["schedule"] for item in rows if item["group"] == "Inter Regional"), 3),
                "interregional_actual": round(sum(item["actual"] for item in rows if item["group"] == "Inter Regional"), 3),
                "transnational_schedule": round(sum(item["schedule"] for item in rows if item["group"] == "Transnational"), 3),
                "transnational_actual": round(sum(item["actual"] for item in rows if item["group"] == "Transnational"), 3),
            },
        }
    except ValueError:
        return {"success": False, "message": "Invalid date. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.post("/mis/diurnal-curve")
async def get_mis_diurnal_curve(req: DiurnalCurveRequest):
    """Generate processed MIS diurnal demand curves from curve files."""
    try:
        states = [normalize_check_key(item) for item in (req.states or []) if str(item or "").strip()]
        if not states:
            states = ["ER"]
        req.states = states
        if not req.date_ranges:
            today = date.today()
            req.date_ranges = [
                DiurnalCurveRange(
                    label="Last 7 Days",
                    start_date=(today - timedelta(days=7)).strftime("%Y-%m-%d"),
                    end_date=(today - timedelta(days=1)).strftime("%Y-%m-%d"),
                    curve_type="daily",
                )
            ]
        if int(req.block_minutes or 0) <= 0:
            req.block_minutes = 15

        all_dates = []
        for range_req in req.date_ranges:
            all_dates.extend(expand_diurnal_dates(range_req))
        all_dates = sorted(set(all_dates))

        config = get_psp_config_with_curve_defaults()
        db = MongoService()
        config["curve_state_mapping"] = get_curve_state_mapping(db)

        series_by_date_state = {}
        source_meta = []
        for date_str in all_dates:
            series, meta = read_curve_file_series(date_str, states, config)
            source_meta.append({"date": date_str, **(meta or {})})
            if meta.get("available"):
                series_by_date_state[date_str] = series

        result = aggregate_diurnal_series(series_by_date_state, req)
        return {
            "success": True,
            "states": states,
            "block_minutes": req.block_minutes,
            "date_count": len(all_dates),
            "source_meta": source_meta,
            **result,
        }
    except ValueError as exc:
        return {"success": False, "message": str(exc)}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

MIS_VOLTAGE_API_BASE_URL = "http://10.3.230.62:5010"
MIS_REACTOR_SWITCHING_URL = "https://crms.erldc.in/Codebook/TrElementOutageHistoryData"
MIS_ELEMENT_NAMES_URL = "https://crms.erldc.in/Codebook/getElementNamesbyType"
MIS_UNIT_MASTER_URL = "https://mdp.erldc.in/outageapi/API/GeneratingStation/GetAllGeneratingUnits/0"
MIS_PLANNED_OUTAGE_COLLECTION = "MIS_Planned_Outage_Master"
MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE = "GENERATING_UNIT"
MIS_PLANNED_OUTAGE_FALLBACK_ELEMENT_TYPES = [
    "14",
    "TRANSMISSION LINE",
    "AC_TRANSMISSION_LINE_CIRCUIT",
]

def normalize_mis_voltage_datetime(value: str):
    text = str(value or "").strip().replace("T", " ")
    if len(text) == 16:
        return text
    if len(text) >= 19:
        return text[:16]
    return text

def parse_mis_voltage_datetime(value: str):
    text = str(value or "").strip()
    for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None

def make_mis_voltage_time_axis(start_date: str, end_date: str, interval_minutes: int, count: int):
    start_text = normalize_mis_voltage_datetime(start_date)
    end_text = normalize_mis_voltage_datetime(end_date)
    start_dt = parse_mis_voltage_datetime(start_text)
    end_dt = parse_mis_voltage_datetime(end_text)
    if not start_dt:
        return [str(i + 1) for i in range(count)]
    step = max(1, int(interval_minutes or 5))
    values = []
    current = start_dt
    while len(values) < count and (not end_dt or current <= end_dt):
        values.append(current.strftime("%d-%m-%Y %H:%M"))
        current += timedelta(minutes=step)
    while len(values) < count:
        values.append((start_dt + timedelta(minutes=step * len(values))).strftime("%d-%m-%Y %H:%M"))
    return values

def normalize_voltage_extreme(value):
    try:
        voltage = value[0][0] if isinstance(value, list) and value and isinstance(value[0], list) else None
        time_value = value[1][0] if isinstance(value, list) and len(value) > 1 and isinstance(value[1], list) else ""
        return {"value": round(float(voltage), 3), "time": str(time_value or "")}
    except Exception:
        return {"value": None, "time": ""}

def extract_reactor_substation(element_name: str):
    text = " ".join(str(element_name or "").strip().split())
    parts = text.upper().rsplit(" AT ", 1)
    if len(parts) != 2:
        return ""
    return text[-len(parts[1]):].strip()

def normalize_reactor_datetime(value: str):
    dt = parse_mis_voltage_datetime(value)
    return dt.strftime("%Y-%m-%d %H:%M") if dt else str(value or "").strip()

def fetch_reactor_switching_raw_rows(start_date: str, end_date: str):
    params = {
        "start_date": normalize_mis_voltage_datetime(start_date),
        "end_date": normalize_mis_voltage_datetime(end_date),
    }
    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    response = session.get(MIS_REACTOR_SWITCHING_URL, params=params, timeout=180)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, dict):
        data = data.get("data") or data.get("rows") or []
    return data if isinstance(data, list) else [], response.url

def fetch_mis_element_names(element_type: str):
    params = {"elementType": str(element_type or "").strip(), "_": int(datetime.now().timestamp() * 1000)}
    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    response = session.get(
        MIS_ELEMENT_NAMES_URL,
        params=params,
        headers={
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=120,
    )
    response.raise_for_status()
    text = response.text or ""
    try:
        data = response.json()
    except ValueError:
        matches = re.findall(r'"entityElementName"\s*:\s*"([^"]+)"', text)
        data = [{"entityElementName": name} for name in matches]
    if isinstance(data, dict):
        data = data.get("data") or data.get("rows") or []
    rows = []
    seen = set()
    for item in data if isinstance(data, list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("entityElementName") or item.get("ELEMENTNAME") or item.get("name") or "").strip()
        if not name:
            continue
        key = normalize_check_key(name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"id": item.get("id"), "name": name})
    rows.sort(key=lambda item: item["name"])
    return rows, response.url

def fetch_local_mis_element_names(element_type: str):
    db = MongoService().db
    element_type_text = str(element_type or "").strip().upper()
    projections = {
        "_id": 0,
        "Unit_Name": 1,
        "utility_type": 1,
        "state_name": 1,
        "entity_name": 1,
        "element_type": 1,
        "EntityId": 1,
        "plant_id": 1,
    }
    query = {"Unit_Name": {"$exists": True, "$nin": [None, ""]}}
    if element_type_text and ("TRANSMISSION" in element_type_text or element_type_text in {"14", "AC_TRANSMISSION_LINE_CIRCUIT"}):
        query["$or"] = [
            {"utility_type": {"$regex": "TRANSMISSION|LINE", "$options": "i"}},
            {"entity_name": {"$regex": "TRANSMISSION|LINE", "$options": "i"}},
            {"element_type": {"$regex": "TRANSMISSION|LINE", "$options": "i"}},
            {"EntityId": {"$in": ["14", "TRANSMISSION LINE", "AC_TRANSMISSION_LINE_CIRCUIT"]}},
        ]
    rows = []
    seen = set()
    try:
        for doc in db.unit_collection.find(query, projections):
            name = str(doc.get("Unit_Name") or "").strip()
            if not name:
                continue
            key = normalize_check_key(name)
            if key in seen:
                continue
            seen.add(key)
            rows.append({"id": doc.get("plant_id") or doc.get("EntityId"), "name": name})
    except Exception:
        rows = []
    if not rows and element_type_text and ("TRANSMISSION" in element_type_text or element_type_text in {"14", "AC_TRANSMISSION_LINE_CIRCUIT"}):
        try:
            for doc in db.unit_collection.find({"Unit_Name": {"$exists": True, "$nin": [None, ""]}}, projections):
                name = str(doc.get("Unit_Name") or "").strip()
                if not name:
                    continue
                key = normalize_check_key(name)
                if key in seen:
                    continue
                seen.add(key)
                rows.append({"id": doc.get("plant_id") or doc.get("EntityId"), "name": name})
        except Exception:
            rows = []
    rows.sort(key=lambda item: item["name"])
    return rows, "mongo:unit_collection"

def fetch_mdp_unit_master_names():
    session = requests.Session()
    session.mount("https://", LegacySSLAdapter())
    response = session.get(MIS_UNIT_MASTER_URL, timeout=120, verify=False)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, dict):
        data = data.get("data") or data.get("rows") or []
    rows = []
    seen = set()
    for item in data if isinstance(data, list) else []:
        if not isinstance(item, dict):
            continue
        if item.get("ACTIVE") not in (None, 1, "1", True):
            continue
        name = str(item.get("Unit_Name") or item.get("unit_name") or item.get("name") or "").strip()
        if not name:
            continue
        key = normalize_check_key(name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"id": item.get("Id") or item.get("id") or item.get("Unit_Id") or item.get("Unit_Number"), "name": name})
    rows.sort(key=lambda item: item["name"])
    return rows, response.url

def fetch_mis_planned_outage_unit_names(element_type: str):
    element_type_text = str(element_type or MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE).strip()
    if normalize_check_key(element_type_text) in {
        normalize_check_key("GENERATING_UNIT"),
        normalize_check_key("GENERATING UNIT"),
    }:
        rows, source_url = fetch_mdp_unit_master_names()
        return rows, source_url, MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE, "crms_generating_unit_master"

    candidates = [element_type_text]
    normalized = normalize_check_key(element_type_text)
    if normalized in {normalize_check_key("AC_TRANSMISSION_LINE_CIRCUIT"), normalize_check_key("TRANSMISSION LINE"), "14"}:
        for fallback_type in MIS_PLANNED_OUTAGE_FALLBACK_ELEMENT_TYPES:
            if fallback_type not in candidates:
                candidates.append(fallback_type)

    for candidate in candidates:
        try:
            rows, source_url = fetch_mis_element_names(candidate)
        except Exception:
            rows, source_url = [], ""
        if rows and "accounts/login" not in str(source_url or "").lower():
            return rows, source_url, candidate, "crms"

    local_rows, source_url = fetch_local_mis_element_names(element_type_text)
    if local_rows:
        return local_rows, source_url, element_type_text, "mongo"

    mdp_rows, source_url = fetch_mdp_unit_master_names()
    return mdp_rows, source_url, element_type_text, "mdp"

def get_mis_planned_outage_collection():
    db = MongoService().db
    if MIS_PLANNED_OUTAGE_COLLECTION not in db.list_collection_names():
        db.create_collection(MIS_PLANNED_OUTAGE_COLLECTION)
    return db[MIS_PLANNED_OUTAGE_COLLECTION]

def normalize_planned_outage_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return ""

def serialize_planned_outage_entry(doc):
    history = doc.get("history") or []
    legacy_date = doc.get("planned_outage_date") or ""
    from_date = doc.get("planned_outage_from_date") or legacy_date
    to_date = doc.get("planned_outage_to_date") or from_date
    return {
        "id": str(doc["_id"]),
        "element_type": doc.get("element_type") or "",
        "unit_name": doc.get("unit_name") or "",
        "planned_outage_date": legacy_date or from_date,
        "planned_outage_from_date": from_date,
        "planned_outage_to_date": to_date,
        "reason": doc.get("reason") or "",
        "remarks": doc.get("remarks") or "",
        "source": doc.get("source") or "CRMS",
        "history": [
            {
                "unit_name": item.get("unit_name") or "",
                "planned_outage_date": item.get("planned_outage_date") or item.get("planned_outage_from_date") or "",
                "planned_outage_from_date": item.get("planned_outage_from_date") or item.get("planned_outage_date") or "",
                "planned_outage_to_date": item.get("planned_outage_to_date") or item.get("planned_outage_from_date") or item.get("planned_outage_date") or "",
                "reason": item.get("reason") or "",
                "remarks": item.get("remarks") or "",
                "changed_at": item.get("changed_at").isoformat() if isinstance(item.get("changed_at"), datetime) else item.get("changed_at") or "",
            }
            for item in history
            if isinstance(item, dict)
        ],
        "created_at": doc.get("created_at").isoformat() if isinstance(doc.get("created_at"), datetime) else doc.get("created_at") or "",
        "updated_at": doc.get("updated_at").isoformat() if isinstance(doc.get("updated_at"), datetime) else doc.get("updated_at") or "",
    }

@router.get("/mis/planned-outage/unit-names")
async def get_mis_planned_outage_unit_names(element_type: str = MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE):
    try:
        rows, source_url, resolved_element_type, source_kind = fetch_mis_planned_outage_unit_names(
            element_type or MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE
        )
        return {
            "success": True,
            "element_type": resolved_element_type,
            "source_kind": source_kind,
            "source_url": source_url,
            "units": rows,
            "unit_count": len(rows),
        }
    except Exception as exc:
        traceback.print_exc()
        return {
            "success": False,
            "message": str(exc),
            "units": [],
            "unit_count": 0,
        }

@router.get("/mis/planned-outage/entries")
async def get_mis_planned_outage_entries(element_type: str = ""):
    try:
        collection = get_mis_planned_outage_collection()
        query = {}
        if element_type:
            query["element_type"] = element_type
        docs = list(collection.find(query).sort([("planned_outage_from_date", 1), ("planned_outage_date", 1), ("updated_at", -1)]))
        return {
            "success": True,
            "rows": [serialize_planned_outage_entry(doc) for doc in docs],
            "count": len(docs),
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc), "rows": []}

@router.post("/mis/planned-outage/entries")
async def create_mis_planned_outage_entry(req: MisPlannedOutageRequest):
    try:
        element_type = str(req.element_type or MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE).strip()
        unit_name = str(req.unit_name or "").strip()
        reason = str(req.reason or "").strip()
        remarks = str(req.remarks or "").strip()
        planned_outage_from_date = normalize_planned_outage_date(
            req.planned_outage_from_date or req.planned_outage_date
        )
        planned_outage_to_date = normalize_planned_outage_date(
            req.planned_outage_to_date or req.planned_outage_from_date or req.planned_outage_date
        )
        if not unit_name:
            return {"success": False, "message": "Unit name is required."}
        if not reason:
            return {"success": False, "message": "Outage reason is required."}
        if not planned_outage_from_date or not planned_outage_to_date:
            return {"success": False, "message": "Valid planned outage From and To dates are required."}
        selected_from_date = date.fromisoformat(planned_outage_from_date)
        selected_to_date = date.fromisoformat(planned_outage_to_date)
        if selected_from_date < date.today():
            return {"success": False, "message": "Planned outage From date must be today or a future date."}
        if selected_to_date < selected_from_date:
            return {"success": False, "message": "Planned outage To date cannot be before the From date."}

        units, _, _, source_kind = fetch_mis_planned_outage_unit_names(element_type)
        normalized_unit = ""
        needle = normalize_check_key(unit_name)
        for item in units:
            if normalize_check_key(item.get("name")) == needle:
                normalized_unit = item.get("name") or unit_name
                break
        if not normalized_unit:
            return {"success": False, "message": "Unit name was not found in CRMS master data."}

        now = datetime.utcnow()
        doc = {
            "element_type": element_type,
            "unit_name": normalized_unit,
            "planned_outage_date": planned_outage_from_date,
            "planned_outage_from_date": planned_outage_from_date,
            "planned_outage_to_date": planned_outage_to_date,
            "reason": reason,
            "remarks": remarks,
            "source": source_kind,
            "history": [],
            "created_at": now,
            "updated_at": now,
        }
        collection = get_mis_planned_outage_collection()
        result = collection.insert_one(doc)
        return {
            "success": True,
            "message": "Planned outage entry saved.",
            "id": str(result.inserted_id),
            "entry": serialize_planned_outage_entry({**doc, "_id": result.inserted_id}),
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.put("/mis/planned-outage/entries/{entry_id}")
async def update_mis_planned_outage_entry(entry_id: str, req: MisPlannedOutageRequest):
    try:
        if not ObjectId.is_valid(entry_id):
            return {"success": False, "message": "Invalid planned outage entry."}
        collection = get_mis_planned_outage_collection()
        existing = collection.find_one({"_id": ObjectId(entry_id)})
        if not existing:
            return {"success": False, "message": "Planned outage entry not found."}

        element_type = str(req.element_type or existing.get("element_type") or MIS_PLANNED_OUTAGE_DEFAULT_ELEMENT_TYPE).strip()
        unit_name = str(req.unit_name or "").strip()
        reason = str(req.reason or "").strip()
        remarks = str(req.remarks or "").strip()
        planned_outage_from_date = normalize_planned_outage_date(
            req.planned_outage_from_date or req.planned_outage_date
        )
        planned_outage_to_date = normalize_planned_outage_date(
            req.planned_outage_to_date or req.planned_outage_from_date or req.planned_outage_date
        )
        if not unit_name:
            return {"success": False, "message": "Unit name is required."}
        if not reason:
            return {"success": False, "message": "Outage reason is required."}
        if not planned_outage_from_date or not planned_outage_to_date:
            return {"success": False, "message": "Valid planned outage From and To dates are required."}
        selected_from_date = date.fromisoformat(planned_outage_from_date)
        selected_to_date = date.fromisoformat(planned_outage_to_date)
        if selected_from_date < date.today():
            return {"success": False, "message": "Planned outage From date must be today or a future date."}
        if selected_to_date < selected_from_date:
            return {"success": False, "message": "Planned outage To date cannot be before the From date."}

        units, _, _, source_kind = fetch_mis_planned_outage_unit_names(element_type)
        normalized_unit = ""
        needle = normalize_check_key(unit_name)
        for item in units:
            if normalize_check_key(item.get("name")) == needle:
                normalized_unit = item.get("name") or unit_name
                break
        if not normalized_unit:
            return {"success": False, "message": "Unit name was not found in CRMS master data."}

        now = datetime.utcnow()
        history_entry = {
            "unit_name": existing.get("unit_name") or "",
            "planned_outage_date": existing.get("planned_outage_date") or "",
            "planned_outage_from_date": existing.get("planned_outage_from_date") or existing.get("planned_outage_date") or "",
            "planned_outage_to_date": existing.get("planned_outage_to_date") or existing.get("planned_outage_from_date") or existing.get("planned_outage_date") or "",
            "reason": existing.get("reason") or "",
            "remarks": existing.get("remarks") or "",
            "changed_at": now,
        }
        collection.update_one(
            {"_id": ObjectId(entry_id)},
            {
                "$set": {
                    "element_type": element_type,
                    "unit_name": normalized_unit,
                    "planned_outage_date": planned_outage_from_date,
                    "planned_outage_from_date": planned_outage_from_date,
                    "planned_outage_to_date": planned_outage_to_date,
                    "reason": reason,
                    "remarks": remarks,
                    "source": source_kind or existing.get("source") or "CRMS",
                    "updated_at": now,
                },
                "$push": {"history": history_entry},
            },
        )
        updated = collection.find_one({"_id": ObjectId(entry_id)})
        return {
            "success": True,
            "message": "Planned outage entry updated.",
            "entry": serialize_planned_outage_entry(updated),
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

OUTAGE_REASON_STOPWORDS = {
    "THE", "AND", "FOR", "FROM", "WITH", "DUE", "TO", "OF", "IN", "AT", "ON", "BY", "IS", "WAS",
    "ARE", "BE", "AS", "END", "LINE", "KV", "PH", "PHASE", "KM", "KA", "FAULT", "TRIPPED",
    "OUTAGE", "SHUTDOWN", "WORK", "REASON", "DETAILS", "ZONE", "RECEIVED", "RELAY", "FD", "FC",
}

OUTAGE_CATEGORY_RULES = [
    ("Voltage Regulation", ("VOLTAGE REGULATION", "H/T ON VOLTAGE", "LOW VOLTAGE", "OVER VOLTAGE", "OVERVOLTAGE")),
    ("Physical Regulation", ("PHYSICAL REGULATION", "H/T ON PHYSICAL")),
    ("Protection / Tripping", ("PROTECTION", "TRIP", "TRIPPED", "AUTO TRIP", "A/R SUCCESS", "FAULT", "RELAY", "DISTANCE PROTECTION", "BUS BAR")),
    ("Emergency / Equipment Attention", ("EMERGENCY", "HOTSPOT", "HOT SPOT", "SF6", "OIL LEAK", "FIRE", "FLASHOVER", "DECAP", "CRACK", "BURST", "ALARM")),
    ("Planned / Maintenance", ("OCC SHUTDOWN", "NON-OCC", "PLANNED", "MAINTENANCE", "AMP", "TESTING", "TREE", "PRUNING", "INSULATOR", "OPGW", "METER", "FIRMWARE")),
    ("System Requirement", ("SYSTEM REQUIREMENT", "LOADING", "OVERLOADING", "POWER ORDER", "CONTROL FLOW", "NORMALLY KEPT OPEN")),
]

def classify_outage_reason(reason: str, outage_type: str, entity_name: str):
    combined = normalize_check_key(f"{outage_type} {reason} {entity_name}")
    if "GENERATOR" in combined or entity_name and normalize_check_key(entity_name) == "GENERATING UNIT":
        if "PLANNED" in combined:
            return "Generator Planned"
        return "Generator Forced"
    for category, needles in OUTAGE_CATEGORY_RULES:
        if any(needle in combined for needle in needles):
            return category
    return "Other"

def outage_duration_bucket(hours):
    if hours is None:
        return "Unknown"
    if hours < 1:
        return "< 1 hr"
    if hours < 6:
        return "1-6 hr"
    if hours < 24:
        return "6-24 hr"
    if hours < 72:
        return "1-3 days"
    return "> 3 days"

def format_outage_duration(hours):
    if hours is None:
        return "-"
    if hours < 1:
        return f"{round(hours * 60)} min"
    if hours < 24:
        return f"{hours:.1f} hr"
    days = int(hours // 24)
    rem = hours - (days * 24)
    return f"{days}d {rem:.1f}h"

OLD_LOGBOOK_ANALYSIS_START = date(2019, 10, 1)
OLD_LOGBOOK_ANALYSIS_END = date(2025, 3, 31)
OLD_LOGBOOK_ELEMENT_TYPE_MAP = {
    "14": "AC_TRANSMISSION_LINE_CIRCUIT",
    "TRANSMISSION LINE": "AC_TRANSMISSION_LINE_CIRCUIT",
    "9": "TRANSFORMER",
    "TRANSFORMER": "TRANSFORMER",
    "4": "BUS_REACTOR",
    "BUS REACTOR": "BUS_REACTOR",
    "5": "LINE_REACTOR",
    "LINE REACTOR": "LINE_REACTOR",
    "16": "BUS",
    "BUS": "BUS",
    "25": "BAY",
    "BAY": "BAY",
    "8": "GENERATING_UNIT",
    "GENERATING UNIT": "GENERATING_UNIT",
    "26": "AUTO_RECLOSER",
    "AUTO RECLOSER": "AUTO_RECLOSER",
    "15": "HVDC_POLE",
    "HVDC POLE": "HVDC_POLE",
    "STATCOM": "STATCOM",
}

def old_logbook_analysis_window(start_dt, end_dt):
    if not start_dt or not end_dt:
        return None
    start_day = max(start_dt.date(), OLD_LOGBOOK_ANALYSIS_START)
    end_day = min(end_dt.date(), OLD_LOGBOOK_ANALYSIS_END)
    if start_day > end_day:
        return None
    return start_day, end_day

def old_logbook_datetime_text(doc: dict, date_key: str, time_key: str):
    parsed_date = parse_logbook_date(doc.get(date_key))
    if not parsed_date:
        return ""
    time_text = clean_text(doc.get(time_key)) or "00:00"
    time_match = re.search(r"\d{1,2}:\d{2}(?::\d{2})?", time_text)
    normalized_time = time_match.group(0) if time_match else "00:00"
    if len(normalized_time) == 5:
        normalized_time = f"{normalized_time}:00"
    return f"{parsed_date.isoformat()} {normalized_time[:5]}"

def first_old_logbook_value(doc: dict, *field_names: str):
    for field_name in field_names:
        value = clean_text(doc.get(field_name))
        if value:
            return value
    return ""

def normalize_old_logbook_element_type(value: str):
    text = clean_text(value)
    if not text:
        return ""
    key = normalize_check_key(text)
    return OLD_LOGBOOK_ELEMENT_TYPE_MAP.get(key, OLD_LOGBOOK_ELEMENT_TYPE_MAP.get(text, text))

def old_logbook_doc_to_outage_raw(doc: dict, kind: str):
    config = COLLECTION_CONFIG[kind]
    raw = to_jsonable(doc)
    outage_dt = old_logbook_datetime_text(raw, config["outage_date"], config["outage_time"])
    if not outage_dt:
        return None
    reason = combine_fields(raw, config.get("reason_fields", []))
    if kind == "shutdown" and clean_text(raw.get("Reason")):
        reason = clean_text(raw.get("Reason"))
    element_type = clean_text(raw.get(config.get("element_type", "Type")))
    if not element_type and kind != "shutdown":
        element_type = clean_text(raw.get("EntityId"))
    element_type = normalize_old_logbook_element_type(element_type)
    requesting_entity = first_old_logbook_value(
        raw,
        "RequestingEntity",
        "Requesting Entity",
        "CodeIssuedBy",
        "Code Issued By",
        "IssuedBy",
        "Issued By",
        "CreatedBy",
        "UserName",
    )
    owner = first_old_logbook_value(raw, "owner", "Owner", "OwnerName", "Utility", "EntityName", "Entity")
    return {
        "ELEMENTNAME": clean_text(raw.get(config["element"])),
        "ELEMENT_NAME": clean_text(raw.get(config["element"])),
        "ENTITY_NAME": element_type,
        "OUTAGE_TYPE": config["label"],
        "REASON": reason,
        "RequestingEntity": requesting_entity,
        "owner": owner,
        "OUTAGE_DATE_TIME": outage_dt,
        "REVIVED_DATE_TIME": old_logbook_datetime_text(raw, config["revival_date"], config["revival_time"]),
        "SOURCE": "Old Logbook",
        "SOURCE_COLLECTION": config["collection"],
        "OLD_LOGBOOK_ID": clean_text(raw.get("_id")),
    }

def fetch_old_logbook_outage_raw_rows(start_dt, end_dt):
    window = old_logbook_analysis_window(start_dt, end_dt)
    if not window:
        return [], {"enabled": False, "count": 0}
    start_day, end_day = window
    mongo = MongoService()
    db = mongo.client[OLD_LOGBOOK_DB]
    rows = []
    scanned = defaultdict(int)
    for kind, config in COLLECTION_CONFIG.items():
        collection = db[config["collection"]]
        date_key = config["outage_date"]
        docs = list(collection.find({}))
        scanned[config["collection"]] = len(docs)
        for doc in docs:
            parsed_day = parse_logbook_date(doc.get(date_key))
            if not parsed_day or parsed_day < start_day or parsed_day > end_day:
                continue
            row = old_logbook_doc_to_outage_raw(doc, kind)
            if row:
                rows.append(row)
    return rows, {
        "enabled": True,
        "count": len(rows),
        "start_date": start_day.isoformat(),
        "end_date": end_day.isoformat(),
        "collections": dict(scanned),
    }

def summarize_outage_group(rows, key):
    buckets = defaultdict(lambda: {"count": 0, "open_count": 0, "total_duration_hours": 0.0, "durations": []})
    for row in rows:
        name = row.get(key) or "Unspecified"
        item = buckets[name]
        item["count"] += 1
        if row.get("status") == "Open":
            item["open_count"] += 1
        duration = row.get("duration_hours")
        if isinstance(duration, (int, float)):
            item["total_duration_hours"] += duration
            item["durations"].append(duration)
    summary = []
    for name, item in buckets.items():
        durations = item.pop("durations")
        count = item["count"]
        item["name"] = name
        item["avg_duration_hours"] = round(sum(durations) / len(durations), 2) if durations else None
        item["max_duration_hours"] = round(max(durations), 2) if durations else None
        item["total_duration_hours"] = round(item["total_duration_hours"], 2)
        item["share_percent"] = 0
        summary.append(item)
    total = max(1, len(rows))
    for item in summary:
        item["share_percent"] = round((item["count"] / total) * 100, 1)
    return sorted(summary, key=lambda item: (-item["count"], item["name"]))[:20]

def analyze_outage_rows(raw_rows, req: MisOutageAnalysisRequest):
    start_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.start_date))
    end_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.end_date))
    if start_dt and len(normalize_mis_voltage_datetime(req.start_date)) <= 10:
        start_dt = datetime.combine(start_dt.date(), datetime.min.time())
    if end_dt:
        end_dt = datetime.combine(end_dt.date(), datetime.max.time()) if len(normalize_mis_voltage_datetime(req.end_date)) <= 10 else end_dt
    now_dt = datetime.now()
    analysis_dt = now_dt

    entity_filter = {normalize_check_key(item) for item in req.entity_names if str(item or "").strip()}
    element_filter = {normalize_check_key(item) for item in req.element_names if str(item or "").strip()}
    type_filter = {normalize_check_key(item) for item in req.outage_types if str(item or "").strip()}
    excluded_type_filter = {normalize_check_key(item) for item in req.excluded_outage_types if str(item or "").strip()}
    requesting_filter = {normalize_check_key(item) for item in req.requesting_entities if str(item or "").strip()}
    owner_filter = {normalize_check_key(item) for item in req.owners if str(item or "").strip()}
    reason_query = normalize_check_key(req.reason_query)

    rows = []
    entity_options = set()
    outage_type_options = set()
    requesting_options = set()
    owner_options = set()
    keyword_counter = Counter()

    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        element_name = str(item.get("ELEMENTNAME") or item.get("ELEMENT_NAME") or "").strip()
        entity_name = str(item.get("ENTITY_NAME") or "").strip()
        outage_type = str(item.get("OUTAGE_TYPE") or "").strip()
        reason = str(item.get("REASON") or "").strip()
        requesting_entity = str(item.get("RequestingEntity") or "").strip()
        owner = str(item.get("owner") or "").strip()
        source = str(item.get("SOURCE") or item.get("source") or "CRMS").strip() or "CRMS"
        source_collection = str(item.get("SOURCE_COLLECTION") or item.get("source_collection") or "").strip()
        outage_dt = parse_mis_voltage_datetime(item.get("OUTAGE_DATE_TIME"))
        revived_dt = parse_mis_voltage_datetime(item.get("REVIVED_DATE_TIME"))

        if excluded_type_filter and normalize_check_key(outage_type) in excluded_type_filter:
            continue

        if entity_name:
            entity_options.add(entity_name)
        if outage_type:
            outage_type_options.add(outage_type)
        if requesting_entity:
            requesting_options.add(requesting_entity)
        for owner_part in [part.strip() for part in owner.split(",") if part.strip()]:
            owner_options.add(owner_part)

        if not outage_dt:
            continue
        if start_dt and revived_dt and revived_dt < start_dt:
            continue
        if end_dt and outage_dt > end_dt:
            continue

        if element_filter and normalize_check_key(element_name) not in element_filter:
            continue
        if entity_filter and normalize_check_key(entity_name) not in entity_filter:
            continue
        if type_filter and normalize_check_key(outage_type) not in type_filter:
            continue
        if requesting_filter and normalize_check_key(requesting_entity) not in requesting_filter:
            continue
        if owner_filter:
            owner_keys = {normalize_check_key(part) for part in owner.split(",") if part.strip()}
            if not owner_keys.intersection(owner_filter):
                continue
        search_blob = normalize_check_key(f"{element_name} {entity_name} {outage_type} {reason} {requesting_entity} {owner}")
        if reason_query and reason_query not in search_blob:
            continue

        duration_end = revived_dt or analysis_dt
        duration_hours = max(0.0, round((duration_end - outage_dt).total_seconds() / 3600, 2))
        category = classify_outage_reason(reason, outage_type, entity_name)
        bucket = outage_duration_bucket(duration_hours)
        for word in re.findall(r"[A-Z0-9]{3,}", normalize_check_key(reason)):
            if word not in OUTAGE_REASON_STOPWORDS and not word.isdigit():
                keyword_counter[word] += 1

        rows.append({
            "element_name": element_name,
            "entity_name": entity_name or "Unspecified",
            "outage_type": outage_type or "Unspecified",
            "reason": reason,
            "reason_category": category,
            "requesting_entity": requesting_entity or "Unspecified",
            "owner": owner or "Unspecified",
            "outage_time": outage_dt.strftime("%Y-%m-%d %H:%M"),
            "revived_time": revived_dt.strftime("%Y-%m-%d %H:%M") if revived_dt else "",
            "status": "Closed" if revived_dt else "Open",
            "duration_hours": duration_hours,
            "duration_label": format_outage_duration(duration_hours),
            "duration_bucket": bucket,
            "source": source,
            "source_collection": source_collection,
        })

    rows.sort(key=lambda row: row["outage_time"], reverse=True)
    durations = [row["duration_hours"] for row in rows if isinstance(row.get("duration_hours"), (int, float))]
    category_summary = summarize_outage_group(rows, "reason_category")
    outage_type_summary = summarize_outage_group(rows, "outage_type")
    keyword_summary = [{"keyword": word, "count": count} for word, count in keyword_counter.most_common(20)]
    return {
        "success": True,
        "start_date": normalize_mis_voltage_datetime(req.start_date),
        "end_date": normalize_mis_voltage_datetime(req.end_date),
        "total_rows": len(raw_rows),
        "filtered_rows": len(rows),
        "excluded_outage_types": req.excluded_outage_types,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "events": len(rows),
            "open": sum(1 for row in rows if row["status"] == "Open"),
            "closed": sum(1 for row in rows if row["status"] == "Closed"),
            "avg_duration_hours": round(sum(durations) / len(durations), 2) if durations else None,
            "max_duration_hours": round(max(durations), 2) if durations else None,
            "total_duration_hours": round(sum(durations), 2) if durations else 0,
        },
        "category_summary": category_summary,
        "entity_summary": summarize_outage_group(rows, "entity_name"),
        "outage_type_summary": outage_type_summary,
        "duration_summary": summarize_outage_group(rows, "duration_bucket"),
        "requesting_entity_summary": summarize_outage_group(rows, "requesting_entity"),
        "owner_summary": summarize_outage_group(rows, "owner"),
        "keyword_summary": keyword_summary,
        "filter_options": {
            "entity_names": sorted(entity_options),
            "outage_types": sorted(outage_type_options),
            "requesting_entities": sorted(requesting_options),
            "owners": sorted(owner_options),
        },
        "rows": rows,
    }

@router.get("/mis/voltage-names")
async def get_mis_voltage_names(start_date: str, end_date: str):
    try:
        params = {
            "startDate": normalize_mis_voltage_datetime(start_date),
            "endDate": normalize_mis_voltage_datetime(end_date),
        }
        response = requests.get(f"{MIS_VOLTAGE_API_BASE_URL}/VoltageNames", params=params, timeout=60)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, dict):
            data = data.get("data") or data.get("stations") or data.get("stationNames") or []
        stations = sorted({
            str(item or "").strip()
            for item in (data if isinstance(data, list) else [])
            if str(item or "").strip()
        })
        return {
            "success": True,
            "start_date": params["startDate"],
            "end_date": params["endDate"],
            "count": len(stations),
            "stations": stations,
            "source_url": response.url,
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.get("/mis/element-names")
async def get_mis_element_names(element_type: str = ""):
    try:
        element_type_text = str(element_type or "").strip()
        target_types = [element_type_text] if element_type_text else MIS_OUTAGE_ELEMENT_TYPES
        merged = {}
        source_urls = []
        for target_type in target_types:
            rows, source_url = fetch_mis_element_names(target_type)
            source_urls.append(source_url)
            for row in rows:
                name = str(row.get("name") or "").strip()
                if name:
                    merged[normalize_check_key(name)] = {**row, "name": name, "element_type": target_type}
        rows = sorted(merged.values(), key=lambda item: item["name"])
        return {
            "success": True,
            "element_type": element_type_text,
            "element_types": target_types,
            "count": len(rows),
            "elements": rows,
            "source_url": source_urls[0] if len(source_urls) == 1 else "",
            "source_urls": source_urls,
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.post("/mis/voltage-profile")
async def get_mis_voltage_profile(req: MisVoltageProfileRequest):
    try:
        stations = [str(item or "").strip() for item in (req.station_names or []) if str(item or "").strip()]
        if not stations:
            return {"success": False, "message": "Select at least one substation."}
        interval = max(1, int(req.time or 5))
        bus_key = req.voltage_bus if req.voltage_bus in {"voltageBus1", "voltageBus2"} else "voltageBus1"
        params = {
            "startDate": normalize_mis_voltage_datetime(req.start_date),
            "endDate": normalize_mis_voltage_datetime(req.end_date),
            "stationName": ",".join(stations),
            "time": interval,
        }
        response = requests.get(f"{MIS_VOLTAGE_API_BASE_URL}/GetVoltageData", params=params, timeout=180)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, dict):
            data = data.get("data") or data.get("rows") or []
        raw_rows = data if isinstance(data, list) else []

        series = []
        max_len = 0
        for item in raw_rows:
            if not isinstance(item, dict):
                continue
            station_name = str(item.get("stationName") or item.get("StationName") or "").strip()
            values = item.get(bus_key) or []
            if not isinstance(values, list):
                values = []
            parsed_values = []
            for value in values:
                try:
                    parsed_values.append(round(float(value), 3))
                except (TypeError, ValueError):
                    parsed_values.append(None)
            if not any((value or 0) > 0 for value in parsed_values):
                continue
            max_len = max(max_len, len(parsed_values))
            avg_key = "avg_v1" if bus_key == "voltageBus1" else "avg_v2"
            min_key = "min_v1" if bus_key == "voltageBus1" else "min_v2"
            max_key = "max_v1" if bus_key == "voltageBus1" else "max_v2"
            series.append({
                "key": f"v{len(series)}",
                "label": station_name or f"Station {len(series) + 1}",
                "station_name": station_name,
                "voltage_bus": bus_key,
                "values": parsed_values,
                "avg": round(float(item.get(avg_key)), 3) if item.get(avg_key) is not None else None,
                "min": normalize_voltage_extreme(item.get(min_key)),
                "max": normalize_voltage_extreme(item.get(max_key)),
            })

        explicit_times = []
        for item in raw_rows:
            if isinstance(item, dict) and isinstance(item.get("Date_Time"), list):
                explicit_times = [str(value or "") for value in item.get("Date_Time")]
                break
        time_axis = explicit_times[:max_len] if explicit_times else make_mis_voltage_time_axis(req.start_date, req.end_date, interval, max_len)
        chart_rows = []
        for index, time_label in enumerate(time_axis):
            row = {"time": time_label}
            for item in series:
                row[item["key"]] = item["values"][index] if index < len(item["values"]) else None
            chart_rows.append(row)

        return {
            "success": True,
            "master_point": req.master_point,
            "start_date": params["startDate"],
            "end_date": params["endDate"],
            "time": interval,
            "voltage_bus": bus_key,
            "station_count": len(stations),
            "series": [{k: v for k, v in item.items() if k != "values"} for item in series],
            "chart_rows": chart_rows,
            "source_url": response.url,
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.post("/mis/reactor-switching")
async def get_mis_reactor_switching(req: MisReactorSwitchingRequest):
    try:
        raw_rows, source_url = fetch_reactor_switching_raw_rows(req.start_date, req.end_date)
        start_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.start_date))
        end_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.end_date))

        def in_requested_window(value):
            if not value:
                return False
            if start_dt and value < start_dt:
                return False
            if end_dt and value > end_dt:
                return False
            return True

        station_filter = {str(item or "").strip().upper() for item in (req.stations or []) if str(item or "").strip()}
        rows = []
        events = []
        for item in raw_rows:
            if not isinstance(item, dict):
                continue
            entity_name = normalize_check_key(item.get("ENTITY_NAME"))
            reason = normalize_check_key(item.get("REASON"))
            if entity_name not in {"BUS REACTOR", "LINE REACTOR"}:
                continue
            if reason != "H/T ON VOLTAGE REGULATION":
                continue
            element_name = item.get("ELEMENTNAME") or item.get("ELEMENT_NAME") or ""
            station = extract_reactor_substation(element_name)
            station_upper = station.upper()
            if station_filter and not any(
                station_upper == target or station_upper in target or target in station_upper
                for target in station_filter
            ):
                continue
            normalized = {
                "element_name": str(element_name or "").strip(),
                "entity_name": str(item.get("ENTITY_NAME") or "").strip(),
                "reason": str(item.get("REASON") or "").strip(),
                "requesting_entity": str(item.get("RequestingEntity") or "").strip(),
                "outage_type": str(item.get("OUTAGE_TYPE") or "").strip(),
                "station": station,
                "open_time": normalize_reactor_datetime(item.get("OUTAGE_DATE_TIME")),
                "close_time": normalize_reactor_datetime(item.get("REVIVED_DATE_TIME")),
                "owner": str(item.get("owner") or "").strip(),
            }
            rows.append(normalized)
            open_dt = parse_mis_voltage_datetime(item.get("OUTAGE_DATE_TIME"))
            close_dt = parse_mis_voltage_datetime(item.get("REVIVED_DATE_TIME"))
            if in_requested_window(open_dt):
                events.append({
                    "event_type": "Open",
                    "timestamp": int(open_dt.timestamp() * 1000),
                    "time": open_dt.strftime("%Y-%m-%d %H:%M"),
                    **normalized,
                })
            if in_requested_window(close_dt):
                events.append({
                    "event_type": "Close",
                    "timestamp": int(close_dt.timestamp() * 1000),
                    "time": close_dt.strftime("%Y-%m-%d %H:%M"),
                    **normalized,
                })
        stations = sorted({row["station"] for row in rows if row.get("station")})
        events.sort(key=lambda item: (item["timestamp"], item["station"], item["event_type"]))
        return {
            "success": True,
            "start_date": normalize_mis_voltage_datetime(req.start_date),
            "end_date": normalize_mis_voltage_datetime(req.end_date),
            "source_url": source_url,
            "stations": stations,
            "rows": rows,
            "events": events,
            "summary": {
                "reactors": len(rows),
                "events": len(events),
                "stations": len(stations),
            },
        }
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.post("/mis/outage-analysis")
async def get_mis_outage_analysis(req: MisOutageAnalysisRequest):
    try:
        start_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.start_date))
        end_dt = parse_mis_voltage_datetime(normalize_mis_voltage_datetime(req.end_date))
        raw_rows = []
        source_url = ""
        source_warnings = []

        if end_dt and end_dt.date() > OLD_LOGBOOK_ANALYSIS_END:
            crms_start_dt = start_dt
            historical_next_day = datetime.combine(OLD_LOGBOOK_ANALYSIS_END + timedelta(days=1), datetime.min.time())
            if not crms_start_dt or crms_start_dt < historical_next_day:
                crms_start_dt = historical_next_day
            crms_start = crms_start_dt.strftime("%Y-%m-%d %H:%M")
            try:
                crms_rows, source_url = fetch_reactor_switching_raw_rows(crms_start, req.end_date)
                for row in crms_rows:
                    if isinstance(row, dict) and not row.get("SOURCE"):
                        row["SOURCE"] = "CRMS"
                raw_rows.extend(crms_rows)
            except Exception as crms_exc:
                source_warnings.append(f"CRMS fetch failed: {str(crms_exc)}")

        old_rows, old_summary = fetch_old_logbook_outage_raw_rows(start_dt, end_dt)
        raw_rows.extend(old_rows)

        result = analyze_outage_rows(raw_rows, req)
        result["source_url"] = source_url
        result["source_summary"] = {
            "crms_rows": sum(1 for row in raw_rows if isinstance(row, dict) and str(row.get("SOURCE") or "").upper() == "CRMS"),
            "old_logbook_rows": len(old_rows),
            "old_logbook": old_summary,
            "warnings": source_warnings,
        }
        return result
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

def build_mis_psp_snapshot_output(range_items: list[dict]):
    db = MongoService()
    states = ["ER", "BIHAR", "DVC", "JHARKHAND", "ODISHA", "SIKKIM", "WEST BENGAL"]

    def empty_bucket():
        return {
            "date_count": 0,
            "max_demand_met": None,
            "max_demand_date": "",
            "max_demand_time": "",
            "total_consumption": 0.0,
            "average_energy_consumption": None,
            "daily_maximum_consumption": None,
            "daily_maximum_consumption_date": "",
        }

    def date_range_strings(range_start, range_end):
        start_dt = date.fromisoformat(range_start)
        end_dt = date.fromisoformat(range_end)
        return [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((end_dt - start_dt).days + 1)]

    def summarize_range(item, index):
        range_start = item.get("start_date")
        range_end = item.get("end_date")
        start_dt = date.fromisoformat(range_start)
        end_dt = date.fromisoformat(range_end)
        if start_dt > end_dt:
            raise ValueError("Start date must be less than or equal to end date.")

        expected_dates = date_range_strings(range_start, range_end)
        label = str(item.get("label") or "").strip() or f"{range_start} to {range_end}"
        docs = list(db.psp_collection.find(
            {"date": {"$gte": range_start, "$lte": range_end}},
            {
                "_id": 0,
                "date": 1,
                "pspstateloaddetailsER": 1,
                "pspstatedemandrequirement": 1,
                "pspstatedemandforecast": 1,
            }
        ).sort("date", 1))
        fetched_dates = [doc.get("date") for doc in docs if doc.get("date")]
        missing_dates = [day for day in expected_dates if day not in set(fetched_dates)]

        summary = {state: empty_bucket() for state in states}
        daily_rows = []

        for doc in docs:
            date_value = doc.get("date")
            values = extract_operational_check_values(doc)
            for state in states:
                item_values = values.get(state, {}) or {}
                demand = item_values.get("max_demand")
                energy = item_values.get("energy")
                energy_value = float(energy or 0.0)
                bucket = summary[state]
                bucket["date_count"] += 1
                bucket["total_consumption"] += energy_value
                if demand is not None and (
                    bucket["max_demand_met"] is None or float(demand) > bucket["max_demand_met"]
                ):
                    bucket["max_demand_met"] = float(demand)
                    bucket["max_demand_date"] = date_value
                    bucket["max_demand_time"] = item_values.get("max_demand_time") or ""
                if bucket["daily_maximum_consumption"] is None or energy_value > bucket["daily_maximum_consumption"]:
                    bucket["daily_maximum_consumption"] = energy_value
                    bucket["daily_maximum_consumption_date"] = date_value
                daily_rows.append({
                    "range_label": label,
                    "date": date_value,
                    "state": state,
                    "maximum_demand_met": round(float(demand), 3) if demand is not None else None,
                    "maximum_demand_time": item_values.get("max_demand_time") or "",
                    "energy_consumption": round(energy_value, 3),
                })

        rows = []
        for state in states:
            bucket = summary[state]
            if bucket["date_count"]:
                bucket["total_consumption"] = round(bucket["total_consumption"], 3)
                bucket["average_energy_consumption"] = round(bucket["total_consumption"] / bucket["date_count"], 3)
                if bucket["daily_maximum_consumption"] is not None:
                    bucket["daily_maximum_consumption"] = round(bucket["daily_maximum_consumption"], 3)
                if bucket["max_demand_met"] is not None:
                    bucket["max_demand_met"] = round(bucket["max_demand_met"], 3)
            rows.append({"range_label": label, "range_index": index, "state": state, **bucket})

        totals = {
            "date_count": len(docs),
            "er_max_demand": next((row["max_demand_met"] for row in rows if row["state"] == "ER"), None),
            "er_total_consumption": next((row["total_consumption"] for row in rows if row["state"] == "ER"), 0.0),
            "er_average_consumption": next((row["average_energy_consumption"] for row in rows if row["state"] == "ER"), None),
            "er_daily_maximum_consumption": next((row["daily_maximum_consumption"] for row in rows if row["state"] == "ER"), None),
        }
        return {
            "label": label,
            "start_date": range_start,
            "end_date": range_end,
            "expected_date_count": len(expected_dates),
            "date_count": len(docs),
            "fetched_dates": fetched_dates,
            "missing_dates": missing_dates,
            "rows": rows,
            "daily_rows": daily_rows,
            "totals": totals,
        }

    ranges = [summarize_range(item, index) for index, item in enumerate(range_items)]
    by_range_state = [
        {row["state"]: row for row in range_data["rows"]}
        for range_data in ranges
    ]
    comparison_rows = []
    for state in states:
        row = {"state": state}
        for range_data, state_map in zip(ranges, by_range_state):
            state_values = state_map.get(state, {})
            prefix = range_data["label"]
            for key in ("max_demand_met", "average_energy_consumption", "total_consumption", "daily_maximum_consumption"):
                row[f"{prefix}|{key}"] = state_values.get(key)
        comparison_rows.append(row)

    return {
        "success": True,
        "source": "Mongo PSP snapshot collection",
        "collections": {
            "maximum_demand_met": "pspstatedemandrequirement.MAX_DEMAND",
            "energy_consumption": "pspstateloaddetailsER.CONSUMPTION",
            "er_energy": "sum of state CONSUMPTION rows",
        },
        "ranges": ranges,
        "comparison_rows": comparison_rows,
    }

@router.post("/mis/psp-snapshot-output")
async def post_mis_psp_snapshot_output(req: SnapshotOutputRequest):
    try:
        range_items = [item.dict() for item in (req.ranges or [])]
        if not range_items:
            return {"success": False, "message": "At least one date range is required."}
        return build_mis_psp_snapshot_output(range_items)
    except ValueError as exc:
        return {"success": False, "message": str(exc)}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}

@router.get("/mis/psp-snapshot-output")
async def get_mis_psp_snapshot_output(
    start_date: str,
    end_date: str,
    compare_start_date: str = None,
    compare_end_date: str = None,
):
    """Range data output from PSP snapshots already saved in Mongo."""
    try:
        start_dt = date.fromisoformat(start_date)
        end_dt = date.fromisoformat(end_date)
        if start_dt > end_dt:
            return {"success": False, "message": "Start date must be less than or equal to end date."}
        if bool(compare_start_date) != bool(compare_end_date):
            return {"success": False, "message": "Provide both compare_start_date and compare_end_date."}
        if compare_start_date and compare_end_date:
            compare_start_dt = date.fromisoformat(compare_start_date)
            compare_end_dt = date.fromisoformat(compare_end_date)
            if compare_start_dt > compare_end_dt:
                return {"success": False, "message": "Compare start date must be less than or equal to compare end date."}

        db = MongoService()
        states = ["ER", "BIHAR", "DVC", "JHARKHAND", "ODISHA", "SIKKIM", "WEST BENGAL"]

        def empty_bucket():
            return {
                "date_count": 0,
                "max_demand_met": None,
                "max_demand_date": "",
                "max_demand_time": "",
                "total_consumption": 0.0,
                "average_energy_consumption": None,
                "daily_maximum_consumption": None,
                "daily_maximum_consumption_date": "",
            }

        def summarize_range(label, range_start, range_end):
            docs = list(db.psp_collection.find(
                {"date": {"$gte": range_start, "$lte": range_end}},
                {
                    "_id": 0,
                    "date": 1,
                    "pspstateloaddetailsER": 1,
                    "pspstatedemandrequirement": 1,
                    "pspstatedemandforecast": 1,
                }
            ).sort("date", 1))

            summary = {state: empty_bucket() for state in states}
            daily_rows = []

            for doc in docs:
                date_value = doc.get("date")
                values = extract_operational_check_values(doc)
                for state in states:
                    item = values.get(state, {}) or {}
                    demand = item.get("max_demand")
                    energy = item.get("energy")
                    energy_value = float(energy or 0.0)
                    bucket = summary[state]
                    bucket["date_count"] += 1
                    bucket["total_consumption"] += energy_value
                    if demand is not None and (
                        bucket["max_demand_met"] is None or float(demand) > bucket["max_demand_met"]
                    ):
                        bucket["max_demand_met"] = float(demand)
                        bucket["max_demand_date"] = date_value
                        bucket["max_demand_time"] = item.get("max_demand_time") or ""
                    if bucket["daily_maximum_consumption"] is None or energy_value > bucket["daily_maximum_consumption"]:
                        bucket["daily_maximum_consumption"] = energy_value
                        bucket["daily_maximum_consumption_date"] = date_value
                    daily_rows.append({
                        "range_label": label,
                        "date": date_value,
                        "state": state,
                        "maximum_demand_met": round(float(demand), 3) if demand is not None else None,
                        "maximum_demand_time": item.get("max_demand_time") or "",
                        "energy_consumption": round(energy_value, 3),
                    })

            rows = []
            for state in states:
                bucket = summary[state]
                if bucket["date_count"]:
                    bucket["total_consumption"] = round(bucket["total_consumption"], 3)
                    bucket["average_energy_consumption"] = round(bucket["total_consumption"] / bucket["date_count"], 3)
                    if bucket["daily_maximum_consumption"] is not None:
                        bucket["daily_maximum_consumption"] = round(bucket["daily_maximum_consumption"], 3)
                    if bucket["max_demand_met"] is not None:
                        bucket["max_demand_met"] = round(bucket["max_demand_met"], 3)
                rows.append({"range_label": label, "state": state, **bucket})

            totals = {
                "date_count": len(docs),
                "er_max_demand": next((row["max_demand_met"] for row in rows if row["state"] == "ER"), None),
                "er_total_consumption": next((row["total_consumption"] for row in rows if row["state"] == "ER"), 0.0),
                "er_average_consumption": next((row["average_energy_consumption"] for row in rows if row["state"] == "ER"), None),
                "er_daily_maximum_consumption": next((row["daily_maximum_consumption"] for row in rows if row["state"] == "ER"), None),
            }
            return {
                "label": label,
                "start_date": range_start,
                "end_date": range_end,
                "date_count": len(docs),
                "rows": rows,
                "daily_rows": daily_rows,
                "totals": totals,
            }

        ranges = [summarize_range("Base Range", start_date, end_date)]
        if compare_start_date and compare_end_date:
            ranges.append(summarize_range("Compare Range", compare_start_date, compare_end_date))

        comparison_rows = []
        if len(ranges) == 2:
            compare_by_state = {row["state"]: row for row in ranges[1]["rows"]}
            for base_row in ranges[0]["rows"]:
                compare_row = compare_by_state.get(base_row["state"], {})
                item = {"state": base_row["state"]}
                for key in ("max_demand_met", "average_energy_consumption", "total_consumption", "daily_maximum_consumption"):
                    base_value = base_row.get(key)
                    compare_value = compare_row.get(key)
                    delta = (
                        round(float(base_value) - float(compare_value), 3)
                        if base_value is not None and compare_value is not None
                        else None
                    )
                    item[f"base_{key}"] = base_value
                    item[f"compare_{key}"] = compare_value
                    item[f"delta_{key}"] = delta
                comparison_rows.append(item)

        return {
            "success": True,
            "source": "Mongo PSP snapshot collection",
            "collections": {
                "maximum_demand_met": "pspstatedemandrequirement.MAX_DEMAND",
                "energy_consumption": "pspstateloaddetailsER.CONSUMPTION",
                "er_energy": "sum of state CONSUMPTION rows",
            },
            "ranges": ranges,
            "comparison_rows": comparison_rows,
        }
    except ValueError:
        return {"success": False, "message": "Invalid date. Use YYYY-MM-DD."}
    except Exception as exc:
        traceback.print_exc()
        return {"success": False, "message": str(exc)}
